"""缺陷管理系统 FastAPI 主应用"""
from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Request, Form, Body
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session, joinedload, noload, selectinload
from sqlalchemy import func, and_, or_
from typing import List, Optional, Any, Dict
from datetime import datetime, timedelta, date
import time
from pathlib import Path
import json
import io
import csv
import requests
import re
import hashlib
import base64
import os
import subprocess
import warnings
from decimal import Decimal
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# 过滤 Pydantic 的受保护命名空间警告
warnings.filterwarnings('ignore', message='.*has conflict with protected namespace.*')

import models
import schemas
from config import engine, get_db, SessionLocal
from permissions import check_permission, require_permission, get_user_permissions, ROLE_NAMES
from auth import hash_password, verify_password, create_access_token, decode_access_token, get_current_user, CurrentUser
from swagger_parser import OpenAPIParser, parse_swagger_file
from data_generator import TestDataGenerator


# 创建数据库表
models.Base.metadata.create_all(bind=engine)

# 初始化默认管理员用户
def init_default_admin():
    """初始化默认管理员用户（如果不存在或密码不正确）"""
    from config import SessionLocal
    from auth import hash_password, verify_password
    
    db = SessionLocal()
    try:
        # 检查 admin 用户是否存在
        admin_user = db.query(models.User).filter(models.User.username == 'admin').first()
        if not admin_user:
            # 创建默认管理员用户
            admin_password_hash = hash_password('admin123')
            admin_user = models.User(
                username='admin',
                email='admin@example.com',
                password=admin_password_hash,
                display_name='系统管理员',
                roles=['admin'],
                status='active'
            )
            db.add(admin_user)
            db.commit()
            print("✅ 已创建默认管理员用户: admin/admin123")
        else:
            # 检查密码是否正确，如果不正确则更新
            if not verify_password('admin123', admin_user.password):
                print("⚠️  admin 用户密码不正确，正在更新...")
                admin_user.password = hash_password('admin123')
                db.commit()
                print("✅ 已更新 admin 用户密码为: admin123")
            else:
                print("ℹ️  管理员用户已存在，密码正确")
    except Exception as e:
        import traceback
        print(f"⚠️  初始化管理员用户时出错: {e}")
        traceback.print_exc()
        db.rollback()
    finally:
        db.close()

# 应用启动时初始化默认管理员
init_default_admin()

app = FastAPI(title="缺陷管理系统", version="1.0.0")

# 配置日志：过滤 401 未授权访问日志
import logging

# 自定义日志过滤器：过滤 401 响应
class Filter401(logging.Filter):
    def filter(self, record: logging.LogRecord) -> bool:
        # 检查日志消息中是否包含 401 Unauthorized
        message = record.getMessage()
        # 匹配格式: "GET /api/xxx HTTP/1.1" 401 Unauthorized
        if ' 401 ' in message or '401 Unauthorized' in message:
            return False
        # 检查是否有 status_code 属性（uvicorn 可能会设置）
        if hasattr(record, 'status_code') and record.status_code == 401:
            return False
        return True

# 配置 uvicorn 的访问日志过滤器
# 注意：需要在 uvicorn 启动后配置，所以这里先设置，启动时会被应用
try:
    access_logger = logging.getLogger("uvicorn.access")
    if not any(isinstance(f, Filter401) for f in access_logger.filters):
        access_logger.addFilter(Filter401())
except Exception:
    pass  # 如果配置失败，不影响应用运行

# ==================== 权限检查辅助函数 ====================

def check_project_member_permission(
    user: models.User, 
    project: models.Project, 
    action: str = "操作"
) -> None:
    """
    检查用户是否是项目成员（admin 可以操作所有项目）
    
    Args:
        user: 用户对象
        project: 项目对象（需要已加载 members 关联）
        action: 操作描述（用于错误消息）
    
    Raises:
        HTTPException: 如果用户不是项目成员且不是 admin
    """
    # admin 可以操作所有项目
    if user.roles and 'admin' in user.roles:
        return
    
    # 加载项目成员（如果还没加载）
    if not hasattr(project, 'members') or project.members is None:
        raise HTTPException(
            status_code=500, 
            detail="项目成员信息未加载，请使用 joinedload"
        )
    
    # 检查是否是项目成员
    if user not in project.members:
        raise HTTPException(
            status_code=403, 
            detail=f"您不是该项目成员，无权{action}"
        )

# 配置CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== 项目管理 ====================

@app.get("/api/projects")
def get_projects(
    page: int = 1,
    page_size: int = 10,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取项目列表（服务端分页）"""
    query = db.query(models.Project).options(joinedload(models.Project.members))
    if keyword:
        query = query.filter(models.Project.name.contains(keyword))
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}

@app.post("/api/projects", response_model=schemas.Project)
def create_project(
    project: schemas.ProjectCreate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建项目"""
    # 检查权限：只有管理员可以创建项目
    require_permission(current_user.role, "projects", "create")
    
    # 检查项目名称是否已存在
    if db.query(models.Project).filter(models.Project.name == project.name).first():
        raise HTTPException(status_code=400, detail="项目名称已存在")
    
    project_data = project.model_dump(exclude={'member_ids'})
    db_project = models.Project(**project_data)
    db.add(db_project)
    db.flush()  # 获取项目ID
    
    # 添加项目成员
    if project.member_ids:
        for user_id in project.member_ids:
            user = db.query(models.User).filter(models.User.id == user_id).first()
            if user:
                db_project.members.append(user)
    
    # 如果创建者是admin，自动添加为成员；否则也添加创建者
    creator = db.query(models.User).filter(models.User.id == current_user.id).first()
    if creator and creator not in db_project.members:
        db_project.members.append(creator)
    
    db.commit()
    db.refresh(db_project)
    # 重新加载以包含成员信息
    db_project = db.query(models.Project).options(joinedload(models.Project.members)).filter(models.Project.id == db_project.id).first()
    return db_project

@app.get("/api/projects/{project_id}", response_model=schemas.Project)
def get_project(
    project_id: int, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取项目详情"""
    project = db.query(models.Project).options(joinedload(models.Project.members)).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    return project

@app.put("/api/projects/{project_id}", response_model=schemas.Project)
def update_project(
    project_id: int, 
    project: schemas.ProjectUpdate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新项目"""
    # 检查权限：只有管理员可以更新项目
    require_permission(current_user.role, "projects", "update")
    
    db_project = db.query(models.Project).options(joinedload(models.Project.members)).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    update_data = project.model_dump(exclude_unset=True, exclude={'member_ids'})
    for key, value in update_data.items():
        setattr(db_project, key, value)
    
    # 更新项目成员
    if 'member_ids' in project.model_dump(exclude_unset=True):
        member_ids = project.member_ids or []
        # 清除现有成员
        db_project.members.clear()
        # 添加新成员
        for user_id in member_ids:
            member_user = db.query(models.User).filter(models.User.id == user_id).first()
            if member_user:
                db_project.members.append(member_user)
    
    db.commit()
    db.refresh(db_project)
    # 重新加载以包含成员信息
    db_project = db.query(models.Project).options(joinedload(models.Project.members)).filter(models.Project.id == project_id).first()
    return db_project

@app.delete("/api/projects/{project_id}")
def delete_project(
    project_id: int, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除项目"""
    # 检查权限：只有管理员可以删除项目
    require_permission(current_user.role, "projects", "delete")
    
    project = db.query(models.Project).options(joinedload(models.Project.members)).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    db.delete(project)
    db.commit()
    return {"message": "项目已删除"}

# ==================== 迭代管理 ====================

@app.get("/api/sprints")
def get_sprints(
    project_id: Optional[int] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取迭代列表（服务端分页）"""
    query = db.query(models.Sprint).options(joinedload(models.Sprint.project))
    if project_id:
        query = query.filter(models.Sprint.project_id == project_id)
    if keyword:
        # 支持按迭代名称或负责人模糊搜索
        query = query.filter(
            or_(
                models.Sprint.name.contains(keyword),
                models.Sprint.owner.contains(keyword)
            )
        )
    total = query.count()
    items = query.order_by(models.Sprint.start_date.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}

@app.post("/api/sprints", response_model=schemas.Sprint)
def create_sprint(
    sprint: schemas.SprintCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建迭代"""
    require_permission(current_user.role, "projects", "create")
    
    # 检查项目是否存在
    project = db.query(models.Project).filter(models.Project.id == sprint.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 验证日期
    if sprint.end_date < sprint.start_date:
        raise HTTPException(status_code=400, detail="截止时间不能早于起始时间")
    
    sprint_data = sprint.model_dump()
    db_sprint = models.Sprint(**sprint_data)
    db.add(db_sprint)
    db.commit()
    db.refresh(db_sprint)
    db_sprint = db.query(models.Sprint).options(joinedload(models.Sprint.project)).filter(models.Sprint.id == db_sprint.id).first()
    return db_sprint

@app.get("/api/sprints/{sprint_id}", response_model=schemas.Sprint)
def get_sprint(
    sprint_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取迭代详情"""
    sprint = db.query(models.Sprint).options(joinedload(models.Sprint.project)).filter(models.Sprint.id == sprint_id).first()
    if not sprint:
        raise HTTPException(status_code=404, detail="迭代不存在")
    return sprint

@app.put("/api/sprints/{sprint_id}", response_model=schemas.Sprint)
def update_sprint(
    sprint_id: int,
    sprint: schemas.SprintUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新迭代"""
    require_permission(current_user.role, "projects", "update")
    
    db_sprint = db.query(models.Sprint).options(joinedload(models.Sprint.project)).filter(models.Sprint.id == sprint_id).first()
    if not db_sprint:
        raise HTTPException(status_code=404, detail="迭代不存在")
    
    update_data = sprint.model_dump(exclude_unset=True)
    
    # 验证日期
    start_date = update_data.get('start_date', db_sprint.start_date)
    end_date = update_data.get('end_date', db_sprint.end_date)
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="截止时间不能早于起始时间")
    
    for key, value in update_data.items():
        setattr(db_sprint, key, value)
    
    db.commit()
    db.refresh(db_sprint)
    db_sprint = db.query(models.Sprint).options(joinedload(models.Sprint.project)).filter(models.Sprint.id == sprint_id).first()
    return db_sprint

@app.delete("/api/sprints/{sprint_id}")
def delete_sprint(
    sprint_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除迭代"""
    require_permission(current_user.role, "projects", "delete")
    
    sprint = db.query(models.Sprint).filter(models.Sprint.id == sprint_id).first()
    if not sprint:
        raise HTTPException(status_code=404, detail="迭代不存在")
    
    db.delete(sprint)
    db.commit()
    return {"message": "迭代已删除"}

# ==================== 需求管理 ====================

@app.get("/api/requirements")
def get_requirements(
    project_id: Optional[int] = None,
    sprint_id: Optional[int] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取需求列表（仅返回顶级需求，每条含children子需求）"""
    from sqlalchemy.orm import subqueryload
    query = db.query(models.Requirement).options(
        joinedload(models.Requirement.project),
        joinedload(models.Requirement.sprint),
        joinedload(models.Requirement.assignee),
        joinedload(models.Requirement.creator),
        subqueryload(models.Requirement.children).options(
            joinedload(models.Requirement.assignee),
            joinedload(models.Requirement.project),
            joinedload(models.Requirement.sprint),
        ),
    ).filter(models.Requirement.parent_id == None)
    if project_id:
        query = query.filter(models.Requirement.project_id == project_id)
    if sprint_id:
        query = query.filter(models.Requirement.sprint_id == sprint_id)
    if status:
        query = query.filter(models.Requirement.status == status)
    if priority:
        query = query.filter(models.Requirement.priority == priority)
    if keyword:
        query = query.filter(models.Requirement.title.contains(keyword))
    total = query.count()
    items = query.order_by(models.Requirement.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}


@app.get("/api/requirements/{req_id}", response_model=schemas.RequirementOut)
def get_requirement(
    req_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取单个需求（含子需求）"""
    from sqlalchemy.orm import subqueryload
    db_req = db.query(models.Requirement).options(
        joinedload(models.Requirement.project),
        joinedload(models.Requirement.sprint),
        joinedload(models.Requirement.assignee),
        joinedload(models.Requirement.creator),
        subqueryload(models.Requirement.children).options(
            joinedload(models.Requirement.assignee),
            joinedload(models.Requirement.project),
            joinedload(models.Requirement.sprint),
        ),
    ).filter(models.Requirement.id == req_id).first()
    if not db_req:
        raise HTTPException(status_code=404, detail="需求不存在")
    return db_req


@app.post("/api/requirements", response_model=schemas.RequirementOut)
def create_requirement(
    req: schemas.RequirementCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建需求"""
    data = req.model_dump()
    data['created_by'] = current_user.id
    db_req = models.Requirement(**data)
    db.add(db_req)
    db.commit()
    db.refresh(db_req)
    db_req = db.query(models.Requirement).options(
        joinedload(models.Requirement.project),
        joinedload(models.Requirement.sprint),
        joinedload(models.Requirement.assignee),
        joinedload(models.Requirement.creator),
    ).filter(models.Requirement.id == db_req.id).first()
    return db_req


@app.put("/api/requirements/{req_id}", response_model=schemas.RequirementOut)
def update_requirement(
    req_id: int,
    req: schemas.RequirementUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新需求"""
    db_req = db.query(models.Requirement).filter(models.Requirement.id == req_id).first()
    if not db_req:
        raise HTTPException(status_code=404, detail="需求不存在")
    for key, value in req.model_dump(exclude_unset=True).items():
        setattr(db_req, key, value)
    db.commit()
    db.refresh(db_req)
    db_req = db.query(models.Requirement).options(
        joinedload(models.Requirement.project),
        joinedload(models.Requirement.sprint),
        joinedload(models.Requirement.assignee),
        joinedload(models.Requirement.creator),
    ).filter(models.Requirement.id == db_req.id).first()
    return db_req


@app.delete("/api/requirements/{req_id}")
def delete_requirement(
    req_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除需求"""
    db_req = db.query(models.Requirement).filter(models.Requirement.id == req_id).first()
    if not db_req:
        raise HTTPException(status_code=404, detail="需求不存在")
    db.delete(db_req)
    db.commit()
    return {"message": "需求已删除"}


# ==================== 任务管理 ====================

@app.get("/api/worktasks")
def get_worktasks(
    project_id: Optional[int] = None,
    sprint_id: Optional[int] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取任务列表（仅顶级任务，每条含 children 子任务）"""
    from sqlalchemy.orm import subqueryload
    query = db.query(models.WorkTask).options(
        joinedload(models.WorkTask.project),
        joinedload(models.WorkTask.sprint),
        joinedload(models.WorkTask.assignee),
        joinedload(models.WorkTask.creator),
        subqueryload(models.WorkTask.children).options(
            joinedload(models.WorkTask.assignee),
            joinedload(models.WorkTask.project),
            joinedload(models.WorkTask.sprint),
        ),
    ).filter(models.WorkTask.parent_id == None)
    if project_id:
        query = query.filter(models.WorkTask.project_id == project_id)
    if sprint_id:
        query = query.filter(models.WorkTask.sprint_id == sprint_id)
    if status:
        query = query.filter(models.WorkTask.status == status)
    if priority:
        query = query.filter(models.WorkTask.priority == priority)
    if keyword:
        query = query.filter(models.WorkTask.title.contains(keyword))
    total = query.count()
    items = query.order_by(models.WorkTask.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}


@app.get("/api/worktasks/{task_id}", response_model=schemas.WorkTaskOut)
def get_worktask(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取单个任务（含子任务）"""
    from sqlalchemy.orm import subqueryload
    db_task = db.query(models.WorkTask).options(
        joinedload(models.WorkTask.project),
        joinedload(models.WorkTask.sprint),
        joinedload(models.WorkTask.assignee),
        joinedload(models.WorkTask.creator),
        subqueryload(models.WorkTask.children).options(
            joinedload(models.WorkTask.assignee),
            joinedload(models.WorkTask.project),
            joinedload(models.WorkTask.sprint),
        ),
    ).filter(models.WorkTask.id == task_id).first()
    if not db_task:
        raise HTTPException(status_code=404, detail="任务不存在")
    return db_task


@app.post("/api/worktasks", response_model=schemas.WorkTaskOut)
def create_worktask(
    task: schemas.WorkTaskCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建任务"""
    data = task.model_dump()
    data['created_by'] = current_user.id
    db_task = models.WorkTask(**data)
    db.add(db_task)
    db.commit()
    db.refresh(db_task)
    db_task = db.query(models.WorkTask).options(
        joinedload(models.WorkTask.project),
        joinedload(models.WorkTask.sprint),
        joinedload(models.WorkTask.assignee),
        joinedload(models.WorkTask.creator),
    ).filter(models.WorkTask.id == db_task.id).first()
    return db_task


@app.put("/api/worktasks/{task_id}", response_model=schemas.WorkTaskOut)
def update_worktask(
    task_id: int,
    task: schemas.WorkTaskUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新任务"""
    db_task = db.query(models.WorkTask).filter(models.WorkTask.id == task_id).first()
    if not db_task:
        raise HTTPException(status_code=404, detail="任务不存在")
    for key, value in task.model_dump(exclude_unset=True).items():
        setattr(db_task, key, value)
    db.commit()
    db.refresh(db_task)
    db_task = db.query(models.WorkTask).options(
        joinedload(models.WorkTask.project),
        joinedload(models.WorkTask.sprint),
        joinedload(models.WorkTask.assignee),
        joinedload(models.WorkTask.creator),
    ).filter(models.WorkTask.id == db_task.id).first()
    return db_task


@app.delete("/api/worktasks/{task_id}")
def delete_worktask(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除任务（含子任务时子任务 parent_id 由外键置空，变为顶级任务）"""
    db_task = db.query(models.WorkTask).filter(models.WorkTask.id == task_id).first()
    if not db_task:
        raise HTTPException(status_code=404, detail="任务不存在")
    db.delete(db_task)
    db.commit()
    return {"message": "任务已删除"}


# ==================== 用户管理 ====================

@app.get("/api/users")
def get_users(
    page: int = 1,
    page_size: int = 10,
    role: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """获取用户列表（服务端分页）"""
    query = db.query(models.User)
    if role:
        query = query.filter(models.User.role == role)
    if status:
        query = query.filter(models.User.status == status)
    if keyword:
        query = query.filter(
            models.User.username.contains(keyword) | models.User.email.contains(keyword)
        )
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}

@app.post("/api/users", response_model=schemas.User)
def create_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    """创建用户"""
    if db.query(models.User).filter(models.User.username == user.username).first():
        raise HTTPException(status_code=400, detail="用户名已存在")
    if db.query(models.User).filter(models.User.email == user.email).first():
        raise HTTPException(status_code=400, detail="邮箱已存在")
    
    # 如果没有提供密码，默认使用用户名作为密码
    if not user.password:
        password = user.username
    else:
        password = user.password
    
    # 加密密码
    hashed_password = hash_password(password)
    
    # 使用 model_dump 而不是 dict（Pydantic V2）
    user_data = user.model_dump(exclude={'password'})
    user_data['password'] = hashed_password
    
    db_user = models.User(**user_data)
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.get("/api/users/{user_id}", response_model=schemas.User)
def get_user(user_id: int, db: Session = Depends(get_db)):
    """获取用户详情"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    return user

@app.put("/api/users/{user_id}", response_model=schemas.User)
def update_user(user_id: int, user: schemas.UserUpdate, db: Session = Depends(get_db)):
    """更新用户"""
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    for key, value in user.model_dump(exclude_unset=True).items():
        setattr(db_user, key, value)
    
    db.commit()
    db.refresh(db_user)
    return db_user

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    """删除用户"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    db.delete(user)
    db.commit()
    return {"message": "用户已删除"}

# ==================== 缺陷管理 ====================

def generate_bug_key(db: Session, project_id: int) -> str:
    """生成缺陷唯一Key"""
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    prefix = (project.key or f"project{project.id}").strip()
    if not prefix:
        prefix = f"project{project.id}"

    # 获取同前缀的所有 bug_key（bug_key 全表唯一，需按前缀全局计算避免冲突）
    bugs = db.query(models.Bug).filter(
        models.Bug.bug_key.isnot(None),
        models.Bug.bug_key.like(f"{prefix}-%"),
    ).all()
    
    max_number = 0
    
    for bug in bugs:
        if bug.bug_key and bug.bug_key.lower().startswith(prefix.lower() + '-'):
            try:
                # 提取编号部分
                number_str = bug.bug_key.split('-')[-1]
                number = int(number_str)
                if number > max_number:
                    max_number = number
            except (ValueError, IndexError):
                continue
    
    # 新编号 = 最大编号 + 1
    new_number = max_number + 1
    return f"{prefix}-{new_number:04d}"

@app.get("/api/bugs")
def get_bugs(
    page: int = 1,
    page_size: int = 10,
    project_id: Optional[int] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    assignee_id: Optional[int] = None,
    reporter_id: Optional[int] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """获取缺陷列表（服务端分页）"""
    query = db.query(models.Bug).options(
        joinedload(models.Bug.project),
        joinedload(models.Bug.assignee),
        joinedload(models.Bug.reporter),
    )

    if project_id:
        query = query.filter(models.Bug.project_id == project_id)
    if status:
        query = query.filter(models.Bug.status == status)
    if priority:
        query = query.filter(models.Bug.priority == priority)
    if assignee_id:
        query = query.filter(models.Bug.assignee_id == assignee_id)
    if reporter_id:
        query = query.filter(models.Bug.reporter_id == reporter_id)
    if keyword:
        query = query.filter(
            models.Bug.title.contains(keyword) | models.Bug.description.contains(keyword)
        )

    total = query.count()
    skip = (page - 1) * page_size
    bugs = query.order_by(models.Bug.created_at.desc()).offset(skip).limit(page_size).all()
    return {"total": total, "items": bugs, "page": page, "page_size": page_size}


@app.get("/api/bugs/export")
def export_bugs(
    project_id: Optional[int] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    assignee_id: Optional[int] = None,
    reporter_id: Optional[int] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """导出缺陷列表为 Excel（按照当前查询条件）"""
    query = db.query(models.Bug).options(
        joinedload(models.Bug.project),
        joinedload(models.Bug.assignee),
        joinedload(models.Bug.reporter),
    )

    if project_id:
        query = query.filter(models.Bug.project_id == project_id)
    if status:
        query = query.filter(models.Bug.status == status)
    if priority:
        query = query.filter(models.Bug.priority == priority)
    if assignee_id:
        query = query.filter(models.Bug.assignee_id == assignee_id)
    if reporter_id:
        query = query.filter(models.Bug.reporter_id == reporter_id)
    if keyword:
        query = query.filter(
            models.Bug.title.contains(keyword) | models.Bug.description.contains(keyword)
        )

    bugs = query.order_by(models.Bug.created_at.desc()).all()

    # 导出字段配置：字段名、表头
    columns = [
        ("bug_key", "缺陷编号"),
        ("project_name", "项目"),
        ("title", "标题"),
        ("status", "状态"),
        ("type", "缺陷类型"),
        ("priority", "优先级"),
        ("severity", "缺陷级别"),
        ("assignee_name", "处理人"),
        ("reporter_name", "创建人"),
        ("version", "版本"),
        ("module", "迭代"),
        # 按“标题、环境、页面、描述”的顺序输出
        ("environment", "环境"),
        ("page_url", "页面"),
        ("description", "描述"),
        ("steps_to_reproduce", "复现步骤"),
        ("expected_result", "期望结果"),
        ("actual_result", "实际结果"),
        ("resolution", "解决结果"),
        ("due_date", "截止日期"),
        ("estimated_hours", "预估工时"),
        ("actual_hours", "实际工时"),
        ("tags_text", "标签（逗号分隔）"),
        ("created_at", "创建时间"),
        ("updated_at", "更新时间"),
    ]

    STATUS_CN   = {"open": "待处理", "in_progress": "进行中", "resolved": "已解决", "closed": "已关闭", "reopened": "重新打开", "pending": "待定"}
    TYPE_CN     = {"bug": "缺陷", "defect": "故障", "improvement": "改进", "task": "任务"}
    PRIORITY_CN = {"urgent": "紧急", "high": "高", "medium": "中", "low": "低"}
    SEVERITY_CN = {"fatal": "致命", "serious": "严重", "general": "一般", "slight": "提示", "suggestion": "建议"}
    RESOLUTION_CN = {"fixed": "已修复", "wontfix": "不修复", "duplicate": "重复", "cannot_reproduce": "无法复现", "deferred": "延期处理", "": ""}

    def bug_to_row(bug: models.Bug) -> list:
        tags_text = ""
        if bug.tags:
            try:
                if isinstance(bug.tags, list):
                    tags_text = ",".join(str(t) for t in bug.tags)
                else:
                    tags_text = ",".join(str(t) for t in (bug.tags or []))
            except Exception:
                tags_text = ""
        return [
            bug.bug_key,
            bug.project.name if bug.project else "",
            bug.title,
            STATUS_CN.get(bug.status, bug.status),
            TYPE_CN.get(bug.type, bug.type),
            PRIORITY_CN.get(bug.priority, bug.priority),
            SEVERITY_CN.get(bug.severity, bug.severity),
            (bug.assignee.display_name or bug.assignee.username) if bug.assignee else "",
            (bug.reporter.display_name or bug.reporter.username) if bug.reporter else "",
            bug.version or "",
            bug.module or "",
            bug.environment or "",
            bug.page_url or "",
            bug.description or "",
            bug.steps_to_reproduce or "",
            bug.expected_result or "",
            bug.actual_result or "",
            RESOLUTION_CN.get(bug.resolution or "", bug.resolution or ""),
            bug.due_date.isoformat() if bug.due_date else "",
            str(bug.estimated_hours) if bug.estimated_hours is not None else "",
            str(bug.actual_hours) if bug.actual_hours is not None else "",
            tags_text,
            bug.created_at.strftime("%Y-%m-%d %H:%M:%S") if bug.created_at else "",
            bug.updated_at.strftime("%Y-%m-%d %H:%M:%S") if bug.updated_at else "",
        ]

    filename_ts = datetime.now().strftime("%Y%m%d-%H%M%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "缺陷列表"

    # 表头
    ws.append([col[1] for col in columns])
    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for bug in bugs:
        ws.append(bug_to_row(bug))

    # 列宽
    col_widths = {
        "缺陷编号": 14, "项目": 16, "标题": 36, "状态": 12, "缺陷类型": 12,
        "优先级": 10, "缺陷级别": 12, "处理人": 14, "创建人": 14,
        "版本": 12, "迭代": 14, "环境": 16, "页面": 20,
        "描述": 28, "复现步骤": 28, "期望结果": 28, "实际结果": 28,
        "解决结果": 14, "截止日期": 14, "预估工时": 12, "实际工时": 12,
        "标签（逗号分隔）": 20, "创建时间": 18, "更新时间": 18,
    }
    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col[1], 14)

    ws.freeze_panes = "A2"

    # 下拉选项写入隐藏配置表
    export_dropdown_options = {
        "状态":    ["待处理", "进行中", "已解决", "已关闭", "重新打开", "待定"],
        "缺陷类型": ["缺陷", "故障", "改进", "任务"],
        "优先级":  ["紧急", "高", "中", "低"],
        "缺陷级别": ["致命", "严重", "一般", "提示", "建议"],
        "解决结果": ["已修复", "不修复", "重复", "无法复现", "延期处理"],
    }
    ws_cfg = wb.create_sheet("_配置")
    ws_cfg.sheet_state = "hidden"
    cfg_col = 1
    header_cn_to_col = {col[1]: idx + 1 for idx, col in enumerate(columns)}
    for field_cn, options in export_dropdown_options.items():
        for row_i, opt in enumerate(options, start=1):
            ws_cfg.cell(row=row_i, column=cfg_col, value=opt)
        col_letter = get_column_letter(cfg_col)
        formula_range = f"_配置!${col_letter}$1:${col_letter}${len(options)}"
        cfg_col += 1

        data_col_idx = header_cn_to_col.get(field_cn)
        if not data_col_idx:
            continue
        data_col_letter = get_column_letter(data_col_idx)
        total_rows = len(bugs) + 1
        dv = DataValidation(
            type="list",
            formula1=formula_range,
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"{data_col_letter}2:{data_col_letter}{max(total_rows, 2)}"
        ws.add_data_validation(dv)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="bugs-{filename_ts}.xlsx"'
        },
    )


BUG_IMPORT_COLUMNS = [
    # 按“标题、环境、页面、描述”的顺序
    ("title", "标题", True),
    ("environment", "环境", False),
    ("page_url", "页面", False),
    ("description", "描述", False),
    ("status", "状态", False),
    ("type", "缺陷类型", False),
    ("priority", "优先级", False),
    ("severity", "缺陷级别", False),
    ("resolution", "解决结果", False),
    ("assignee_username", "处理人用户名", False),
    ("version", "版本", False),
    ("fix_version", "修复版本", False),
    ("module", "迭代", False),
    ("steps_to_reproduce", "复现步骤", False),
    ("expected_result", "期望结果", False),
    ("actual_result", "实际结果", False),
    ("due_date", "截止日期(YYYY-MM-DD)", False),
    ("estimated_hours", "预估工时", False),
    ("actual_hours", "实际工时", False),
    ("tags", "标签（逗号分隔）", False),
]


@app.get("/api/bugs/import/template")
def get_bug_import_template():
    """下载缺陷导入模板（Excel），字段与新建缺陷表单一致"""
    headers_cn = [col[1] for col in BUG_IMPORT_COLUMNS]
    filename_ts = datetime.now().strftime("%Y%m%d-%H%M%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "缺陷导入模板"

    # 写入表头
    ws.append(headers_cn)

    # 表头样式：蓝色背景、白色加粗字体、居中
    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col in enumerate(BUG_IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col[2]:
            cell.value = f"* {cell.value}"

    # 列宽
    col_widths = {
        "标题": 30, "环境": 15, "页面": 20, "描述": 25,
        "状态": 12, "缺陷类型": 12, "优先级": 10, "缺陷级别": 12,
        "解决结果": 15, "处理人用户名": 15, "版本": 12, "修复版本": 12,
        "迭代": 12, "复现步骤": 25, "期望结果": 25, "实际结果": 25,
        "截止日期(YYYY-MM-DD)": 22, "预估工时": 12, "实际工时": 12,
        "标签（逗号分隔）": 20,
    }
    for col_idx, col in enumerate(BUG_IMPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col[1], 15)

    ws.freeze_panes = "A2"

    # 下拉选项写入隐藏配置表，避免中文在 inline formula 中失效
    dropdown_options = {
        "状态":    ["待处理", "进行中", "已解决", "已关闭", "重新打开", "待定"],
        "缺陷类型": ["缺陷", "故障", "改进", "任务"],
        "优先级":  ["紧急", "高", "中", "低"],
        "缺陷级别": ["致命", "严重", "一般", "提示", "建议"],
    }
    ws_cfg = wb.create_sheet("_配置")
    ws_cfg.sheet_state = "hidden"
    cfg_col = 1
    dv_ranges: dict[str, str] = {}
    for field_cn, options in dropdown_options.items():
        for row_i, opt in enumerate(options, start=1):
            ws_cfg.cell(row=row_i, column=cfg_col, value=opt)
        col_letter = get_column_letter(cfg_col)
        dv_ranges[field_cn] = f"_配置!${col_letter}$1:${col_letter}${len(options)}"
        cfg_col += 1

    header_cn_to_col = {col[1]: idx + 1 for idx, col in enumerate(BUG_IMPORT_COLUMNS)}
    for field_cn, formula_range in dv_ranges.items():
        col_idx = header_cn_to_col.get(field_cn)
        if not col_idx:
            continue
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=formula_range,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="输入不合法",
            error=f"请从下拉列表中选择合法的{field_cn}值",
        )
        dv.sqref = f"{col_letter}2:{col_letter}10000"
        ws.add_data_validation(dv)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="bug-import-template-{filename_ts}.xlsx"'
        },
    )


@app.post("/api/bugs/import")
async def import_bugs(
    project_id: int = Query(..., description="导入到的项目ID（通常为当前筛选项目）"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """从 Excel 导入缺陷，字段与模板一致"""
    # 校验项目和权限（与 create_bug 一致）
    project = db.query(models.Project).options(
        joinedload(models.Project.members)
    ).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")

    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建缺陷")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    # 解析表头 -> 英文字段名
    header_map = {}  # index -> field_name
    rows_data: list[dict] = []

    cn_to_field = {cn: field for field, cn, _ in BUG_IMPORT_COLUMNS}
    # 表头可能带有必填星号前缀，兼容处理
    cn_to_field.update({f"* {cn}": field for field, cn, _ in BUG_IMPORT_COLUMNS})

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    for idx, h in enumerate(headers):
        h = (h or "").strip() if isinstance(h, str) else ""
        if h in cn_to_field:
            header_map[idx] = cn_to_field[h]

    # 校验表头是否包含必填列：标题、描述、状态、优先级、缺陷级别
    required_fields_header = {
        "title": "标题",
        "description": "描述",
        "status": "状态",
        "priority": "优先级",
        "severity": "缺陷级别",
    }
    present_fields = set(header_map.values())
    missing_headers = [cn for field, cn in required_fields_header.items() if field not in present_fields]
    if missing_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Excel 表头缺少必填列：{'、'.join(missing_headers)}，请使用最新模板重新导入",
        )

    for row in rows_iter:
        if not any(row):
            continue
        row_dict: dict = {}
        for idx, cell in enumerate(row):
            field = header_map.get(idx)
            if not field:
                continue
            if cell is None:
                continue
            if isinstance(cell, (int, float, Decimal)):
                value = str(cell)
            elif isinstance(cell, date):
                value = cell.isoformat()
            else:
                value = str(cell).strip()
            row_dict[field] = value
        if row_dict:
            rows_data.append(row_dict)

    if not rows_data:
        raise HTTPException(status_code=400, detail="未解析到任何数据行，请检查模板与内容是否匹配")

    imported = 0
    errors: list[str] = []

    # 预先计算同前缀的 bug_key 最大编号，并缓存已占用的 key
    prefix = (project.key or f"project{project.id}").strip()
    if not prefix:
        prefix = f"project{project.id}"

    existing_bugs = db.query(models.Bug).filter(
        models.Bug.bug_key.isnot(None),
        models.Bug.bug_key.like(f"{prefix}-%"),
    ).all()
    existing_bug_keys = {b.bug_key for b in existing_bugs if b.bug_key}

    # 按标题去重：该项目下已存在的标题不再导入
    existing_titles = {r[0].strip() for r in db.query(models.Bug.title).filter(models.Bug.project_id == project_id).all() if r[0] and str(r[0]).strip()}
    imported_titles_in_batch: set[str] = set()  # 本文件内已导入的标题，避免同文件重复

    max_number = 0
    for bug in existing_bugs:
        if bug.bug_key and bug.bug_key.lower().startswith(prefix.lower() + "-"):
            try:
                number_str = bug.bug_key.split("-")[-1]
                number = int(number_str)
                if number > max_number:
                    max_number = number
            except (ValueError, IndexError):
                continue

    for idx, row in enumerate(rows_data, start=2):  # 从第2行开始（第1行为表头）
        title = (row.get("title") or "").strip()
        description_text = (row.get("description") or "").strip()
        status_text = (row.get("status") or "").strip()
        priority_text = (row.get("priority") or "").strip()
        severity_text = (row.get("severity") or "").strip()

        missing_cells: list[str] = []
        if not title:
            missing_cells.append("标题")
        if not description_text:
            missing_cells.append("描述")
        if not status_text:
            missing_cells.append("状态")
        if not priority_text:
            missing_cells.append("优先级")
        if not severity_text:
            missing_cells.append("缺陷级别")
        if missing_cells:
            errors.append(f"第 {idx} 行：{ '、'.join(missing_cells) } 不能为空")
            continue

        # 按标题去重：标题已存在则跳过（项目内 + 本文件内）
        if title in existing_titles:
            errors.append(f"第 {idx} 行：标题已存在，已跳过")
            continue
        if title in imported_titles_in_batch:
            errors.append(f"第 {idx} 行：标题在本文件中重复，已跳过")
            continue

        CN_PRIORITY = {'紧急': 'urgent', '高': 'high', '中': 'medium', '低': 'low'}
        CN_SEVERITY = {'致命': 'fatal', '严重': 'serious', '一般': 'general', '轻微': 'slight', '提示': 'slight', '建议': 'suggestion'}
        CN_STATUS = {
            '待处理': 'open', '打开': 'open',
            '进行中': 'in_progress', '处理中': 'in_progress',
            '已解决': 'resolved', '已修复': 'resolved',
            '已关闭': 'closed', '关闭': 'closed',
            '重新打开': 'reopened', '重开': 'reopened',
            '待定': 'pending',
            '待验证': 'pending', '验证中': 'pending',
            '已拒绝': 'closed', '拒绝': 'closed',
        }
        CN_TYPE = {
            '缺陷': 'bug',
            '错误': 'bug',
            '功能缺陷': 'bug',
            'UI 界面问题': 'bug',
            '改进': 'improvement',
            '任务': 'task',
            '故障': 'defect',
        }

        VALID_PRIORITY = {'urgent', 'high', 'medium', 'low'}
        VALID_SEVERITY = {'fatal', 'serious', 'general', 'slight', 'suggestion'}
        VALID_STATUS = {'open', 'in_progress', 'resolved', 'closed', 'reopened', 'pending'}
        VALID_TYPE = {'bug', 'defect', 'improvement', 'task'}

        # 已保证这几个字段非空，这里只做中英文映射与合法性校验
        raw_priority = CN_PRIORITY.get(priority_text, priority_text)
        raw_severity = CN_SEVERITY.get(severity_text, severity_text)
        raw_status = CN_STATUS.get(status_text, status_text)

        type_text = (row.get("type") or "").strip()
        raw_type = CN_TYPE.get(type_text, type_text or "bug")

        PRIORITY_HINT = "紧急/高/中/低（或 urgent/high/medium/low）"
        SEVERITY_HINT = "缺陷级别：致命/严重/一般/提示/建议（或 fatal/serious/general/slight/suggestion）"
        STATUS_HINT = "状态：待处理/进行中/已解决/已关闭/重新打开/待定（或 open/in_progress/resolved/closed/reopened/pending）"

        enum_errors = []
        if raw_priority not in VALID_PRIORITY:
            enum_errors.append(f"优先级 '{row.get('priority')}' 不合法，可填：{PRIORITY_HINT}")
        if raw_severity not in VALID_SEVERITY:
            enum_errors.append(f"缺陷级别 '{row.get('severity')}' 不合法，可填：{SEVERITY_HINT}")
        if raw_status not in VALID_STATUS:
            enum_errors.append(f"状态 '{row.get('status')}' 不合法，可填：{STATUS_HINT}")
        # 缺陷类型：未知值一律按 bug 处理，避免数据库枚举错误
        if raw_type not in VALID_TYPE:
            raw_type = "bug"
        if enum_errors:
            errors.append(f"第 {idx} 行：" + "；".join(enum_errors))
            continue

        # 解决结果：仅允许数据库枚举值，否则置空
        resolution_text = (row.get("resolution") or "").strip()
        CN_RESOLUTION = {
            "已修复": "fixed", "修复": "fixed", "fixed": "fixed",
            "不修复": "wontfix", "无法修复": "wontfix", "wontfix": "wontfix",
            "重复": "duplicate", "duplicate": "duplicate",
            "无法复现": "cannot_reproduce", "cannot_reproduce": "cannot_reproduce",
            "延期": "deferred", "deferred": "deferred",
        }
        VALID_RESOLUTION = {"fixed", "wontfix", "duplicate", "cannot_reproduce", "deferred", ""}
        raw_resolution = CN_RESOLUTION.get(resolution_text, resolution_text or "")
        if raw_resolution not in VALID_RESOLUTION:
            raw_resolution = ""

        bug_data: dict = {
            "project_id": project_id,
            "title": title,
            "page_url": row.get("page_url") or None,
            "environment": row.get("environment") or None,
            "description": row.get("description") or None,
            "status": raw_status,
            "type": raw_type,
            "priority": raw_priority,
            "severity": raw_severity,
            "resolution": raw_resolution,
            "version": row.get("version") or None,
            "fix_version": row.get("fix_version") or None,
            "module": row.get("module") or None,
            "steps_to_reproduce": row.get("steps_to_reproduce") or None,
            "expected_result": row.get("expected_result") or None,
            "actual_result": row.get("actual_result") or None,
            "tags": None,
            "due_date": None,
            "estimated_hours": None,
            "actual_hours": None,
        }

        # 处理处理人（按用户名查找）
        assignee_username = (row.get("assignee_username") or "").strip()
        if assignee_username:
            assignee = db.query(models.User).filter(models.User.username == assignee_username).first()
            if not assignee:
                errors.append(f"第 {idx} 行：处理人用户名 '{assignee_username}' 不存在，已忽略该行")
                continue
            bug_data["assignee_id"] = assignee.id

        # 日期与数字字段解析
        if row.get("due_date"):
            try:
                bug_data["due_date"] = datetime.strptime(row["due_date"], "%Y-%m-%d").date()
            except Exception:
                errors.append(f"第 {idx} 行：截止日期格式错误，应为 YYYY-MM-DD")
                continue

        for field_name in ("estimated_hours", "actual_hours"):
            if row.get(field_name):
                try:
                    bug_data[field_name] = Decimal(str(row[field_name]))
                except Exception:
                    errors.append(f"第 {idx} 行：{field_name} 解析失败，应为数字")
                    continue

        tags_text = row.get("tags") or ""
        if tags_text:
            bug_data["tags"] = [t.strip() for t in tags_text.split(",") if t.strip()]

        try:
            bug_create = schemas.BugCreate(
                reporter_id=current_user.id,
                **bug_data,
            )
        except Exception as e:
            errors.append(f"第 {idx} 行：数据校验失败 - {e}")
            continue

        # 生成 bug_key 并创建（在本次导入批次内递增，避免重复；同时与历史数据去重）
        while True:
            max_number += 1
            bug_key = f"{prefix}-{max_number:04d}"
            if bug_key not in existing_bug_keys:
                existing_bug_keys.add(bug_key)
                break
        db_bug = models.Bug(**bug_create.model_dump(), bug_key=bug_key)
        db.add(db_bug)
        imported += 1
        imported_titles_in_batch.add(title)

    if imported > 0:
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            err_msg = str(e) if e else "文件格式不合法，请参考模版后重新导入"
            raise HTTPException(status_code=400, detail=err_msg)

    return {
        "message": f"导入完成，成功 {imported} 条，失败 {len(errors)} 条",
        "imported": imported,
        "errors": errors,
    }

@app.post("/api/bugs", response_model=schemas.Bug)
def create_bug(
    bug: schemas.BugCreate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建缺陷"""
    # 检查项目是否存在并加载成员信息
    project = db.query(models.Project).options(
        joinedload(models.Project.members)
    ).filter(models.Project.id == bug.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查权限：只有项目成员可以创建缺陷
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建缺陷")
    
    # 生成bug_key
    bug_key = generate_bug_key(db, bug.project_id)
    
    # 创建缺陷 - 使用 Pydantic V2 的 model_dump
    bug_data = bug.model_dump()
    bug_data['bug_key'] = bug_key
    db_bug = models.Bug(**bug_data)
    db.add(db_bug)
    db.commit()
    db.refresh(db_bug)
    return db_bug

@app.get("/api/bugs/{bug_id}", response_model=schemas.Bug)
def get_bug(bug_id: int, db: Session = Depends(get_db)):
    """获取缺陷详情"""
    bug = db.query(models.Bug).filter(models.Bug.id == bug_id).first()
    if not bug:
        raise HTTPException(status_code=404, detail="缺陷不存在")
    return bug

@app.put("/api/bugs/{bug_id}", response_model=schemas.Bug)
def update_bug(
    bug_id: int, 
    bug: schemas.BugUpdate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新缺陷"""
    db_bug = db.query(models.Bug).options(
        joinedload(models.Bug.project).joinedload(models.Project.members)
    ).filter(models.Bug.id == bug_id).first()
    if not db_bug:
        raise HTTPException(status_code=404, detail="缺陷不存在")
    
    # 检查权限：只有项目成员可以更新缺陷
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_bug.project, "更新缺陷")
    
    # 使用 Pydantic V2 的 model_dump
    bug_data = bug.model_dump(exclude_unset=True)
    
    # 记录变更历史（跳过附件字段，因为数据太大）
    for key, value in bug_data.items():
        old_value = getattr(db_bug, key)
        if old_value != value:
            # 跳过附件字段，避免数据过大
            if key != 'attachments':
                # 记录历史（简化版，实际应该记录操作用户）
                history = models.BugHistory(
                    bug_id=bug_id,
                    user_id=current_user.id,  # 使用当前操作用户
                    field=key,
                    old_value=str(old_value) if old_value is not None else None,
                    new_value=str(value) if value is not None else None
                )
                db.add(history)
        setattr(db_bug, key, value)
    
    # 更新状态时间戳
    if bug.status == 'resolved' and not db_bug.resolved_at:
        db_bug.resolved_at = datetime.now()
    if bug.status == 'closed' and not db_bug.closed_at:
        db_bug.closed_at = datetime.now()
    
    db.commit()
    db.refresh(db_bug)
    return db_bug

@app.delete("/api/bugs/{bug_id}")
def delete_bug(
    bug_id: int, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除缺陷"""
    bug = db.query(models.Bug).options(
        joinedload(models.Bug.project).joinedload(models.Project.members)
    ).filter(models.Bug.id == bug_id).first()
    if not bug:
        raise HTTPException(status_code=404, detail="缺陷不存在")
    
    # 检查权限：只有项目成员可以删除缺陷
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, bug.project, "删除缺陷")
    
    # 删除缺陷相关的图片文件夹
    bug_image_folder = os.path.join(BUG_IMAGE_DIR, bug.bug_key)
    if os.path.exists(bug_image_folder):
        import shutil
        try:
            shutil.rmtree(bug_image_folder)
        except Exception as e:
            print(f"删除图片文件夹失败: {e}")
    
    db.delete(bug)
    db.commit()
    return {"message": "缺陷已删除"}


@app.post("/api/bugs/batch-delete")
def batch_delete_bugs(
    bug_ids: list[int] = Body(..., embed=True, description="要删除的缺陷ID列表"),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """批量删除缺陷"""
    if not bug_ids:
        raise HTTPException(status_code=400, detail="缺陷ID列表不能为空")

    # 查询所有待删除缺陷，带上项目信息和成员
    bugs = db.query(models.Bug).options(
        joinedload(models.Bug.project).joinedload(models.Project.members)
    ).filter(models.Bug.id.in_(bug_ids)).all()

    if not bugs:
        return {"deleted": 0}

    # 当前用户
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")

    # 权限检查：必须是每个缺陷所属项目的成员
    for bug in bugs:
        if not bug.project:
            continue
        check_project_member_permission(user, bug.project, "删除缺陷")

    # 删除相关图片文件夹
    import shutil
    for bug in bugs:
        bug_image_folder = os.path.join(BUG_IMAGE_DIR, bug.bug_key)
        if os.path.exists(bug_image_folder):
            try:
                shutil.rmtree(bug_image_folder)
            except Exception as e:
                print(f"删除图片文件夹失败: {e}")

    # 批量删除
    for bug in bugs:
        db.delete(bug)
    db.commit()

    return {"deleted": len(bugs)}

# ==================== 缺陷图片管理 ====================

@app.post("/api/bugs/{bug_key}/images")
async def upload_bug_image(
    bug_key: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """上传缺陷截图
    
    图片存储目录结构：
    - images/{bug_key}/{uuid}.{ext}
    
    例如：images/BUG-001/a1b2c3d4.png
    """
    import uuid
    
    # 验证缺陷是否存在
    bug = db.query(models.Bug).filter(models.Bug.bug_key == bug_key).first()
    if not bug:
        raise HTTPException(status_code=404, detail="缺陷不存在")
    
    # 验证文件类型
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/bmp']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="不支持的图片格式，请上传 jpg/png/gif/webp/bmp 格式的图片")
    
    # 创建缺陷图片目录
    bug_image_folder = os.path.join(BUG_IMAGE_DIR, bug_key)
    os.makedirs(bug_image_folder, exist_ok=True)
    
    # 生成唯一文件名
    ext = os.path.splitext(file.filename)[1] if file.filename else '.png'
    if not ext:
        # 根据 content_type 推断扩展名
        ext_map = {
            'image/jpeg': '.jpg',
            'image/png': '.png',
            'image/gif': '.gif',
            'image/webp': '.webp',
            'image/bmp': '.bmp'
        }
        ext = ext_map.get(file.content_type, '.png')
    
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(bug_image_folder, unique_name)
    
    # 保存文件
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    
    # 返回图片访问路径（相对路径）
    image_url = f"/api/bugs/{bug_key}/images/{unique_name}"
    
    return {
        "url": image_url,
        "filename": unique_name,
        "size": len(content),
        "content_type": file.content_type
    }

@app.get("/api/bugs/{bug_key}/images/{filename}")
async def get_bug_image(
    bug_key: str,
    filename: str
):
    """获取缺陷截图"""
    file_path = os.path.join(BUG_IMAGE_DIR, bug_key, filename)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="图片不存在")
    
    # 根据扩展名确定 MIME 类型
    ext = os.path.splitext(filename)[1].lower()
    mime_map = {
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.webp': 'image/webp',
        '.bmp': 'image/bmp'
    }
    media_type = mime_map.get(ext, 'application/octet-stream')
    
    return FileResponse(file_path, media_type=media_type)

@app.delete("/api/bugs/{bug_key}/images/{filename}")
async def delete_bug_image(
    bug_key: str,
    filename: str,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除缺陷截图"""
    file_path = os.path.join(BUG_IMAGE_DIR, bug_key, filename)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="图片不存在")
    
    try:
        os.remove(file_path)
        
        # 如果文件夹为空，删除文件夹
        bug_image_folder = os.path.join(BUG_IMAGE_DIR, bug_key)
        if os.path.isdir(bug_image_folder) and not os.listdir(bug_image_folder):
            os.rmdir(bug_image_folder)
            
        return {"message": "图片已删除"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除图片失败: {str(e)}")

@app.get("/api/bugs/{bug_key}/images")
async def list_bug_images(
    bug_key: str,
    db: Session = Depends(get_db)
):
    """获取缺陷的所有截图列表"""
    bug_image_folder = os.path.join(BUG_IMAGE_DIR, bug_key)
    
    if not os.path.exists(bug_image_folder):
        return {"images": []}
    
    images = []
    for filename in os.listdir(bug_image_folder):
        file_path = os.path.join(bug_image_folder, filename)
        if os.path.isfile(file_path):
            images.append({
                "url": f"/api/bugs/{bug_key}/images/{filename}",
                "filename": filename,
                "size": os.path.getsize(file_path)
            })
    
    return {"images": images}

# ==================== 评论管理 ====================

@app.get("/api/bugs/{bug_id}/comments", response_model=List[schemas.Comment])
def get_comments(bug_id: int, db: Session = Depends(get_db)):
    """获取缺陷评论列表"""
    comments = db.query(models.Comment).filter(
        models.Comment.bug_id == bug_id
    ).order_by(models.Comment.created_at.desc()).all()
    return comments

@app.post("/api/comments", response_model=schemas.Comment)
def create_comment(
    comment: schemas.CommentCreate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建评论"""
    # 获取关联的缺陷并检查项目权限
    bug = db.query(models.Bug).options(
        joinedload(models.Bug.project).joinedload(models.Project.members)
    ).filter(models.Bug.id == comment.bug_id).first()
    if not bug:
        raise HTTPException(status_code=404, detail="缺陷不存在")
    
    # 检查权限：只有项目成员可以创建评论
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, bug.project, "创建评论")
    
    db_comment = models.Comment(**comment.model_dump())
    db.add(db_comment)
    db.commit()
    db.refresh(db_comment)
    return db_comment

# ==================== 统计分析 ====================

@app.get("/api/statistics", response_model=schemas.BugStatistics)
def get_statistics(
    project_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """获取缺陷统计信息"""
    query = db.query(models.Bug)
    if project_id:
        query = query.filter(models.Bug.project_id == project_id)
    
    all_bugs = query.all()
    
    # 统计各状态数量
    status_count = {
        'open': len([b for b in all_bugs if b.status == 'open']),
        'in_progress': len([b for b in all_bugs if b.status == 'in_progress']),
        'resolved': len([b for b in all_bugs if b.status == 'resolved']),
        'closed': len([b for b in all_bugs if b.status == 'closed']),
    }
    
    # 按优先级统计
    priority_count = {}
    for bug in all_bugs:
        priority_count[bug.priority] = priority_count.get(bug.priority, 0) + 1
    
    # 按严重程度统计
    severity_count = {}
    for bug in all_bugs:
        severity_count[bug.severity] = severity_count.get(bug.severity, 0) + 1
    
    # 按类型统计
    type_count = {}
    for bug in all_bugs:
        type_count[bug.type] = type_count.get(bug.type, 0) + 1
    
    return {
        'total': len(all_bugs),
        'open': status_count['open'],
        'in_progress': status_count['in_progress'],
        'resolved': status_count['resolved'],
        'closed': status_count['closed'],
        'by_priority': priority_count,
        'by_severity': severity_count,
        'by_type': type_count
    }

# ==================== 测试用例管理 ====================

def generate_case_key(db: Session, project_id: int) -> str:
    """生成用例唯一标识"""
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 查询该项目下已有的用例，找出最大编号
    existing_cases = db.query(models.TestCase).filter(
        models.TestCase.project_id == project_id
    ).all()
    
    max_number = 0
    for case in existing_cases:
        # 从 case_key 中提取数字部分，格式如 "PROJ-TC-0001"
        try:
            number = int(case.case_key.split('-')[-1])
            if number > max_number:
                max_number = number
        except (ValueError, IndexError):
            continue
    
    next_number = max_number + 1
    return f"{project.key}-TC-{next_number:04d}"

@app.get("/api/testcases")
def get_testcases(
    request: Request,
    project_id: Optional[int] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    module: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=10000),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取测试用例列表（服务端分页）"""
    require_permission(current_user.role, "testcases", "read")

    query = db.query(models.TestCase).options(
        joinedload(models.TestCase.project),
        joinedload(models.TestCase.creator),
        joinedload(models.TestCase.updater)
    )

    if project_id:
        query = query.filter(models.TestCase.project_id == project_id)
    if status:
        query = query.filter(models.TestCase.status == status)
    if priority:
        query = query.filter(models.TestCase.priority == priority)
    if module is not None:
        # 匹配该目录及所有子目录下的用例（前缀匹配）
        query = query.filter(
            (models.TestCase.module == module) |
            (models.TestCase.module.like(f"{module}/%"))
        )
    if search:
        query = query.filter(
            (models.TestCase.case_key.like(f"%{search}%")) |
            (models.TestCase.title.like(f"%{search}%")) |
            (models.TestCase.module.like(f"%{search}%"))
        )

    total = query.count()
    items = query.order_by(models.TestCase.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}

@app.get("/api/testcases/export")
def export_testcases(
    project_id: Optional[int] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    module: Optional[str] = None,
    search: Optional[str] = None,
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """导出测试用例为 Excel 或 CSV"""
    require_permission(current_user.role, "testcases", "read")

    query = db.query(models.TestCase).options(
        joinedload(models.TestCase.project),
        joinedload(models.TestCase.creator)
    )
    if project_id:
        query = query.filter(models.TestCase.project_id == project_id)
    if status:
        query = query.filter(models.TestCase.status == status)
    if priority:
        query = query.filter(models.TestCase.priority == priority)
    if module is not None:
        # 匹配该目录及所有子目录下的用例（前缀匹配）
        query = query.filter(
            (models.TestCase.module == module) |
            (models.TestCase.module.like(f"{module}/%"))
        )
    if search:
        query = query.filter(
            (models.TestCase.case_key.like(f"%{search}%")) |
            (models.TestCase.title.like(f"%{search}%"))
        )

    testcases = query.order_by(models.TestCase.created_at.desc()).all()

    # 计算最多步骤数，最少保留 3 步
    MIN_STEPS = 3
    max_steps = MIN_STEPS
    for tc in testcases:
        if tc.steps and len(tc.steps) > max_steps:
            max_steps = len(tc.steps)

    # 固定基础列：标题、分组、等级、前置条件
    base_headers = ["标题", "分组", "等级", "前置条件"]
    step_headers = []
    for i in range(1, max_steps + 1):
        step_headers += [f"步骤{i}", f"预期结果{i}"]
    headers_cn = base_headers + step_headers

    def build_row(tc):
        steps = tc.steps or []
        row = [
            tc.title or "",
            tc.module or "",
            tc.priority or "P2",
            tc.precondition or "",
        ]
        for i in range(max_steps):
            if i < len(steps):
                s = steps[i] if isinstance(steps[i], dict) else {}
                row.append(s.get("description", "") or "")
                row.append(s.get("expected_result", "") or "")
            else:
                row += ["", ""]
        return row

    filename_ts = datetime.now().strftime("%Y%m%d-%H%M%S")

    if format == "csv":
        file_stream = io.StringIO()
        writer = csv.writer(file_stream)
        writer.writerow(headers_cn)
        for tc in testcases:
            writer.writerow(build_row(tc))
        csv_bytes = file_stream.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="testcases-{filename_ts}.csv"'},
        )

    # xlsx 格式
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(headers_cn)

    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx in range(1, len(headers_cn) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for tc in testcases:
        ws.append(build_row(tc))

    # 列宽：基础4列 + 每对步骤列
    base_widths = [40, 20, 10, 24]
    step_widths = [30, 30] * max_steps
    col_widths = base_widths + step_widths
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="testcases-{filename_ts}.xlsx"'},
    )

@app.post("/api/testcases", response_model=schemas.TestCase)
def create_testcase(
    testcase: schemas.TestCaseCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建测试用例"""
    require_permission(current_user.role, "testcases", "create")
    
    # 检查项目是否存在并加载成员信息
    project = db.query(models.Project).options(
        joinedload(models.Project.members)
    ).filter(models.Project.id == testcase.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查权限：只有项目成员可以创建测试用例
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建测试用例")
    
    # 生成用例唯一标识
    case_key = generate_case_key(db, testcase.project_id)
    
    testcase_data = testcase.model_dump()
    testcase_data['case_key'] = case_key
    
    db_testcase = models.TestCase(**testcase_data)
    db.add(db_testcase)
    db.commit()
    db.refresh(db_testcase)
    return db_testcase

@app.get("/api/testcases/{testcase_id}", response_model=schemas.TestCase)
def get_testcase(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取测试用例详情"""
    require_permission(current_user.role, "testcases", "read")
    
    testcase = db.query(models.TestCase).filter(models.TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    return testcase

@app.put("/api/testcases/{testcase_id}", response_model=schemas.TestCase)
def update_testcase(
    testcase_id: int,
    testcase: schemas.TestCaseUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新测试用例"""
    require_permission(current_user.role, "testcases", "update")
    
    db_testcase = db.query(models.TestCase).options(
        joinedload(models.TestCase.project).joinedload(models.Project.members)
    ).filter(models.TestCase.id == testcase_id).first()
    if not db_testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    # 检查权限：只有项目成员可以更新测试用例
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_testcase.project, "更新测试用例")
    
    update_data = testcase.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_testcase, field, value)
    
    db.commit()
    db.refresh(db_testcase)
    return db_testcase

@app.delete("/api/testcases/{testcase_id}")
def delete_testcase(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除测试用例"""
    require_permission(current_user.role, "testcases", "delete")
    
    db_testcase = db.query(models.TestCase).options(
        joinedload(models.TestCase.project).joinedload(models.Project.members)
    ).filter(models.TestCase.id == testcase_id).first()
    if not db_testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    # 检查权限：只有项目成员可以删除测试用例
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_testcase.project, "删除测试用例")
    
    db.delete(db_testcase)
    db.commit()
    return {"message": "测试用例已删除"}

@app.post("/api/testcases/batch-delete")
def batch_delete_testcases(
    ids: List[int] = Body(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """批量删除测试用例"""
    require_permission(current_user.role, "testcases", "delete")

    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")

    deleted = 0
    for testcase_id in ids:
        db_testcase = db.query(models.TestCase).options(
            joinedload(models.TestCase.project).joinedload(models.Project.members)
        ).filter(models.TestCase.id == testcase_id).first()
        if db_testcase:
            try:
                check_project_member_permission(user, db_testcase.project, "删除测试用例")
                db.delete(db_testcase)
                deleted += 1
            except Exception:
                pass
    db.commit()
    return {"message": f"成功删除 {deleted} 个测试用例", "deleted": deleted}


# 导入支持的基础字段（只读取这4个，其余列自动忽略）
# 基础字段映射（标题必填，其余选填）
TC_BASE_IMPORT_FIELDS = {
    "标题":     "title",
    "分组":     "module",
    "等级":     "priority",
    "前置条件": "precondition",
    # 单列格式：内容为 "[1]xxx\n[2]xxx" 形式
    "步骤":     "_steps_raw",
    "预期结果": "_expected_raw",
}
# 多列格式：步骤1/步骤2... 与 预期结果1/预期结果2...
TC_STEP_COL_RE     = re.compile(r"^步骤(\d+)$")
TC_EXPECTED_COL_RE = re.compile(r"^预期结果(\d+)$")
# [N]text 标记解析
TC_STEP_MARKER_RE  = re.compile(r"\[(\d+)\]\s*(.*?)(?=\[(\d+)\]|$)", re.DOTALL)

def _parse_numbered_text(text: str) -> dict:
    """
    解析 '[1]描述1\n[2]描述2' 格式，返回 {1: '描述1', 2: '描述2'}。
    兼容步骤之间有或无换行的情况。
    """
    result = {}
    if not text:
        return result
    # 先尝试正则全文匹配 [N]content
    for m in TC_STEP_MARKER_RE.finditer(text):
        n   = int(m.group(1))
        val = m.group(2).strip()
        if val:
            result[n] = val
    return result

def _parse_tc_headers(headers_row):
    """
    解析表头，返回 (base_map, step_map, expected_map)
      base_map:     col_idx -> field_name（含 _steps_raw / _expected_raw）
      step_map:     col_idx -> step_number（多列格式：步骤N）
      expected_map: col_idx -> step_number（多列格式：预期结果N）
    """
    base_map:     dict = {}
    step_map:     dict = {}
    expected_map: dict = {}
    for idx, h in enumerate(headers_row):
        h_str = str(h or "").strip()
        if h_str in TC_BASE_IMPORT_FIELDS:
            base_map[idx] = TC_BASE_IMPORT_FIELDS[h_str]
        elif m := TC_STEP_COL_RE.match(h_str):
            step_map[idx] = int(m.group(1))
        elif m := TC_EXPECTED_COL_RE.match(h_str):
            expected_map[idx] = int(m.group(1))
    return base_map, step_map, expected_map

def _build_steps(step_descs: dict, step_expected: dict) -> list:
    """根据步骤描述和预期结果字典（{step_num: text}）合并成步骤列表"""
    all_nums = sorted(set(step_descs) | set(step_expected))
    steps = []
    for n in all_nums:
        desc = step_descs.get(n, "").strip()
        exp  = step_expected.get(n, "").strip()
        if desc or exp:
            steps.append({"description": desc, "expected_result": exp})
    return steps

def _parse_cell(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, (int, float)):
        return str(cell)
    return str(cell).strip()

@app.post("/api/testcases/import")
async def import_testcases(
    project_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """从 Excel / CSV 导入测试用例（只解析：标题、分组、等级、前置条件、步骤N、预期结果N）"""
    require_permission(current_user.role, "testcases", "create")

    project = db.query(models.Project).options(
        joinedload(models.Project.members)
    ).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")

    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建测试用例")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    filename_lower = (file.filename or "").lower()
    is_csv = filename_lower.endswith(".csv")

    rows_data: list[dict] = []

    def process_rows(headers_raw, data_rows_iter, get_cell):
        base_map, step_map, expected_map = _parse_tc_headers(headers_raw)
        if "title" not in base_map.values():
            raise HTTPException(
                status_code=400,
                detail="文件表头缺少必填列「标题」，请使用导出的文件重新导入"
            )
        for row in data_rows_iter:
            row_base: dict = {}
            step_descs:    dict = {}
            step_expected: dict = {}
            has_content = False
            for idx, cell in enumerate(row):
                v = get_cell(cell)
                if not v:
                    continue
                has_content = True
                if idx in base_map:
                    field = base_map[idx]
                    row_base[field] = v
                elif idx in step_map:
                    step_descs[step_map[idx]] = v
                elif idx in expected_map:
                    step_expected[expected_map[idx]] = v

            # 处理单列 [N]text 格式（"步骤" / "预期结果" 列）
            if "_steps_raw" in row_base:
                parsed = _parse_numbered_text(row_base.pop("_steps_raw"))
                # 不覆盖多列格式已有的值
                for n, v in parsed.items():
                    step_descs.setdefault(n, v)
            if "_expected_raw" in row_base:
                parsed = _parse_numbered_text(row_base.pop("_expected_raw"))
                for n, v in parsed.items():
                    step_expected.setdefault(n, v)

            if has_content and row_base.get("title", "").strip():
                row_base["steps"] = _build_steps(step_descs, step_expected)
                rows_data.append(row_base)

    if is_csv:
        text = content.decode("utf-8-sig").strip()
        reader = csv.reader(io.StringIO(text))
        try:
            headers_raw = next(reader)
        except StopIteration:
            raise HTTPException(status_code=400, detail="CSV 文件为空")
        process_rows(headers_raw, reader, lambda c: c.strip() if isinstance(c, str) else "")
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers_raw = next(rows_iter)
        except StopIteration:
            raise HTTPException(status_code=400, detail="Excel 文件为空")
        process_rows(headers_raw, rows_iter, _parse_cell)

    if not rows_data:
        raise HTTPException(status_code=400, detail="未解析到有效数据行，请检查文件内容")

    # 预加载该项目下已有的 (module, title) 组合，用于去重
    existing_keys: set[tuple[str, str]] = {
        (tc.module or "", tc.title or "")
        for tc in db.query(models.TestCase.module, models.TestCase.title)
                     .filter(models.TestCase.project_id == project_id)
                     .all()
    }
    # 同一批文件内也去重（避免文件中重复行）
    imported_keys: set[tuple[str, str]] = set()

    priority_valid = {"P0", "P1", "P2", "P3", "P4"}
    imported = 0
    skipped = 0
    errors: list[str] = []

    for i, row in enumerate(rows_data, start=2):
        title = row.get("title", "").strip()
        if not title:
            errors.append(f"第 {i} 行：标题为空，已跳过")
            continue
        try:
            module = row.get("module", "").strip()
            dedup_key = (module, title)

            # 分组 + 标题 已存在，跳过
            if dedup_key in existing_keys or dedup_key in imported_keys:
                skipped += 1
                continue

            priority = row.get("priority", "P2").strip()
            if priority not in priority_valid:
                priority = "P2"

            case_key = generate_case_key(db, project_id)
            tc = models.TestCase(
                case_key=case_key,
                project_id=project_id,
                title=title,
                module=module,
                precondition=row.get("precondition", "").strip(),
                steps=row.get("steps", []),
                expected_result="",
                priority=priority,
                type="functional",
                status="draft",
                tags=[],
                created_by=current_user.id,
            )
            db.add(tc)
            db.flush()
            imported_keys.add(dedup_key)
            imported += 1
        except Exception as e:
            errors.append(f"第 {i} 行（{title}）：{str(e)}")

    db.commit()
    parts = [f"成功导入 {imported} 个用例"]
    if skipped:
        parts.append(f"{skipped} 个已存在跳过")
    if errors:
        parts.append(f"{len(errors)} 个失败")
    msg = "，".join(parts)
    return {"message": msg, "imported": imported, "skipped": skipped, "errors": errors}


# ==================== 用例目录管理 ====================

@app.get("/api/testcase-directories", response_model=List[schemas.TestCaseDirectoryResponse])
def get_testcase_directories(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取指定项目的所有用例目录"""
    require_permission(current_user.role, "testcases", "read")
    return db.query(models.TestCaseDirectory).filter(
        models.TestCaseDirectory.project_id == project_id
    ).order_by(models.TestCaseDirectory.path).all()


@app.post("/api/testcase-directories", response_model=schemas.TestCaseDirectoryResponse)
def create_testcase_directory(
    data: schemas.TestCaseDirectoryCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建用例目录"""
    require_permission(current_user.role, "testcases", "create")
    # 同项目下路径唯一
    exists = db.query(models.TestCaseDirectory).filter(
        models.TestCaseDirectory.project_id == data.project_id,
        models.TestCaseDirectory.path == data.path
    ).first()
    if exists:
        raise HTTPException(status_code=400, detail="该目录已存在")
    directory = models.TestCaseDirectory(
        project_id=data.project_id,
        path=data.path,
        name=data.name
    )
    db.add(directory)
    db.commit()
    db.refresh(directory)
    return directory


@app.put("/api/testcase-directories/{directory_id}", response_model=schemas.TestCaseDirectoryResponse)
def update_testcase_directory(
    directory_id: int,
    data: schemas.TestCaseDirectoryUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新目录（重命名时同步更新子目录路径）"""
    require_permission(current_user.role, "testcases", "update")
    directory = db.query(models.TestCaseDirectory).filter(
        models.TestCaseDirectory.id == directory_id
    ).first()
    if not directory:
        raise HTTPException(status_code=404, detail="目录不存在")
    old_path = directory.path
    if data.path is not None:
        new_path = data.path
        # 同步更新所有子目录的路径前缀
        sub_dirs = db.query(models.TestCaseDirectory).filter(
            models.TestCaseDirectory.project_id == directory.project_id,
            models.TestCaseDirectory.path.like(f"{old_path}/%")
        ).all()
        for sub in sub_dirs:
            sub.path = new_path + sub.path[len(old_path):]
        # 同步更新该目录下用例的 module 字段
        db.query(models.TestCase).filter(
            models.TestCase.project_id == directory.project_id,
            models.TestCase.module.like(f"{old_path}/%") | (models.TestCase.module == old_path)
        ).all()
        testcases = db.query(models.TestCase).filter(
            models.TestCase.project_id == directory.project_id,
        ).all()
        for tc in testcases:
            if tc.module == old_path:
                tc.module = new_path
            elif tc.module and tc.module.startswith(old_path + "/"):
                tc.module = new_path + tc.module[len(old_path):]
        directory.path = new_path
    if data.name is not None:
        directory.name = data.name
    db.commit()
    db.refresh(directory)
    return directory


@app.delete("/api/testcase-directories/{directory_id}")
def delete_testcase_directory(
    directory_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除目录及其所有子目录"""
    require_permission(current_user.role, "testcases", "delete")
    directory = db.query(models.TestCaseDirectory).filter(
        models.TestCaseDirectory.id == directory_id
    ).first()
    if not directory:
        raise HTTPException(status_code=404, detail="目录不存在")
    path = directory.path
    # 删除该目录及所有子目录记录
    db.query(models.TestCaseDirectory).filter(
        models.TestCaseDirectory.project_id == directory.project_id,
        (models.TestCaseDirectory.path == path) |
        models.TestCaseDirectory.path.like(f"{path}/%")
    ).delete(synchronize_session=False)
    # 删除该目录下的所有用例
    db.query(models.TestCase).filter(
        models.TestCase.project_id == directory.project_id,
        (models.TestCase.module == path) |
        models.TestCase.module.like(f"{path}/%")
    ).delete(synchronize_session=False)
    db.commit()
    return {"message": "目录已删除"}


# ==================== 用例评审管理 ====================

def calculate_review_status(start_date, end_date):
    """根据当前时间和起始/截止时间计算评审状态"""
    from datetime import date
    today = date.today()
    
    if today < start_date:
        return 'not_started'
    elif start_date <= today <= end_date:
        return 'in_progress'
    else:
        return 'ended'

@app.get("/api/testcase_reviews")
def get_testcase_reviews(
    project_id: Optional[int] = None,
    sprint_id: Optional[int] = None,
    keyword: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取用例评审列表（服务端分页）"""
    query = db.query(models.TestCaseReview).options(
        joinedload(models.TestCaseReview.project),
        joinedload(models.TestCaseReview.sprint),
        joinedload(models.TestCaseReview.initiator),
        noload(models.TestCaseReview.review_items)
    )

    if project_id:
        query = query.filter(models.TestCaseReview.project_id == project_id)
    if sprint_id:
        query = query.filter(models.TestCaseReview.sprint_id == sprint_id)
    if keyword:
        query = query.filter(models.TestCaseReview.name.contains(keyword))

    total = query.count()
    reviews = query.order_by(models.TestCaseReview.start_date.desc()).offset((page - 1) * page_size).limit(page_size).all()

    for review in reviews:
        calculated_status = calculate_review_status(review.start_date, review.end_date)
        if review.status != calculated_status:
            review.status = calculated_status
            db.commit()
            db.refresh(review)

    return {"total": total, "items": reviews, "page": page, "page_size": page_size}

@app.post("/api/testcase_reviews", response_model=schemas.TestCaseReview)
def create_testcase_review(
    review: schemas.TestCaseReviewCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建用例评审"""
    # 检查项目是否存在
    project = db.query(models.Project).filter(models.Project.id == review.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 如果指定了迭代，检查迭代是否存在
    if review.sprint_id:
        sprint = db.query(models.Sprint).filter(models.Sprint.id == review.sprint_id).first()
        if not sprint:
            raise HTTPException(status_code=404, detail="迭代不存在")
    
    # 验证日期
    if review.end_date < review.start_date:
        raise HTTPException(status_code=400, detail="截止时间不能早于发起时间")
    
    # 计算状态
    status = calculate_review_status(review.start_date, review.end_date)
    
    review_data = review.model_dump()
    review_data['status'] = status
    db_review = models.TestCaseReview(**review_data)
    db.add(db_review)
    db.commit()
    db.refresh(db_review)
    
    # 重新加载关联数据
    db_review = db.query(models.TestCaseReview).options(
        joinedload(models.TestCaseReview.project),
        joinedload(models.TestCaseReview.sprint),
        joinedload(models.TestCaseReview.initiator)
    ).filter(models.TestCaseReview.id == db_review.id).first()
    
    return db_review

@app.get("/api/testcase_reviews/{review_id}", response_model=schemas.TestCaseReview)
def get_testcase_review(
    review_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取用例评审详情"""
    # 使用 selectinload 加载 review_items，然后使用 joinedload 加载嵌套关系
    # 注意：不能对同一个关系使用多次 selectinload，所以先加载 review_items，再分别加载嵌套关系
    review = db.query(models.TestCaseReview).options(
        joinedload(models.TestCaseReview.project),
        joinedload(models.TestCaseReview.sprint),
        joinedload(models.TestCaseReview.initiator),
        selectinload(models.TestCaseReview.review_items)
    ).filter(models.TestCaseReview.id == review_id).first()
    
    if review and review.review_items:
        # 使用单独的查询来加载嵌套关系，避免循环引用
        item_ids = [item.id for item in review.review_items]
        # 预加载 testcase 和 reviewer
        db.query(models.TestCaseReviewItem).filter(
            models.TestCaseReviewItem.id.in_(item_ids)
        ).options(
            joinedload(models.TestCaseReviewItem.testcase),
            joinedload(models.TestCaseReviewItem.reviewer)
        ).all()
    
    if not review:
        raise HTTPException(status_code=404, detail="用例评审不存在")
    
    # 更新状态
    calculated_status = calculate_review_status(review.start_date, review.end_date)
    if review.status != calculated_status:
        review.status = calculated_status
        db.commit()
        # 注意：refresh 可能会重新加载关系，所以我们不 refresh，直接使用当前对象
    
    # 手动构建字典，避免循环引用
    # 因为 model_validate 在验证阶段就会检测到循环引用，即使 review 字段为 None
    review_dict = {
        'id': review.id,
        'project_id': review.project_id,
        'sprint_id': review.sprint_id,
        'name': review.name,
        'initiator_id': review.initiator_id,
        'start_date': review.start_date,
        'end_date': review.end_date,
        'status': review.status,
        'created_at': review.created_at,
        'updated_at': review.updated_at,
        'project': schemas.Project.model_validate(review.project).model_dump() if review.project else None,
        'sprint': schemas.Sprint.model_validate(review.sprint).model_dump() if review.sprint else None,
        'initiator': schemas.User.model_validate(review.initiator).model_dump() if review.initiator else None,
    }
    
    # 手动构建 review_items，排除 review 字段以避免循环引用
    if review.review_items:
        review_items_data = []
        for item in review.review_items:
            item_dict = {
                'id': item.id,
                'review_id': item.review_id,
                'testcase_id': item.testcase_id,
                'reviewer_id': item.reviewer_id,
                'status': item.status,
                'comments': item.comments,
                'reviewed_at': item.reviewed_at,
                'created_at': item.created_at,
                'updated_at': item.updated_at,
                'testcase': schemas.TestCase.model_validate(item.testcase).model_dump() if item.testcase else None,
                'reviewer': schemas.User.model_validate(item.reviewer).model_dump() if item.reviewer else None,
                # 不包含 review 字段，避免循环引用
            }
            review_items_data.append(item_dict)
        review_dict['review_items'] = review_items_data
    else:
        review_dict['review_items'] = None
    
    # 验证构建的字典
    return schemas.TestCaseReview.model_validate(review_dict)

@app.put("/api/testcase_reviews/{review_id}", response_model=schemas.TestCaseReview)
def update_testcase_review(
    review_id: int,
    review: schemas.TestCaseReviewUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新用例评审"""
    db_review = db.query(models.TestCaseReview).filter(models.TestCaseReview.id == review_id).first()
    if not db_review:
        raise HTTPException(status_code=404, detail="用例评审不存在")
    
    # 检查项目是否存在
    if review.project_id:
        project = db.query(models.Project).filter(models.Project.id == review.project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查迭代是否存在
    if review.sprint_id:
        sprint = db.query(models.Sprint).filter(models.Sprint.id == review.sprint_id).first()
        if not sprint:
            raise HTTPException(status_code=404, detail="迭代不存在")
    
    # 验证日期
    start_date = review.start_date if review.start_date else db_review.start_date
    end_date = review.end_date if review.end_date else db_review.end_date
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="截止时间不能早于发起时间")
    
    update_data = review.model_dump(exclude_unset=True)
    
    # 如果更新了日期，重新计算状态
    if review.start_date or review.end_date:
        calculated_status = calculate_review_status(start_date, end_date)
        update_data['status'] = calculated_status
    
    for key, value in update_data.items():
        setattr(db_review, key, value)
    
    db.commit()
    db.refresh(db_review)
    
    # 重新加载关联数据
    db_review = db.query(models.TestCaseReview).options(
        joinedload(models.TestCaseReview.project),
        joinedload(models.TestCaseReview.sprint),
        joinedload(models.TestCaseReview.initiator)
    ).filter(models.TestCaseReview.id == db_review.id).first()
    
    return db_review

@app.delete("/api/testcase_reviews/{review_id}")
def delete_testcase_review(
    review_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除用例评审"""
    review = db.query(models.TestCaseReview).filter(models.TestCaseReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="用例评审不存在")
    
    db.delete(review)
    db.commit()
    return {"message": "用例评审已删除"}

# ==================== 用例评审项管理 ====================

@app.get("/api/testcase_reviews/{review_id}/items", response_model=List[schemas.TestCaseReviewItem])
def get_review_items(
    review_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取评审的用例列表"""
    review = db.query(models.TestCaseReview).filter(models.TestCaseReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="用例评审不存在")
    
    items = db.query(models.TestCaseReviewItem).options(
        joinedload(models.TestCaseReviewItem.testcase).joinedload(models.TestCase.creator),
        joinedload(models.TestCaseReviewItem.reviewer)
    ).filter(models.TestCaseReviewItem.review_id == review_id).all()
    
    # 清理循环引用：将所有 items 中的 review 字段设为 None
    for item in items:
        item.review = None
    
    return items

@app.post("/api/testcase_reviews/{review_id}/items", response_model=schemas.TestCaseReviewItem)
def add_review_item(
    review_id: int,
    item: schemas.TestCaseReviewItemCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """添加用例到评审"""
    review = db.query(models.TestCaseReview).filter(models.TestCaseReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="用例评审不存在")
    
    # 检查用例是否存在
    testcase = db.query(models.TestCase).filter(models.TestCase.id == item.testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    # 检查是否已经添加过
    existing = db.query(models.TestCaseReviewItem).filter(
        models.TestCaseReviewItem.review_id == review_id,
        models.TestCaseReviewItem.testcase_id == item.testcase_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="该用例已经在此评审中")
    
    item_data = item.model_dump()
    item_data['review_id'] = review_id
    db_item = models.TestCaseReviewItem(**item_data)
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    
    # 重新加载关联数据（不加载 review 关系以避免循环引用）
    db_item = db.query(models.TestCaseReviewItem).options(
        joinedload(models.TestCaseReviewItem.testcase).joinedload(models.TestCase.creator),
        joinedload(models.TestCaseReviewItem.reviewer)
    ).filter(models.TestCaseReviewItem.id == db_item.id).first()
    
    # 清理循环引用：将 review 字段设为 None
    db_item.review = None
    
    return db_item

@app.put("/api/testcase_reviews/{review_id}/items/{item_id}", response_model=schemas.TestCaseReviewItem)
def update_review_item(
    review_id: int,
    item_id: int,
    item: schemas.TestCaseReviewItemUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新评审结果"""
    db_item = db.query(models.TestCaseReviewItem).filter(
        models.TestCaseReviewItem.id == item_id,
        models.TestCaseReviewItem.review_id == review_id
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="评审项不存在")
    
    update_data = item.model_dump(exclude_unset=True)
    
    # 如果状态改变为已评审，设置评审时间和评审人
    if update_data.get('status') and update_data['status'] != 'pending':
        update_data['reviewed_at'] = datetime.now()
        if not update_data.get('reviewer_id'):
            update_data['reviewer_id'] = current_user.id
    elif update_data.get('status') == 'pending':
        update_data['reviewed_at'] = None
    
    for key, value in update_data.items():
        setattr(db_item, key, value)
    
    db.commit()
    db.refresh(db_item)
    
    # 重新加载关联数据（不加载 review 关系以避免循环引用）
    db_item = db.query(models.TestCaseReviewItem).options(
        joinedload(models.TestCaseReviewItem.testcase).joinedload(models.TestCase.creator),
        joinedload(models.TestCaseReviewItem.reviewer)
    ).filter(models.TestCaseReviewItem.id == db_item.id).first()
    
    # 清理循环引用：将 review 字段设为 None
    db_item.review = None
    
    return db_item

@app.delete("/api/testcase_reviews/{review_id}/items/{item_id}")
def delete_review_item(
    review_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """从评审中移除用例"""
    db_item = db.query(models.TestCaseReviewItem).filter(
        models.TestCaseReviewItem.id == item_id,
        models.TestCaseReviewItem.review_id == review_id
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="评审项不存在")
    
    db.delete(db_item)
    db.commit()
    return {"message": "用例已从评审中移除"}

# ==================== 模型管理 ====================

@app.get("/api/models")
def get_models(
    page: int = 1,
    page_size: int = 10,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取模型列表（服务端分页）"""
    require_permission(current_user.role, "models", "read")
    query = db.query(models.Model).order_by(models.Model.is_default.desc(), models.Model.created_at.desc())
    if keyword:
        query = query.filter(models.Model.name.contains(keyword))
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}

@app.get("/api/models/{model_id}", response_model=schemas.Model)
def get_model(
    model_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取模型详情"""
    require_permission(current_user.role, "models", "read")
    model = db.query(models.Model).filter(models.Model.id == model_id).first()
    if not model:
        raise HTTPException(status_code=404, detail="模型不存在")
    return model

@app.post("/api/models", response_model=schemas.Model)
def create_model(
    model: schemas.ModelCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建模型"""
    require_permission(current_user.role, "models", "create")
    
    # 如果设置为默认，取消其他默认模型
    if model.is_default:
        db.query(models.Model).filter(models.Model.is_default == True).update({"is_default": False})
    
    model_data = model.model_dump()
    db_model = models.Model(**model_data)
    db.add(db_model)
    db.commit()
    db.refresh(db_model)
    return db_model

@app.put("/api/models/{model_id}", response_model=schemas.Model)
def update_model(
    model_id: int,
    model: schemas.ModelUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新模型"""
    require_permission(current_user.role, "models", "update")
    
    db_model = db.query(models.Model).filter(models.Model.id == model_id).first()
    if not db_model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    # 如果设置为默认，取消其他默认模型
    update_data = model.model_dump(exclude_unset=True)
    if update_data.get('is_default') == True:
        db.query(models.Model).filter(
            models.Model.is_default == True,
            models.Model.id != model_id
        ).update({"is_default": False})
    
    for field, value in update_data.items():
        setattr(db_model, field, value)
    
    db.commit()
    db.refresh(db_model)
    return db_model

@app.delete("/api/models/{model_id}")
def delete_model(
    model_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除模型"""
    require_permission(current_user.role, "models", "delete")
    
    db_model = db.query(models.Model).filter(models.Model.id == model_id).first()
    if not db_model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    db.delete(db_model)
    db.commit()
    return {"message": "模型已删除"}

@app.post("/api/models/{model_id}/set-default")
def set_default_model(
    model_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """设置默认模型"""
    require_permission(current_user.role, "models", "update")
    
    db_model = db.query(models.Model).filter(models.Model.id == model_id).first()
    if not db_model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    # 取消其他默认模型
    db.query(models.Model).filter(models.Model.is_default == True).update({"is_default": False})
    
    # 设置当前模型为默认
    db_model.is_default = True
    db.commit()
    return {"message": "已设置为默认模型"}

@app.post("/api/models/{model_id}/test", response_model=schemas.ModelTestResponse)
def test_model(
    model_id: int,
    request: schemas.ModelTestRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """测试模型"""
    require_permission(current_user.role, "models", "read")
    
    db_model = db.query(models.Model).filter(models.Model.id == model_id).first()
    if not db_model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    if db_model.status != 'active':
        raise HTTPException(status_code=400, detail="模型未启用")
    
    try:
        response = call_model_api(db_model, request.prompt)
        return {"response": response}
    except ValueError as e:
        # ValueError 通常是配置错误，返回 400
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        # 其他异常返回 500，并记录详细错误
        import traceback
        error_detail = str(e)
        print(f"❌ 模型调用错误: {error_detail}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"模型调用失败: {error_detail}")

def call_model_api(model: models.Model, prompt: str) -> str:
    """调用模型API"""
    if model.type != 'api':
        raise ValueError("只支持 API 类型的模型，本地模型功能已移除")
    return call_api_model(model, prompt)

def call_api_model(model: models.Model, prompt: str) -> str:
    """调用API模型"""
    import requests
    
    # 验证必要的配置
    if not model.api_key:
        raise ValueError("API Key 未配置")
    
    if not model.model_name:
        raise ValueError("模型标识（model_name）未配置，请填写正确的模型名称")
    
    headers = {
        "Content-Type": "application/json"
    }
    
    # 根据提供商设置API地址和认证
    if model.provider == 'openai':
        api_base = model.api_base or "https://api.openai.com/v1"
        headers["Authorization"] = f"Bearer {model.api_key}"
        url = f"{api_base}/chat/completions"
        payload = {
            "model": model.model_name,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 1000
        }
    elif model.provider == 'deepseek':
        api_base = model.api_base or "https://api.deepseek.com/v1"
        headers["Authorization"] = f"Bearer {model.api_key}"
        url = f"{api_base}/chat/completions"
        payload = {
            "model": model.model_name,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 1000
        }
    elif model.provider == 'qwen':
        api_base = model.api_base or "https://dashscope.aliyuncs.com/api/v1"
        headers["Authorization"] = f"Bearer {model.api_key}"
        url = f"{api_base}/services/aigc/text-generation/generation"
        payload = {
            "model": model.model_name,
            "input": {"messages": [{"role": "user", "content": prompt}]},
            "parameters": {"max_tokens": 1000}
        }
    elif model.provider == 'doubao':
        api_base = model.api_base or "https://ark.cn-beijing.volces.com/api/v3"
        headers["Authorization"] = f"Bearer {model.api_key}"
        url = f"{api_base}/chat/completions"
        payload = {
            "model": model.model_name,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 1000
        }
    else:
        raise ValueError(f"不支持的提供商: {model.provider}")
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=60)
    except requests.exceptions.RequestException as e:
        raise Exception(f"网络请求失败: {str(e)}")
    
    if response.status_code != 200:
        # 尝试解析错误信息
        try:
            error_data = response.json()
            error_msg = error_data.get('error', {})
            if isinstance(error_msg, dict):
                error_detail = error_msg.get('message', error_msg.get('code', '未知错误'))
                error_type = error_msg.get('type', '')
            else:
                error_detail = str(error_msg)
                error_type = ''
            
            # 提供更友好的错误提示
            if "Model Not Exist" in str(error_detail) or "model_not_found" in str(error_detail).lower():
                raise ValueError(
                    f"模型不存在: {model.model_name}\n\n"
                    f"请检查：\n"
                    f"1. 模型名称是否正确（当前: {model.model_name}）\n"
                    f"2. 该模型是否在 {model.provider} 提供商中可用\n"
                    f"3. API Key 是否有权限访问该模型\n\n"
                    f"常见模型名称：\n"
                    f"- OpenAI: gpt-4, gpt-4-turbo-preview, gpt-3.5-turbo\n"
                    f"- DeepSeek: deepseek-chat, deepseek-coder\n"
                    f"- 通义千问: qwen-turbo, qwen-plus, qwen-max\n"
                    f"- 豆包: doubao-pro-32k, doubao-lite-32k"
                )
            elif "invalid_api_key" in str(error_detail).lower() or "unauthorized" in str(error_detail).lower():
                raise ValueError(
                    f"API Key 无效或未授权\n\n"
                    f"请检查：\n"
                    f"1. API Key 是否正确\n"
                    f"2. API Key 是否已过期\n"
                    f"3. API Key 是否有足够的权限"
                )
            else:
                raise Exception(f"API调用失败 ({response.status_code}): {error_detail}")
        except (ValueError, KeyError, json.JSONDecodeError):
            # 如果无法解析JSON，返回原始错误
            raise Exception(f"API调用失败: {response.status_code} - {response.text}")
    
    result = response.json()
    
    # 根据不同提供商的响应格式提取内容
    if model.provider == 'qwen':
        output = result.get('output', {})
        choices = output.get('choices', [])
        if choices:
            message = choices[0].get('message', {})
            content = message.get('content', '无响应')
        else:
            content = result.get('output', {}).get('text', '无响应')
        return content if content else '无响应'
    else:
        choices = result.get('choices', [])
        if choices:
            message = choices[0].get('message', {})
            content = message.get('content', '无响应')
            return content if content else '无响应'
        else:
            # 如果没有choices，尝试其他字段
            return result.get('text', result.get('content', '无响应'))

# 本地模型功能已移除，只支持 API 调用

@app.post("/api/testcases/generate-from-image", response_model=List[schemas.TestCase])
def generate_testcases_from_image(
    project_id: int = Form(...),
    image: UploadFile = File(...),
    model_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """从设计原型图片智能生成测试用例"""
    require_permission(current_user.role, "testcases", "create")
    
    # 检查项目是否存在并加载成员信息
    project = db.query(models.Project).options(
        joinedload(models.Project.members)
    ).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查权限：只有项目成员可以创建测试用例
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建测试用例")
    
    # 验证文件类型
    if not image.content_type or not image.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="请上传图片文件")
    
    # 读取图片内容
    image_content = image.file.read()
    image_base64 = base64.b64encode(image_content).decode('utf-8')
    
    # 调用AI服务分析图片并生成测试用例
    try:
        selected_model = None
        
        # 如果指定了模型ID，使用指定的模型
        if model_id:
            selected_model = db.query(models.Model).filter(
                models.Model.id == model_id,
                models.Model.status == 'active'
            ).first()
            if not selected_model:
                raise HTTPException(status_code=404, detail="指定的模型不存在或未启用")
        else:
            # 如果没有指定模型，检查是否有配置的默认模型
            selected_model = db.query(models.Model).filter(
                models.Model.is_default == True,
                models.Model.status == 'active'
            ).first()
        
        if selected_model:
            # 使用配置的模型
            test_cases_data = analyze_image_with_model(db, selected_model, image_base64, image.content_type)
        else:
            # 使用原来的逻辑（后端直接解析，不使用大模型）
            test_cases_data = analyze_image_and_generate_testcases(image_base64, image.content_type)
    except HTTPException:
        raise
    except Exception as e:
        print(f"AI分析图片失败: {e}")
        raise HTTPException(status_code=500, detail=f"AI分析失败: {str(e)}")
    
    # 创建测试用例
    created_testcases = []
    user_id = current_user.id
    
    for case_data in test_cases_data:
        # 生成用例唯一标识
        case_key = generate_case_key(db, project_id)
        
        # 构建测试用例数据
        testcase_data = {
            'case_key': case_key,
            'project_id': project_id,
            'title': case_data.get('title', '未命名用例'),
            'module': case_data.get('module', ''),
            'precondition': case_data.get('precondition', ''),
            'steps': case_data.get('steps', []),
            'expected_result': case_data.get('expected_result', ''),
            'priority': case_data.get('priority', 'P2'),
            'type': case_data.get('type', 'functional'),
            'status': 'draft',
            'tags': case_data.get('tags', []),
            'created_by': user_id
        }
        
        db_testcase = models.TestCase(**testcase_data)
        db.add(db_testcase)
        created_testcases.append(db_testcase)
    
    db.commit()
    
    # 刷新所有创建的测试用例
    for testcase in created_testcases:
        db.refresh(testcase)
    
    return created_testcases

def analyze_image_with_model(db: Session, model: models.Model, image_base64: str, content_type: str) -> List[Dict[str, Any]]:
    """使用配置的模型分析图片并生成测试用例"""
    # 构建提示词
    prompt = """请分析这张设计原型图片，生成详细的测试用例。

要求：
1. 识别图片中的所有功能模块和交互元素
2. 为每个主要功能生成至少1个测试用例
3. 每个测试用例应包含：
   - title：简洁描述测试场景
   - module：模块名称（可选）
   - precondition：执行测试前需要满足的条件（可选）
   - steps：详细的操作步骤数组，每个步骤必须包含：
     * step_number：步骤序号（从1开始）
     * description：步骤描述
     * expected_result：该步骤的预期结果
   - expected_result：测试用例的整体预期结果
   - priority：优先级，P0-P4（P0最高，P4最低）
   - type：类型，functional（功能）或non-functional（非功能）
   - tags：标签数组（可选）

请以JSON格式返回，格式如下（必须严格遵循）：
[
  {
    "title": "用例标题",
    "module": "模块名称",
    "precondition": "前置条件",
    "steps": [
      {"step_number": 1, "description": "步骤描述", "expected_result": "该步骤预期结果"}
    ],
    "expected_result": "整体预期结果",
    "priority": "P2",
    "type": "functional",
    "tags": ["标签1", "标签2"]
  }
]

注意：steps数组中的每个对象必须包含step_number、description和expected_result三个字段。"""
    
    if model.type == 'api':
        # 调用API模型（需要支持vision的模型）
        if model.provider == 'openai':
            # OpenAI Vision API
            response_text = call_openai_vision_api(model, image_base64, content_type, prompt)
        elif model.provider == 'deepseek':
            # DeepSeek Vision API（使用deepseek-chat-v2或deepseek-v2等视觉模型）
            response_text = call_deepseek_vision_api(model, image_base64, content_type, prompt)
        else:
            # 其他API提供商可能不支持vision，使用文本描述
            response_text = call_model_api(model, f"{prompt}\n\n请根据以上要求生成测试用例。")
    else:
        # 本地模型功能已移除，只支持 API 调用
        raise HTTPException(status_code=400, detail="本地模型功能已移除，请使用 API 类型的模型")
    
    # 解析响应
    try:
        # 尝试从响应中提取JSON
        content = response_text.strip()
        if content.startswith('```json'):
            content = content[7:]
        if content.startswith('```'):
            content = content[3:]
        if content.endswith('```'):
            content = content[:-3]
        content = content.strip()
        
        test_cases = json.loads(content)
        if not isinstance(test_cases, list):
            test_cases = [test_cases]
        return test_cases
    except json.JSONDecodeError as e:
        print(f"JSON解析失败: {e}, 内容: {content[:500]}")
        # 如果解析失败，返回模拟数据
        return generate_mock_testcases()

def call_openai_vision_api(model: models.Model, image_base64: str, content_type: str, prompt: str) -> str:
    """调用OpenAI Vision API"""
    import requests
    
    api_base = model.api_base or "https://api.openai.com/v1"
    headers = {
        "Authorization": f"Bearer {model.api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": model.model_name or "gpt-4-vision-preview",
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{content_type};base64,{image_base64}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 2000
    }
    
    response = requests.post(
        f"{api_base}/chat/completions",
        headers=headers,
        json=payload,
        timeout=60
    )
    
    if response.status_code != 200:
        raise Exception(f"OpenAI API错误: {response.status_code} - {response.text}")
    
    result = response.json()
    return result['choices'][0]['message']['content']

def call_deepseek_vision_api(model: models.Model, image_base64: str, content_type: str, prompt: str) -> str:
    """调用DeepSeek Vision API分析图片"""
    import requests
    
    # DeepSeek 支持的视觉模型名称
    vision_models = ['deepseek-vl', 'deepseek-chat-v2', 'deepseek-v2', 'deepseek-reasoner-v2']
    model_name = model.model_name or "deepseek-vl"
    
    # 如果用户配置的模型不是视觉模型，尝试使用默认的视觉模型
    if model_name not in vision_models:
        print(f"⚠️ 警告: 模型 {model_name} 可能不支持视觉功能，尝试使用 deepseek-vl")
        model_name = "deepseek-vl"
    
    api_base = model.api_base or "https://api.deepseek.com/v1"
    headers = {
        "Authorization": f"Bearer {model.api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{content_type};base64,{image_base64}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 2000
    }
    
    try:
        response = requests.post(
            f"{api_base}/chat/completions",
            headers=headers,
            json=payload,
            timeout=120  # 视觉模型可能需要更长时间
        )
        
        if response.status_code != 200:
            error_text = response.text
            try:
                error_data = response.json()
                error_msg = error_data.get('error', {})
                if isinstance(error_msg, dict):
                    error_detail = error_msg.get('message', error_msg.get('code', '未知错误'))
                    error_type = error_msg.get('type', '')
                else:
                    error_detail = str(error_msg)
                    error_type = ''
            except:
                error_detail = error_text
                error_type = ''
            
            # 提供更友好的错误提示
            if "Model Not Exist" in str(error_detail) or "model_not_found" in str(error_detail).lower():
                raise ValueError(
                    f"DeepSeek 视觉模型不存在: {model_name}\n\n"
                    f"请检查：\n"
                    f"1. 模型名称是否正确（当前: {model_name}）\n"
                    f"2. 是否使用了支持视觉的模型\n\n"
                    f"DeepSeek 支持的视觉模型：\n"
                    f"- deepseek-vl（推荐，多模态视觉模型）\n"
                    f"- deepseek-chat-v2（如果支持视觉）\n"
                    f"- deepseek-v2（如果支持视觉）\n\n"
                    f"注意：如果您的模型配置为普通对话模型（如 deepseek-chat），"
                    f"请改为视觉模型名称，或在模型管理中重新配置。"
                )
            else:
                raise Exception(f"DeepSeek Vision API错误 ({response.status_code}): {error_detail}")
        
        result = response.json()
        return result['choices'][0]['message']['content']
    except requests.exceptions.RequestException as e:
        raise Exception(f"DeepSeek Vision API请求失败: {str(e)}")

def analyze_image_and_generate_testcases(image_base64: str, content_type: str) -> List[Dict[str, Any]]:
    """
    使用AI分析图片并生成测试用例
    
    这里可以使用OpenAI Vision API或其他AI服务
    如果API不可用，则返回模拟数据
    """
    # 尝试使用OpenAI API
    openai_api_key = os.getenv("OPENAI_API_KEY")
    
    if openai_api_key:
        try:
            return analyze_with_openai(image_base64, content_type, openai_api_key)
        except Exception as e:
            print(f"OpenAI API调用失败: {e}")
            # 如果OpenAI失败，使用模拟数据
            return generate_mock_testcases()
    else:
        # 如果没有配置OpenAI API，使用模拟数据
        print("未配置OPENAI_API_KEY，使用模拟数据")
        return generate_mock_testcases()

def analyze_with_openai(image_base64: str, content_type: str, api_key: str) -> List[Dict[str, Any]]:
    """使用OpenAI Vision API分析图片"""
    import requests
    
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    # 构建提示词
    prompt = """请分析这张设计原型图片，生成详细的测试用例。

要求：
1. 识别图片中的所有功能模块和交互元素
2. 为每个主要功能生成至少1个测试用例
3. 每个测试用例应包含：
   - title：简洁描述测试场景
   - module：模块名称（可选）
   - precondition：执行测试前需要满足的条件（可选）
   - steps：详细的操作步骤数组，每个步骤必须包含：
     * step_number：步骤序号（从1开始）
     * description：步骤描述
     * expected_result：该步骤的预期结果
   - expected_result：测试用例的整体预期结果
   - priority：优先级，P0-P4（P0最高，P4最低）
   - type：类型，functional（功能）或non-functional（非功能）
   - tags：标签数组（可选）

请以JSON格式返回，格式如下（必须严格遵循）：
[
  {
    "title": "用例标题",
    "module": "模块名称",
    "precondition": "前置条件",
    "steps": [
      {"step_number": 1, "description": "步骤描述", "expected_result": "该步骤预期结果"}
    ],
    "expected_result": "整体预期结果",
    "priority": "P2",
    "type": "functional",
    "tags": ["标签1", "标签2"]
  }
]

注意：steps数组中的每个对象必须包含step_number、description和expected_result三个字段。"""
    
    payload = {
        "model": "gpt-4-vision-preview",
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{content_type};base64,{image_base64}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 2000
    }
    
    response = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers=headers,
        json=payload,
        timeout=60
    )
    
    if response.status_code != 200:
        raise Exception(f"OpenAI API错误: {response.status_code} - {response.text}")
    
    result = response.json()
    content = result['choices'][0]['message']['content']
    
    # 尝试从响应中提取JSON
    # 移除可能的markdown代码块标记
    content = content.strip()
    if content.startswith('```json'):
        content = content[7:]
    if content.startswith('```'):
        content = content[3:]
    if content.endswith('```'):
        content = content[:-3]
    content = content.strip()
    
    try:
        test_cases = json.loads(content)
        if not isinstance(test_cases, list):
            test_cases = [test_cases]
        return test_cases
    except json.JSONDecodeError as e:
        print(f"JSON解析失败: {e}, 内容: {content[:500]}")
        # 如果解析失败，返回模拟数据
        return generate_mock_testcases()

def generate_mock_testcases() -> List[Dict[str, Any]]:
    """生成模拟测试用例（当AI服务不可用时使用）"""
    return [
        {
            "title": "登录功能测试",
            "module": "用户认证",
            "precondition": "用户已注册账号",
            "steps": [
                {
                    "step_number": 1,
                    "description": "打开登录页面",
                    "expected_result": "显示登录表单，包含用户名/密码输入框和登录按钮"
                },
                {
                    "step_number": 2,
                    "description": "输入正确的用户名和密码",
                    "expected_result": "输入框正常显示输入内容"
                },
                {
                    "step_number": 3,
                    "description": "点击登录按钮",
                    "expected_result": "系统验证成功，跳转到主页"
                }
            ],
            "expected_result": "用户成功登录系统",
            "priority": "P0",
            "type": "functional",
            "tags": ["登录", "认证"]
        },
        {
            "title": "表单提交功能测试",
            "module": "表单管理",
            "precondition": "用户已登录系统",
            "steps": [
                {
                    "step_number": 1,
                    "description": "填写必填字段",
                    "expected_result": "字段验证通过，可以继续填写"
                },
                {
                    "step_number": 2,
                    "description": "点击提交按钮",
                    "expected_result": "表单数据提交成功，显示成功提示"
                }
            ],
            "expected_result": "表单数据成功保存",
            "priority": "P1",
            "type": "functional",
            "tags": ["表单", "提交"]
        }
    ]

# ==================== 认证管理 ====================

@app.post("/api/auth/login", response_model=schemas.LoginResponse)
def login(request: schemas.LoginRequest, http_request: Request, db: Session = Depends(get_db)):
    """用户登录"""
    try:
        # 查找用户
        user = db.query(models.User).filter(models.User.username == request.username).first()
        if not user:
            raise HTTPException(status_code=401, detail="用户名或密码错误")
        
        # 验证密码
        if not verify_password(request.password, user.password):
            raise HTTPException(status_code=401, detail="用户名或密码错误")
        
        # 检查用户状态
        if user.status != 'active':
            raise HTTPException(status_code=403, detail="用户已被禁用")
        
        # 创建访问令牌
        # 从 roles 数组中取第一个角色（或优先使用 'admin'）
        primary_role = 'guest'
        if user.roles:
            if 'admin' in user.roles:
                primary_role = 'admin'
            else:
                primary_role = user.roles[0]
        access_token = create_access_token(
            data={"sub": str(user.id), "username": user.username, "role": primary_role}
        )
        
        # 计算 token 的哈希值（用于在数据库中存储）
        token_hash = hashlib.sha256(access_token.encode()).hexdigest()
        
        # 获取客户端信息
        ip_address = http_request.client.host if http_request.client else None
        user_agent = http_request.headers.get("user-agent", "")
        
        # 计算过期时间（24小时）
        expires_at = datetime.utcnow() + timedelta(minutes=60 * 24)
        
        # 创建会话记录（如果表存在）
        try:
            session = models.UserSession(
                user_id=user.id,
                token_hash=token_hash,
                ip_address=ip_address,
                user_agent=user_agent,
                expires_at=expires_at,
                is_active=True
            )
            db.add(session)
            db.commit()
        except Exception as e:
            # 如果 UserSession 表不存在或创建失败，记录错误但继续登录流程
            print(f"⚠️ 创建会话记录失败: {e}")
            import traceback
            traceback.print_exc()
            db.rollback()
            # 继续执行，不阻止登录
        
        # 返回用户信息（不包含密码）
        user_dict = {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "display_name": user.display_name,
            "avatar_url": user.avatar_url,
            "roles": user.roles or [],
            "status": user.status,
            "created_at": user.created_at,
            "updated_at": user.updated_at
        }
        
        return {
            "access_token": access_token,
            "token_type": "bearer",
            "user": user_dict
        }
    except HTTPException:
        # 重新抛出 HTTP 异常
        raise
    except Exception as e:
        # 捕获其他所有异常，记录并返回 500 错误
        import traceback
        error_msg = f"登录时发生错误: {str(e)}"
        print(f"❌ {error_msg}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=error_msg)

@app.get("/api/auth/me", response_model=schemas.User)
def get_current_user_info(
    http_request: Request,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取当前登录用户信息"""
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 更新会话的最后活动时间（如果提供了 token）
    if http_request:
        auth_header = http_request.headers.get("authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
            token_hash = hashlib.sha256(token.encode()).hexdigest()
            session = db.query(models.UserSession).filter(
                models.UserSession.token_hash == token_hash,
                models.UserSession.user_id == current_user.id,
                models.UserSession.is_active == True
            ).first()
            if session:
                session.last_activity_at = datetime.utcnow()
                db.commit()
    
    return user


@app.post("/api/auth/logout")
def logout(
    http_request: Request,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """用户登出"""
    if http_request:
        auth_header = http_request.headers.get("authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
            token_hash = hashlib.sha256(token.encode()).hexdigest()
            session = db.query(models.UserSession).filter(
                models.UserSession.token_hash == token_hash,
                models.UserSession.user_id == current_user.id,
                models.UserSession.is_active == True
            ).first()
            if session:
                session.is_active = False
                db.commit()
    
    return {"message": "登出成功"}

@app.put("/api/auth/current-project")
def update_current_project(
    project_id: Optional[int] = Query(None),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """更新当前用户的当前项目"""
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 如果提供了项目ID，验证项目是否存在
    if project_id is not None:
        project = db.query(models.Project).filter(models.Project.id == project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="项目不存在")
        user.current_project_id = project_id
    else:
        user.current_project_id = None
    
    db.commit()
    db.refresh(user)
    
    return {"message": "当前项目已更新", "current_project_id": user.current_project_id}

@app.post("/api/auth/change-password")
def change_password(
    user_id: int,
    request: schemas.ChangePasswordRequest,
    db: Session = Depends(get_db)
):
    """修改密码"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 验证旧密码
    if not verify_password(request.old_password, user.password):
        raise HTTPException(status_code=400, detail="旧密码错误")
    
    # 更新密码
    user.password = hash_password(request.new_password)
    db.commit()
    
    return {"message": "密码修改成功"}

# ==================== 权限管理 ====================

@app.get("/api/permissions/{user_id}")
def get_user_role_permissions(user_id: int, db: Session = Depends(get_db)):
    """获取用户权限"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 从 roles 数组中取第一个角色（或优先使用 'admin'）
    primary_role = 'guest'
    if user.roles:
        if 'admin' in user.roles:
            primary_role = 'admin'
        else:
            primary_role = user.roles[0]
    
    return {
        "user_id": user_id,
        "username": user.username,
        "role": primary_role,
        "role_name": ROLE_NAMES.get(primary_role, primary_role),
        "permissions": get_user_permissions(primary_role)
    }

@app.get("/api/roles")
def get_all_roles():
    """获取所有角色列表"""
    return {
        "roles": [
            {"value": role, "label": label} 
            for role, label in ROLE_NAMES.items()
        ]
    }

# ==================== 接口测试平台 ====================

# ==================== 环境管理 ====================

@app.get("/api/environments", response_model=List[schemas.ApiEnvironment])
def get_api_environments(
    project_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db)
):
    """获取所有环境"""
    query = db.query(models.ApiEnvironment).options(joinedload(models.ApiEnvironment.project))
    
    if project_id:
        query = query.filter(models.ApiEnvironment.project_id == project_id)
    
    if keyword:
        query = query.filter(
            models.ApiEnvironment.name.contains(keyword) |
            models.ApiEnvironment.base_url.contains(keyword)
        )
    
    return query.order_by(models.ApiEnvironment.created_at.desc()).offset(skip).limit(limit).all()

@app.get("/api/environments/{environment_id}", response_model=schemas.ApiEnvironment)
def get_api_environment(environment_id: int, db: Session = Depends(get_db)):
    """获取单个环境"""
    environment = db.query(models.ApiEnvironment).options(joinedload(models.ApiEnvironment.project)).filter(models.ApiEnvironment.id == environment_id).first()
    if not environment:
        raise HTTPException(status_code=404, detail="环境不存在")
    return environment

@app.post("/api/environments", response_model=schemas.ApiEnvironment)
def create_api_environment(
    environment: schemas.ApiEnvironmentCreate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建环境"""
    # 检查项目是否存在并加载成员信息
    project = db.query(models.Project).options(
        joinedload(models.Project.members)
    ).filter(models.Project.id == environment.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查权限：只有项目成员可以创建环境
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建环境")
    
    # 检查环境信息（base_url）在该项目下是否已存在（项目+环境信息组合唯一）
    if db.query(models.ApiEnvironment).filter(
        models.ApiEnvironment.project_id == environment.project_id,
        models.ApiEnvironment.base_url == environment.base_url
    ).first():
        raise HTTPException(status_code=400, detail="该项目下环境信息已存在，不能重复")
    
    db_environment = models.ApiEnvironment(**environment.model_dump())
    db.add(db_environment)
    db.commit()
    db.refresh(db_environment)
    return db_environment

@app.put("/api/environments/{environment_id}", response_model=schemas.ApiEnvironment)
def update_api_environment(
    environment_id: int, 
    environment: schemas.ApiEnvironmentUpdate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新环境"""
    db_environment = db.query(models.ApiEnvironment).options(
        joinedload(models.ApiEnvironment.project).joinedload(models.Project.members)
    ).filter(models.ApiEnvironment.id == environment_id).first()
    if not db_environment:
        raise HTTPException(status_code=404, detail="环境不存在")
    
    # 检查权限：只有项目成员可以更新环境
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_environment.project, "更新环境")
    
    update_data = environment.model_dump(exclude_unset=True)
    
    # 如果更新了项目ID，检查项目是否存在并检查新项目的权限
    if 'project_id' in update_data:
        new_project = db.query(models.Project).options(
            joinedload(models.Project.members)
        ).filter(models.Project.id == update_data['project_id']).first()
        if not new_project:
            raise HTTPException(status_code=404, detail="项目不存在")
        # 检查新项目的权限
        check_project_member_permission(user, new_project, "更新环境到该项目")
        
    # 如果更新了环境信息（base_url）或项目ID，检查组合是否已存在（排除当前环境）
    if 'base_url' in update_data or 'project_id' in update_data:
        check_project_id = update_data.get('project_id', db_environment.project_id)
        check_base_url = update_data.get('base_url', db_environment.base_url)
        if db.query(models.ApiEnvironment).filter(
            models.ApiEnvironment.project_id == check_project_id,
            models.ApiEnvironment.base_url == check_base_url,
            models.ApiEnvironment.id != environment_id
        ).first():
            raise HTTPException(status_code=400, detail="该项目下环境信息已存在，不能重复")
    
    for key, value in update_data.items():
        setattr(db_environment, key, value)
    
    db.commit()
    db.refresh(db_environment)
    return db_environment

@app.delete("/api/environments/{environment_id}")
def delete_api_environment(
    environment_id: int, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除环境"""
    db_environment = db.query(models.ApiEnvironment).options(
        joinedload(models.ApiEnvironment.project).joinedload(models.Project.members)
    ).filter(models.ApiEnvironment.id == environment_id).first()
    if not db_environment:
        raise HTTPException(status_code=404, detail="环境不存在")
    
    # 检查权限：只有项目成员可以删除环境
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_environment.project, "删除环境")
    
    db.delete(db_environment)
    db.commit()
    return {"message": "环境已删除"}

# ==================== 代码扫描管理 ====================

@app.get("/api/code-scans")
def get_code_scans(
    project_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    result: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取代码扫描任务列表（服务端分页）"""
    query = db.query(models.CodeScan).options(joinedload(models.CodeScan.project))

    if project_id:
        query = query.filter(models.CodeScan.project_id == project_id)
    if keyword:
        query = query.filter(
            models.CodeScan.project_name.contains(keyword) |
            models.CodeScan.branch.contains(keyword) |
            models.CodeScan.scan_path.contains(keyword)
        )
    if result:
        query = query.filter(models.CodeScan.result == result)

    total = query.count()
    items = query.order_by(models.CodeScan.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}

@app.get("/api/code-scans/{scan_id}", response_model=schemas.CodeScan)
def get_code_scan(scan_id: int, db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user)):
    """获取单个扫描任务"""
    scan = db.query(models.CodeScan).options(joinedload(models.CodeScan.project)).filter(models.CodeScan.id == scan_id).first()
    if not scan:
        raise HTTPException(status_code=404, detail="扫描任务不存在")
    return scan

@app.post("/api/code-scans", response_model=schemas.CodeScan)
def create_code_scan(
    scan: schemas.CodeScanCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建代码扫描任务"""
    # 检查项目是否存在
    project = db.query(models.Project).filter(models.Project.id == scan.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查权限
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建代码扫描任务")
    
    db_scan = models.CodeScan(**scan.model_dump())
    db.add(db_scan)
    db.commit()
    db.refresh(db_scan)
    return db_scan

@app.put("/api/code-scans/{scan_id}", response_model=schemas.CodeScan)
def update_code_scan(
    scan_id: int,
    scan: schemas.CodeScanUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新代码扫描任务"""
    db_scan = db.query(models.CodeScan).options(
        joinedload(models.CodeScan.project).joinedload(models.Project.members)
    ).filter(models.CodeScan.id == scan_id).first()
    if not db_scan:
        raise HTTPException(status_code=404, detail="扫描任务不存在")
    
    # 检查权限
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_scan.project, "更新代码扫描任务")
    
    update_data = scan.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_scan, key, value)
    
    db.commit()
    db.refresh(db_scan)
    return db_scan

@app.delete("/api/code-scans/{scan_id}")
def delete_code_scan(
    scan_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除代码扫描任务"""
    db_scan = db.query(models.CodeScan).options(
        joinedload(models.CodeScan.project).joinedload(models.Project.members)
    ).filter(models.CodeScan.id == scan_id).first()
    if not db_scan:
        raise HTTPException(status_code=404, detail="扫描任务不存在")
    
    # 检查权限
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_scan.project, "删除代码扫描任务")
    
    db.delete(db_scan)
    db.commit()
    return {"message": "扫描任务已删除"}

@app.post("/api/code-scans/{scan_id}/execute")
def execute_code_scan(
    scan_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """调用 SonarQube 质量门 API，成功则根据返回结果更新扫描状态（不执行 mvn）"""
    db_scan = db.query(models.CodeScan).options(
        joinedload(models.CodeScan.project).joinedload(models.Project.members)
    ).filter(models.CodeScan.id == scan_id).first()
    if not db_scan:
        raise HTTPException(status_code=404, detail="扫描任务不存在")
    
    # 检查权限
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_scan.project, "执行代码扫描")
    
    # 检查是否配置了 Sonar Host
    if not db_scan.sonar_host:
        raise HTTPException(status_code=400, detail="请先配置 Sonar Host")
    
    # 构建 Sonar 项目 Key
    project_key = db_scan.sonar_project_key or f"{db_scan.project_name}:{db_scan.branch}"
    
    sonar_host = db_scan.sonar_host.rstrip('/')
    
    # 仅调用 Sonar 质量门 API，调用成功即更新状态（不在此接口执行 mvn，避免长时间等待）
    api_url = f"{sonar_host}/api/qualitygates/project_status"
    
    # 指标名称映射
    metric_names = {
        'new_coverage': '新代码覆盖率',
        'coverage': '代码覆盖率',
        'new_duplicated_lines_density': '新代码重复率',
        'duplicated_lines_density': '代码重复率',
        'new_reliability_rating': '新代码可靠性',
        'reliability_rating': '可靠性等级',
        'new_security_rating': '新代码安全性',
        'security_rating': '安全性等级',
        'new_maintainability_rating': '新代码可维护性',
        'sqale_rating': '可维护性等级',
        'new_bugs': '新增Bug数',
        'bugs': 'Bug数',
        'new_vulnerabilities': '新增漏洞数',
        'vulnerabilities': '漏洞数',
        'new_code_smells': '新增代码异味',
        'code_smells': '代码异味数',
        'new_security_hotspots_reviewed': '新代码安全热点审查率',
        'security_hotspots_reviewed': '安全热点审查率'
    }
    
    # 比较运算符映射
    comparators = {
        'LT': '<',
        'GT': '>',
        'EQ': '=',
        'NE': '≠',
        'LE': '≤',
        'GE': '≥'
    }
    
    try:
        headers = {}
        if db_scan.sonar_login:
            # 使用 Token 认证
            auth_string = f"{db_scan.sonar_login}:"
            auth_bytes = base64.b64encode(auth_string.encode()).decode()
            headers['Authorization'] = f'Basic {auth_bytes}'
        
        response = requests.get(
            api_url,
            params={'projectKey': project_key},
            headers=headers,
            timeout=30
        )
        
        if response.status_code == 200:
            data = response.json()
            project_status = data.get('projectStatus', {})
            status = project_status.get('status', '')
            conditions = project_status.get('conditions', [])
            
            # 扫描状态以调用 Sonar 质量门 API 成功后的结果为准（与 mvn 是否执行无关）
            db_scan.result = 'passed' if status == 'OK' else 'failed'
            db_scan.scan_time = datetime.now()
            
            # 解析不通过的条件
            error_details = []
            conditions_out = []
            if status != 'OK' and conditions:
                for cond in conditions:
                    if cond.get('status') == 'ERROR':
                        metric_key = cond.get('metricKey', '')
                        metric_name = metric_names.get(metric_key, metric_key)
                        comparator = comparators.get(cond.get('comparator', ''), cond.get('comparator', ''))
                        threshold = cond.get('errorThreshold', '')
                        actual = cond.get('actualValue', '')
                        error_details.append(f"{metric_name}: {actual} (要求{comparator}{threshold})")
                        conditions_out.append({
                            "metric_key": metric_key,
                            "metric_name": metric_name,
                            "actual_value": actual,
                            "error_threshold": threshold,
                            "comparator": comparator,
                        })
            
            # 存储错误信息，截断过长内容；无具体条件时也给出可读说明并写日志
            if error_details:
                error_msg = '; '.join(error_details)
                if len(error_msg) > 500:
                    error_msg = error_msg[:497] + '...'
                db_scan.error_message = error_msg
            else:
                if status != 'OK':
                    db_scan.error_message = "质量门未通过（未返回具体条件）。请点「详情」→ 右上角「打开 Sonar」进入项目页面，在 Sonar 里查看质量门与未通过项"
                else:
                    db_scan.error_message = None
            
            # 扫描不通过时写一条日志（含原始 conditions），便于排查
            if status != 'OK':
                import logging
                _log = logging.getLogger(__name__)
                _log.info(
                    "code_scan execute: 质量门未通过 scan_id=%s project_key=%s sonar_status=%s error_message=%s",
                    scan_id, project_key, status, db_scan.error_message or "(无详情)"
                )
                if not error_details and conditions:
                    _log.info("code_scan execute: raw conditions from SonarQube: %s", conditions)
            
            # 记录本次扫描结果历史（状态与扫描结果一致：质量门 OK 则为 completed）
            history = models.CodeScanResult(
                scan_id=db_scan.id,
                status='completed' if db_scan.result == 'passed' else 'failed',
                error_message=db_scan.error_message,
            )
            db.add(history)
            
            db.commit()
            db.refresh(db_scan)
            return {
                "message": "扫描完成",
                "result": db_scan.result,
                "scan_time": db_scan.scan_time.isoformat() if db_scan.scan_time else None,
                "sonar_status": status,
                "error_message": db_scan.error_message,
                "conditions": conditions_out,
            }
        elif response.status_code == 404:
            # 项目在 SonarQube 中不存在
            error_msg = f"SonarQube 中找不到项目: {project_key}"
            db_scan.result = 'failed'
            db_scan.error_message = error_msg
            db_scan.scan_time = datetime.now()
            db.add(models.CodeScanResult(
                scan_id=db_scan.id,
                status='failed',
                error_message=error_msg,
            ))
            db.commit()
            raise HTTPException(
                status_code=404, 
                detail=f"{error_msg}，请先在 SonarQube 中执行扫描"
            )
        elif response.status_code == 401:
            error_msg = "SonarQube 认证失败，请检查 Token"
            db_scan.result = 'failed'
            db_scan.error_message = error_msg
            db_scan.scan_time = datetime.now()
            db.add(models.CodeScanResult(
                scan_id=db_scan.id,
                status='failed',
                error_message=error_msg,
            ))
            db.commit()
            raise HTTPException(status_code=401, detail=error_msg)
        else:
            error_msg = f"SonarQube API 调用失败: {response.status_code}"
            db_scan.result = 'failed'
            db_scan.error_message = error_msg
            db_scan.scan_time = datetime.now()
            db.add(models.CodeScanResult(
                scan_id=db_scan.id,
                status='failed',
                error_message=error_msg,
            ))
            db.commit()
            raise HTTPException(
                status_code=500, 
                detail=f"{error_msg} - {response.text}"
            )
    except requests.exceptions.Timeout:
        error_msg = "SonarQube API 请求超时"
        db_scan.result = 'failed'
        db_scan.error_message = error_msg
        db_scan.scan_time = datetime.now()
        db.add(models.CodeScanResult(
            scan_id=db_scan.id,
            status='failed',
            error_message=error_msg,
        ))
        db.commit()
        raise HTTPException(status_code=504, detail=error_msg)
    except requests.exceptions.ConnectionError:
        error_msg = f"无法连接到 SonarQube: {sonar_host}"
        db_scan.result = 'failed'
        db_scan.error_message = error_msg
        db_scan.scan_time = datetime.now()
        db.add(models.CodeScanResult(
            scan_id=db_scan.id,
            status='failed',
            error_message=error_msg,
        ))
        db.commit()
        raise HTTPException(status_code=503, detail=error_msg)
    except HTTPException:
        raise
    except Exception as e:
        error_msg = f"查询 SonarQube 状态失败: {str(e)}"
        db_scan.result = 'failed'
        db_scan.error_message = error_msg
        db_scan.scan_time = datetime.now()
        db.add(models.CodeScanResult(
            scan_id=db_scan.id,
            status='failed',
            error_message=error_msg,
        ))
        db.commit()
        raise HTTPException(status_code=500, detail=error_msg)

@app.get("/api/code-scans/{scan_id}/results")
def get_code_scan_results(
    scan_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取代码扫描历史记录"""
    db_scan = db.query(models.CodeScan).filter(models.CodeScan.id == scan_id).first()
    if not db_scan:
        raise HTTPException(status_code=404, detail="扫描任务不存在")

    histories = (
        db.query(models.CodeScanResult)
        .filter(models.CodeScanResult.scan_id == scan_id)
        .order_by(models.CodeScanResult.created_at.desc())
        .all()
    )

    return [
        {
            "id": h.id,
            "status": h.status,
            "error_message": h.error_message,
            "created_at": h.created_at.isoformat() if h.created_at else None,
            "updated_at": h.updated_at.isoformat() if h.updated_at else None,
        }
        for h in histories
    ]

@app.get("/api/code-scans/{scan_id}/status")
def get_code_scan_status(
    scan_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取扫描任务的最新状态（从 SonarQube 实时查询）"""
    db_scan = db.query(models.CodeScan).filter(models.CodeScan.id == scan_id).first()
    if not db_scan:
        raise HTTPException(status_code=404, detail="扫描任务不存在")
    
    if not db_scan.sonar_host:
        return {
            "result": db_scan.result,
            "scan_time": db_scan.scan_time.isoformat() if db_scan.scan_time else None,
            "message": "未配置 Sonar Host"
        }
    
    project_key = db_scan.sonar_project_key or f"{db_scan.project_name}:{db_scan.branch}"
    sonar_host = db_scan.sonar_host.rstrip('/')
    api_url = f"{sonar_host}/api/qualitygates/project_status"
    
    try:
        headers = {}
        if db_scan.sonar_login:
            auth_string = f"{db_scan.sonar_login}:"
            auth_bytes = base64.b64encode(auth_string.encode()).decode()
            headers['Authorization'] = f'Basic {auth_bytes}'
        
        response = requests.get(
            api_url,
            params={'projectKey': project_key},
            headers=headers,
            timeout=10
        )
        
        if response.status_code == 200:
            data = response.json()
            status = data.get('projectStatus', {}).get('status', '')
            new_result = 'passed' if status == 'OK' else 'failed'
            if db_scan.result != new_result:
                db_scan.result = new_result
                db_scan.scan_time = datetime.now()
                db.commit()
            return {
                "result": new_result,
                "scan_time": db_scan.scan_time.isoformat() if db_scan.scan_time else None,
                "sonar_status": status
            }
        elif response.status_code == 404:
            return {
                "result": None,
                "scan_time": None,
                "message": f"SonarQube 中找不到项目: {project_key}"
            }
        else:
            return {
                "result": db_scan.result,
                "scan_time": db_scan.scan_time.isoformat() if db_scan.scan_time else None,
                "message": f"SonarQube API 调用失败: {response.status_code}"
            }
    except Exception as e:
        return {
            "result": db_scan.result,
            "scan_time": db_scan.scan_time.isoformat() if db_scan.scan_time else None,
            "message": f"查询失败: {str(e)}"
        }

# ==================== 接口端点管理 ====================

@app.get("/api/api-endpoints")
def get_api_endpoints(
    project_id: Optional[int] = None,
    method: Optional[str] = None,
    tag: Optional[str] = None,
    keyword: Optional[str] = None,
    is_favorite: Optional[bool] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=10000),
    db: Session = Depends(get_db)
):
    """获取接口列表（服务端分页）"""
    query = db.query(models.ApiEndpoint).options(joinedload(models.ApiEndpoint.project))

    if project_id:
        query = query.filter(models.ApiEndpoint.project_id == project_id)
    if method:
        query = query.filter(models.ApiEndpoint.method == method.upper())
    if tag:
        query = query.filter(models.ApiEndpoint.tags.contains([tag]))
    if keyword:
        query = query.filter(
            models.ApiEndpoint.name.contains(keyword) |
            models.ApiEndpoint.path.contains(keyword) |
            models.ApiEndpoint.description.contains(keyword)
        )
    if is_favorite is not None:
        query = query.filter(models.ApiEndpoint.is_favorite == is_favorite)

    total = query.count()
    items = query.order_by(models.ApiEndpoint.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}

# 注意：record 路由必须定义在 {endpoint_id} 路由之前，否则会被路径参数路由匹配
@app.post("/api/api-endpoints/record")
def record_api_from_url(
    request_data: schemas.RecordApiRequest,
    db: Session = Depends(get_db)
):
    """从URL录制接口（爬取所有子页面的REST接口）"""
    import requests
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin, urlparse
    import re
    from collections import deque
    
    # 获取环境和项目
    environment = db.query(models.ApiEnvironment).filter(models.ApiEnvironment.id == request_data.environment_id).first()
    if not environment:
        raise HTTPException(status_code=404, detail="环境不存在")
    
    project = db.query(models.Project).filter(models.Project.id == request_data.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    start_url = request_data.start_url.rstrip('/')
    max_depth = request_data.max_depth
    
    # 验证URL
    try:
        parsed_url = urlparse(start_url)
        if not parsed_url.scheme or not parsed_url.netloc:
            raise HTTPException(status_code=400, detail="无效的URL格式")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"URL解析失败: {str(e)}")
    
    # 用于存储已访问的URL和发现的接口
    visited_urls = set()
    discovered_apis = set()  # 使用set去重，存储 (method, path) 元组
    url_queue = deque([(start_url, 0)])  # (url, depth)
    
    try:
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,application/json,*/*;q=0.8'
        })
        
        # 如果需要登录，先执行登录
        if request_data.login_url and request_data.login_username:
            try:
                login_data = request_data.login_data or {}
                if not login_data:
                    # 如果没有提供自定义登录数据，使用默认格式
                    login_data = {
                        "username": request_data.login_username,
                        "password": request_data.login_password or ""
                    }
                else:
                    # 如果提供了自定义登录数据，替换用户名和密码字段
                    login_data = dict(login_data)  # 复制一份
                    # 尝试替换常见的用户名和密码字段
                    for key in login_data.keys():
                        if 'username' in key.lower() or 'user' in key.lower() or 'account' in key.lower():
                            login_data[key] = request_data.login_username
                        if 'password' in key.lower() or 'pwd' in key.lower() or 'pass' in key.lower():
                            login_data[key] = request_data.login_password or ""
                
                login_response = session.post(
                    request_data.login_url,
                    json=login_data,
                    timeout=10,
                    allow_redirects=True
                )
                
                if login_response.status_code in [200, 201]:
                    # 登录成功，session会自动保存cookie
                    print(f"✅ 登录成功: {request_data.login_url}")
                else:
                    print(f"⚠️ 登录可能失败，状态码: {login_response.status_code}")
            except Exception as login_error:
                print(f"⚠️ 登录失败，将继续尝试爬取: {str(login_error)}")
        
        # ========== 爬取页面提取 API ==========
        base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
        print(f"\n🔍 开始爬取页面: {start_url}")
        print(f"   域名: {base_url}, 最大深度: {max_depth}")
        
        # 定义更广泛的 API 路径模式
        api_path_patterns = [
            r'/api/',
            r'/v\d+/',
            r'/rest/',
            r'/service/',
            r'/services/',
            r'/graphql',
        ]
        
        # 定义检查路径是否像 API 的函数
        def is_api_path(path):
            if not path:
                return False
            for pattern in api_path_patterns:
                if re.search(pattern, path, re.IGNORECASE):
                    return True
            # 也检查以 /api 开头的路径
            if path.startswith('/api'):
                return True
            return False
        
        # 用于存储发现的 baseURL 配置
        discovered_base_urls = set(['/api'])  # 默认包含 /api
        
        # 常见的资源路径（不带 /api 前缀）
        common_resource_paths = [
            'users', 'bugs', 'projects', 'tasks', 'requirements', 'testcases',
            'sprints', 'statistics', 'comments', 'files', 'uploads', 'images',
            'auth', 'login', 'logout', 'register', 'environments', 'flows',
            'endpoints', 'models', 'permissions', 'roles', 'settings', 'config'
        ]
        
        # 爬取所有子页面
        while url_queue:
            current_url, depth = url_queue.popleft()
            
            if current_url in visited_urls or depth > max_depth:
                continue
            
            visited_urls.add(current_url)
            
            try:
                # 请求页面
                response = session.get(current_url, timeout=10, allow_redirects=True)
                if response.status_code != 200:
                    continue
                
                content_type = response.headers.get('Content-Type', '').lower()
                
                # 如果是JSON响应，尝试从中提取API路径（但跳过swagger/openapi文档格式）
                if 'application/json' in content_type:
                    try:
                        json_data = response.json()
                        # 检查是否是swagger/openapi文档格式，如果是则跳过（swagger应该在同步接口功能中处理）
                        if isinstance(json_data, dict):
                            # 检查是否包含swagger/openapi特征字段
                            if 'swagger' in json_data or 'openapi' in json_data or 'info' in json_data:
                                # 这是swagger文档，跳过
                                continue
                        
                        # 递归搜索JSON中的URL路径（用于从API响应中提取其他API路径）
                        def extract_paths_from_json(obj, parent_key=''):
                            if isinstance(obj, dict):
                                for key, value in obj.items():
                                    # 跳过swagger相关字段
                                    if key.lower() in ['swagger', 'openapi', 'info', 'paths', 'components']:
                                        continue
                                    if key.lower() in ['path', 'url', 'endpoint', 'api', 'href', 'link']:
                                        if isinstance(value, str) and value.startswith('/api/'):
                                            discovered_apis.add(('GET', value.split('?')[0].split('#')[0]))
                                    extract_paths_from_json(value, key)
                            elif isinstance(obj, list):
                                for item in obj:
                                    extract_paths_from_json(item, parent_key)
                            elif isinstance(obj, str):
                                # 检查字符串中是否包含API路径
                                api_matches = re.finditer(r'["\'](/api/[^"\']+)["\']', obj)
                                for match in api_matches:
                                    path = match.group(1).split('?')[0].split('#')[0]
                                    discovered_apis.add(('GET', path))
                        extract_paths_from_json(json_data)
                    except:
                        pass
                    continue
                
                if 'text/html' not in content_type:
                    continue
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # 提取页面中的链接（用于继续爬取）
                if depth < max_depth:
                    for link in soup.find_all('a', href=True):
                        href = link['href']
                        absolute_url = urljoin(current_url, href)
                        parsed = urlparse(absolute_url)
                        
                        # 只爬取同一域名的链接
                        if parsed.netloc == parsed_url.netloc and absolute_url not in visited_urls:
                            url_queue.append((absolute_url, depth + 1))
                
                # 从HTML中提取API路径
                # 1. 从form的action属性
                for form in soup.find_all('form', action=True):
                    action = form.get('action')
                    if action:
                        method = form.get('method', 'POST').upper()
                        full_url = urljoin(current_url, action)
                        parsed_api = urlparse(full_url)
                        # 提取所有看起来像API的路径（包含/api/或/v1/等）
                        if '/api/' in parsed_api.path or '/v1/' in parsed_api.path or '/v2/' in parsed_api.path or parsed_api.path.startswith('/api'):
                            path = parsed_api.path
                            discovered_apis.add((method, path))
                
                # 2. 从所有href中提取API路径
                for link in soup.find_all(['a', 'link'], href=True):
                    href = link.get('href')
                    if href and (href.startswith('/api/') or '/api/' in href or '/v1/' in href or '/v2/' in href):
                        full_url = urljoin(current_url, href)
                        parsed_api = urlparse(full_url)
                        if parsed_api.path:
                            path = parsed_api.path
                            discovered_apis.add(('GET', path))
                
                # 从 JavaScript代码中提取API调用
                # 获取所有script标签的内容（包括内联和外部）
                script_contents = []
                script_urls_to_fetch = []
                                
                for script in soup.find_all('script'):
                    if script.string:
                        script_contents.append(script.string)
                    src = script.get('src')
                    if src:
                        script_urls_to_fetch.append(src)
                                
                # 查找 Vite 特有的模块引用（如 type="module" src="/src/main.ts"）
                module_scripts = soup.find_all('script', {'type': 'module'})
                for script in module_scripts:
                    src = script.get('src')
                    if src:
                        script_urls_to_fetch.append(src)
                                
                # 获取所有外部 JS 文件
                for src in set(script_urls_to_fetch):
                    try:
                        script_url = urljoin(current_url, src)
                        if urlparse(script_url).netloc == parsed_url.netloc:
                            print(f"    📄 加载脚本: {src}")
                            script_response = session.get(script_url, timeout=10)
                            if script_response.status_code == 200:
                                script_contents.append(script_response.text)
                                                
                                # 如果是 Vite 开发服务器，尝试加载导入的模块
                                # 匹配 import ... from './xxx' 或 import './xxx'
                                import_matches = re.finditer(
                                    r'(?:import|from)\s+["\'](\.?/[^"\']+\.(?:ts|js|vue))["\']',
                                    script_response.text,
                                    re.IGNORECASE
                                )
                                for imp in import_matches:
                                    imp_path = imp.group(1)
                                    imp_url = urljoin(script_url, imp_path)
                                    try:
                                        imp_response = session.get(imp_url, timeout=5)
                                        if imp_response.status_code == 200:
                                            script_contents.append(imp_response.text)
                                    except:
                                        pass
                    except Exception as e:
                        print(f"    ⚠️ 加载脚本失败: {src} - {str(e)}")
                
                # 也直接从HTML文本中搜索API模式
                html_text = response.text
                script_contents.append(html_text)
                
                for script_content in script_contents:
                    if not script_content:
                        continue
                    
                    # 提取 baseURL 配置
                    base_url_matches = re.finditer(
                        r'(?:baseURL|baseUrl|base_url)\s*[:=]\s*["\'`]([^"\'`]+)["\'`]',
                        script_content,
                        re.IGNORECASE
                    )
                    for match in base_url_matches:
                        base = match.group(1)
                        if base.startswith('/'):
                            discovered_base_urls.add(base.rstrip('/'))
                            print(f"    📌 发现 baseURL: {base}")
                    
                    # 匹配短路径（如 '/bugs', '/users' 等）并拼接 baseURL
                    short_path_patterns = re.finditer(
                        r'(?:request|axios|http|api|fetch)\s*\.\s*(get|post|put|delete|patch)\s*(?:<[^>]*>)?\s*\(\s*["\'`](/[a-zA-Z][a-zA-Z0-9_/-]*)["\'`]',
                        script_content,
                        re.IGNORECASE
                    )
                    for match in short_path_patterns:
                        method = match.group(1).upper()
                        short_path = match.group(2).split('?')[0].split('#')[0]
                        # 移除模板变量
                        short_path = re.sub(r'\$\{[^}]+\}', '', short_path)
                        if short_path and not short_path.startswith('/api'):
                            # 拼接 baseURL
                            for base in discovered_base_urls:
                                full_path = base + short_path
                                discovered_apis.add((method, full_path))
                                print(f"    ✅ 发现 API: {method} {full_path}")
                    
                    # 匹配常见的API调用模式
                    # 1. fetch(url, {method: 'GET', ...}) - 更宽松的匹配
                    fetch_patterns = re.finditer(
                        r'fetch\s*\(\s*["\'`]([^"\'"`]+)["\'`]\s*(?:,\s*\{[^}]*method\s*:\s*["\'`]?(\w+)["\'`]?\s*[^}]*\})?',
                        script_content,
                        re.IGNORECASE | re.DOTALL
                    )
                    for match in fetch_patterns:
                        api_url = match.group(1)
                        method = (match.group(2) or 'GET').upper()
                        full_url = urljoin(current_url, api_url)
                        parsed_api = urlparse(full_url)
                        
                        # 使用 is_api_path 函数检查
                        if is_api_path(parsed_api.path):
                            path = parsed_api.path.split('?')[0].split('#')[0]
                            discovered_apis.add((method, path))
                            print(f"    ✅ 发现 API: {method} {path}")
                    
                    # 2. axios.get/post/put/delete(url) - 更宽松的匹配
                    axios_patterns = re.finditer(
                        r'axios\.?(get|post|put|delete|patch|request)?\s*\(\s*["\'`]([^"\'"`]+)["\'`]',
                        script_content,
                        re.IGNORECASE
                    )
                    for match in axios_patterns:
                        method = (match.group(1) or 'GET').upper()
                        api_url = match.group(2)
                        full_url = urljoin(current_url, api_url)
                        parsed_api = urlparse(full_url)
                        
                        if is_api_path(parsed_api.path):
                            path = parsed_api.path.split('?')[0].split('#')[0]
                            discovered_apis.add((method, path))
                            print(f"    ✅ 发现 API: {method} {path}")
                    
                    # 3. $.ajax({url: ..., type: 'GET', ...}) - 更宽松的匹配
                    ajax_patterns = re.finditer(
                        r'\$\.ajax\s*\(\s*\{[^}]*url\s*:\s*["\'`]([^"\'`]+)["\'`][^}]*(?:type|method)\s*:\s*["\'`]?(\w+)["\'`]?\s*[^}]*\}',
                        script_content,
                        re.IGNORECASE | re.DOTALL
                    )
                    for match in ajax_patterns:
                        api_url = match.group(1)
                        method = match.group(2).upper()
                        full_url = urljoin(current_url, api_url)
                        parsed_api = urlparse(full_url)
                        
                        if is_api_path(parsed_api.path):
                            path = parsed_api.path.split('?')[0].split('#')[0]
                            discovered_apis.add((method, path))
                            print(f"    ✅ 发现 API: {method} {path}")
                    
                    # 4. XMLHttpRequest - 更宽松的匹配
                    xhr_patterns = re.finditer(
                        r'\.open\s*\(\s*["\'`]?(\w+)["\'`]?\s*,\s*["\'`]([^"\'`]+)["\'`]',
                        script_content,
                        re.IGNORECASE
                    )
                    for match in xhr_patterns:
                        method = match.group(1).upper()
                        api_url = match.group(2)
                        full_url = urljoin(current_url, api_url)
                        parsed_api = urlparse(full_url)
                        
                        if is_api_path(parsed_api.path):
                            path = parsed_api.path.split('?')[0].split('#')[0]
                            discovered_apis.add((method, path))
                            print(f"    ✅ 发现 API: {method} {path}")
                    
                    # 5. 直接URL模式 - 更宽松的匹配（包括 /api/, /v1/, /v2/ 等）
                    # 匹配引号内的API路径
                    direct_api_patterns = re.finditer(
                        r'["\'](/api/[^"\']+|/v\d+/[^"\']+)["\']',
                        script_content,
                        re.IGNORECASE
                    )
                    for match in direct_api_patterns:
                        path = match.group(1)
                        # 清理路径
                        if '?' in path:
                            path = path.split('?')[0]
                        if '#' in path:
                            path = path.split('#')[0]
                        # 尝试从上下文推断方法（默认GET）
                        method = 'GET'
                        # 检查前面是否有method提示
                        context_start = max(0, match.start() - 100)
                        context = script_content[context_start:match.start()].lower()
                        if 'post' in context or 'create' in context or 'add' in context or 'save' in context:
                            method = 'POST'
                        elif 'put' in context or 'update' in context or 'edit' in context or 'modify' in context:
                            method = 'PUT'
                        elif 'delete' in context or 'remove' in context or 'del' in context:
                            method = 'DELETE'
                        elif 'patch' in context:
                            method = 'PATCH'
                        discovered_apis.add((method, path))
                    
                    # 5.1. 匹配模板字符串中的API路径（如 `/api/${id}` 或 `/api/users`）
                    template_string_patterns = re.finditer(
                        r'[`"\'](/api/[^`"\']+)[`"\']',
                        script_content,
                        re.IGNORECASE
                    )
                    for match in template_string_patterns:
                        path = match.group(1)
                        # 移除模板变量部分，只保留基础路径
                        path = re.sub(r'\$\{[^}]+\}', '', path)
                        if path.startswith('/api/'):
                            if '?' in path:
                                path = path.split('?')[0]
                            if '#' in path:
                                path = path.split('#')[0]
                            discovered_apis.add(('GET', path))
                    
                    # 5.2. 匹配函数调用中的API路径（如 request('/api/users')）
                    function_call_patterns = re.finditer(
                        r'(?:request|http|api|fetch|get|post|put|delete)\s*\(\s*["\'](/api/[^"\']+)["\']',
                        script_content,
                        re.IGNORECASE
                    )
                    for match in function_call_patterns:
                        path = match.group(1)
                        if '?' in path:
                            path = path.split('?')[0]
                        if '#' in path:
                            path = path.split('#')[0]
                        discovered_apis.add(('GET', path))
                    
                    # 6. 匹配 baseURL + path 模式（如 baseURL + '/api/users'）
                    base_url_patterns = re.finditer(
                        r'(?:baseURL|baseUrl|base_url)\s*\+\s*["\']([^"\']+)["\']',
                        script_content,
                        re.IGNORECASE
                    )
                    for match in base_url_patterns:
                        path = match.group(1)
                        if path.startswith('/api/') or path.startswith('/v'):
                            method = 'GET'
                            context_start = max(0, match.start() - 100)
                            context = script_content[context_start:match.start()].lower()
                            if 'post' in context or 'create' in context:
                                method = 'POST'
                            elif 'put' in context or 'update' in context:
                                method = 'PUT'
                            elif 'delete' in context or 'remove' in context:
                                method = 'DELETE'
                            discovered_apis.add((method, path))
                
                # 从HTML中的data属性提取API信息
                for elem in soup.find_all(attrs={'data-api': True}):
                    api_url = elem.get('data-api')
                    method = elem.get('data-method', 'GET').upper()
                    full_url = urljoin(current_url, api_url)
                    parsed_api = urlparse(full_url)
                    
                    if is_api_path(parsed_api.path):
                        path = parsed_api.path.split('?')[0].split('#')[0]
                        discovered_apis.add((method, path))
                        print(f"    ✅ 发现 API: {method} {path}")
                
                # 从所有可能的API路径模式中提取（更宽松的匹配）
                # 匹配类似 /api/xxx, /v1/xxx, /v2/xxx 等模式
                all_api_patterns = re.finditer(
                    r'["\'`](/api/[^"\'`]+|/v\d+/[^"\'`]+)["\'`]',
                    response.text,
                    re.IGNORECASE
                )
                for match in all_api_patterns:
                    path = match.group(1)
                    # 清理路径（移除查询参数和锚点）
                    path = path.split('?')[0].split('#')[0]
                    # 移除模板变量
                    path = re.sub(r'\$\{[^}]+\}', '', path)
                    if path and is_api_path(path):
                        discovered_apis.add(('GET', path))
                
                # 尝试从页面中提取所有可能的API端点（更激进的匹配）
                # 匹配任何包含 /api/ 的URL模式，即使不在引号中
                aggressive_patterns = re.finditer(
                    r'(?:url|path|endpoint|api|href|src)\s*[:=]\s*["\'`]?([^"\'`\s]+/api/[^"\'`\s\)]+)["\'`]?',
                    response.text,
                    re.IGNORECASE
                )
                for match in aggressive_patterns:
                    potential_path = match.group(1)
                    # 解析URL，提取路径部分
                    try:
                        parsed = urlparse(potential_path)
                        if parsed.path and is_api_path(parsed.path):
                            path = parsed.path.split('?')[0].split('#')[0]
                            discovered_apis.add(('GET', path))
                    except:
                        pass
                        
            except Exception as e:
                print(f"⚠️ 爬取URL失败: {current_url}, 错误: {str(e)}")
                continue
                
        print(f"\n📊 爬取完成！访问了 {len(visited_urls)} 个页面，发现 {len(discovered_apis)} 个 API 接口")
        
        # 基于已发现的API路径，尝试推断和访问相关的API端点
        # 例如：如果发现了 /api/chain-plans，尝试访问 /api/chain-plans/{id} 等
        inferred_apis = set()
        for method, path in list(discovered_apis):
            # 提取路径的基础部分（如 /api/chain-plans）
            path_parts = [p for p in path.split('/') if p]
            if len(path_parts) >= 2 and path_parts[0] == 'api':
                base_path = '/' + '/'.join(path_parts[:2])  # 如 /api/chain-plans
                # 尝试访问基础路径，看是否能获取列表或文档
                try:
                    test_url = urljoin(start_url.rstrip('/'), base_path)
                    test_response = session.get(test_url, timeout=3, allow_redirects=True)
                    if test_response.status_code == 200:
                        content_type = test_response.headers.get('Content-Type', '').lower()
                        if 'application/json' in content_type:
                            try:
                                json_data = test_response.json()
                                # 如果是列表，尝试推断其他可能的端点
                                if isinstance(json_data, list) and len(json_data) > 0:
                                    # 尝试访问第一个元素的详情页
                                    first_item = json_data[0]
                                    if isinstance(first_item, dict):
                                        item_id = first_item.get('id') or first_item.get('_id') or first_item.get('key')
                                        if item_id:
                                            detail_path = f"{base_path}/{item_id}"
                                            inferred_apis.add(('GET', detail_path))
                                            inferred_apis.add(('PUT', detail_path))
                                            inferred_apis.add(('DELETE', detail_path))
                            except:
                                pass
                except:
                    pass
        
        discovered_apis.update(inferred_apis)
        
        # 保存发现的接口到数据库
        imported_count = 0
        
        for method, path in discovered_apis:
            # 检查接口是否已存在（同一项目、方法、路径）
            existing = db.query(models.ApiEndpoint).filter(
                models.ApiEndpoint.project_id == request_data.project_id,
                models.ApiEndpoint.method == method,
                models.ApiEndpoint.path == path
            ).first()
            
            if existing:
                continue
            
            # 从路径提取接口名称
            path_parts = [p for p in path.split('/') if p]
            if path_parts:
                name = path_parts[-1].replace('-', ' ').replace('_', ' ').title()
            else:
                name = f"{method} {path}"
            
            # 创建接口记录
            endpoint = models.ApiEndpoint(
                project_id=request_data.project_id,
                name=name,
                path=path,
                method=method,
                description=f"从 {start_url} 录制发现的接口"
            )
            db.add(endpoint)
            imported_count += 1
        
        db.commit()
        
        return {
            "message": f"录制完成，爬取了 {len(visited_urls)} 个页面，发现 {len(discovered_apis)} 个接口，导入 {imported_count} 个新接口",
            "discovered_count": len(discovered_apis),
            "imported_count": imported_count
        }
        
    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=500, detail=f"网络请求失败: {str(e)}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"录制失败: {str(e)}")

@app.post("/api/api-endpoints", response_model=schemas.ApiEndpoint)
def create_api_endpoint(
    endpoint: schemas.ApiEndpointCreate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建接口"""
    # 检查项目是否存在并加载成员信息
    project = db.query(models.Project).options(
        joinedload(models.Project.members)
    ).filter(models.Project.id == endpoint.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查权限：只有项目成员可以创建接口
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建接口")
    
    db_endpoint = models.ApiEndpoint(**endpoint.model_dump())
    db.add(db_endpoint)
    db.commit()
    db.refresh(db_endpoint)
    return db_endpoint

@app.get("/api/api-endpoints/{endpoint_id}", response_model=schemas.ApiEndpoint)
def get_api_endpoint(endpoint_id: int, db: Session = Depends(get_db)):
    """获取单个接口"""
    endpoint = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.id == endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail="接口不存在")
    return endpoint

@app.put("/api/api-endpoints/{endpoint_id}", response_model=schemas.ApiEndpoint)
def update_api_endpoint(
    endpoint_id: int, 
    endpoint: schemas.ApiEndpointUpdate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新接口"""
    db_endpoint = db.query(models.ApiEndpoint).options(
        joinedload(models.ApiEndpoint.project).joinedload(models.Project.members)
    ).filter(models.ApiEndpoint.id == endpoint_id).first()
    if not db_endpoint:
        raise HTTPException(status_code=404, detail="接口不存在")
    
    # 检查权限：只有项目成员可以更新接口
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_endpoint.project, "更新接口")
    
    # 如果更新了项目ID，检查项目是否存在并检查新项目的权限
    update_data = endpoint.model_dump(exclude_unset=True)
    if 'project_id' in update_data:
        new_project = db.query(models.Project).options(
            joinedload(models.Project.members)
        ).filter(models.Project.id == update_data['project_id']).first()
        if not new_project:
            raise HTTPException(status_code=404, detail="项目不存在")
        # 检查新项目的权限
        check_project_member_permission(user, new_project, "更新接口到该项目")
    
    for key, value in update_data.items():
        setattr(db_endpoint, key, value)
    
    db.commit()
    db.refresh(db_endpoint)
    return db_endpoint

@app.delete("/api/api-endpoints/{endpoint_id}")
def delete_api_endpoint(
    endpoint_id: int, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除接口"""
    endpoint = db.query(models.ApiEndpoint).options(
        joinedload(models.ApiEndpoint.project).joinedload(models.Project.members)
    ).filter(models.ApiEndpoint.id == endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail="接口不存在")
    
    # 检查权限：只有项目成员可以删除接口
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, endpoint.project, "删除接口")
    
    # 检查依赖：是否有测试任务使用该接口
    task_items = db.query(models.TestTaskItem).filter(
        models.TestTaskItem.item_type == 'api',
        models.TestTaskItem.item_id == endpoint_id
    ).all()
    if task_items:
        task_names = []
        for item in task_items:
            task = db.query(models.TestTask).filter(models.TestTask.id == item.task_id).first()
            if task:
                task_names.append(task.name)
        raise HTTPException(
            status_code=400,
            detail=f"无法删除接口：该接口正在被测试任务使用（{', '.join(task_names)}），请先从测试任务中移除该接口"
        )
    
    # 检查依赖：是否有流程使用该接口
    flows = db.query(models.ApiTestFlow).all()
    dependent_flows = []
    for flow in flows:
        if flow.steps:
            steps = flow.steps if isinstance(flow.steps, list) else []
            for step in steps:
                if isinstance(step, dict) and step.get('endpoint_id') == endpoint_id:
                    dependent_flows.append(flow.name)
                    break
    
    if dependent_flows:
        raise HTTPException(
            status_code=400,
            detail=f"无法删除接口：该接口正在被测试流程使用（{', '.join(dependent_flows)}），请先从测试流程中移除该接口"
        )
    
    db.delete(endpoint)
    db.commit()
    return {"message": "接口已删除"}

@app.put("/api/api-endpoints/{endpoint_id}/favorite")
def toggle_favorite_endpoint(
    endpoint_id: int, 
    is_favorite: bool = Query(...), 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """收藏/取消收藏接口"""
    endpoint = db.query(models.ApiEndpoint).options(
        joinedload(models.ApiEndpoint.project).joinedload(models.Project.members)
    ).filter(models.ApiEndpoint.id == endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail="接口不存在")
    
    # 检查权限：只有项目成员可以收藏接口
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, endpoint.project, "收藏接口")
    
    endpoint.is_favorite = is_favorite
    db.commit()
    db.refresh(endpoint)
    return {"message": "操作成功", "is_favorite": endpoint.is_favorite}

# ==================== 测试数据管理 ====================

@app.get("/api/api-test-data", response_model=List[schemas.ApiTestData])
def get_api_test_data_list(endpoint_id: Optional[int] = None, db: Session = Depends(get_db)):
    """获取测试数据列表；如指定 endpoint 且暂无数据，则自动生成一条默认测试数据"""
    query = db.query(models.ApiTestData)
    if endpoint_id:
        query = query.filter(models.ApiTestData.endpoint_id == endpoint_id)
        data = query.order_by(models.ApiTestData.created_at.desc()).all()
        if not data:
            # 如果该接口存在但还没有测试数据，则自动生成一条默认数据
            endpoint = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.id == endpoint_id).first()
            if endpoint:
                default_data = models.ApiTestData(
                    endpoint_id=endpoint_id,
                    name=endpoint.path,  # 使用接口路径作为名称
                    expected_status=200,
                )
                db.add(default_data)
                db.commit()
                db.refresh(default_data)
                return [default_data]
        return data
    return query.order_by(models.ApiTestData.created_at.desc()).all()

@app.get("/api/api-test-data/{test_data_id}", response_model=schemas.ApiTestData)
def get_api_test_data(test_data_id: int, db: Session = Depends(get_db)):
    """获取单个测试数据"""
    test_data = db.query(models.ApiTestData).filter(models.ApiTestData.id == test_data_id).first()
    if not test_data:
        raise HTTPException(status_code=404, detail="测试数据不存在")
    return test_data

@app.post("/api/api-test-data", response_model=schemas.ApiTestData)
def create_api_test_data(
    test_data: schemas.ApiTestDataCreate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建测试数据"""
    # 获取接口并加载项目成员信息
    endpoint = db.query(models.ApiEndpoint).options(
        joinedload(models.ApiEndpoint.project).joinedload(models.Project.members)
    ).filter(models.ApiEndpoint.id == test_data.endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail="接口不存在")
    
    # 检查权限：只有项目成员可以创建测试数据
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, endpoint.project, "创建测试数据")
    
    db_test_data = models.ApiTestData(**test_data.model_dump())
    db.add(db_test_data)
    db.commit()
    db.refresh(db_test_data)
    return db_test_data

@app.put("/api/api-test-data/{test_data_id}", response_model=schemas.ApiTestData)
def update_api_test_data(
    test_data_id: int, 
    test_data: schemas.ApiTestDataUpdate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新测试数据"""
    db_test_data = db.query(models.ApiTestData).options(
        joinedload(models.ApiTestData.endpoint).joinedload(models.ApiEndpoint.project).joinedload(models.Project.members)
    ).filter(models.ApiTestData.id == test_data_id).first()
    if not db_test_data:
        raise HTTPException(status_code=404, detail="测试数据不存在")
    
    # 检查权限：只有项目成员可以更新测试数据
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_test_data.endpoint.project, "更新测试数据")
    
    for key, value in test_data.model_dump(exclude_unset=True).items():
        setattr(db_test_data, key, value)
    
    db.commit()
    db.refresh(db_test_data)
    return db_test_data

@app.delete("/api/api-test-data/{test_data_id}")
def delete_api_test_data(
    test_data_id: int, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除测试数据"""
    test_data = db.query(models.ApiTestData).options(
        joinedload(models.ApiTestData.endpoint).joinedload(models.ApiEndpoint.project).joinedload(models.Project.members)
    ).filter(models.ApiTestData.id == test_data_id).first()
    if not test_data:
        raise HTTPException(status_code=404, detail="测试数据不存在")
    
    # 检查权限：只有项目成员可以删除测试数据
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, test_data.endpoint.project, "删除测试数据")
    
    db.delete(test_data)
    db.commit()
    return {"message": "测试数据已删除"}

# ==================== 接口执行 ====================

@app.post("/api/api-endpoints/{endpoint_id}/execute", response_model=schemas.ApiExecutionRecord)
def execute_api_endpoint(
    endpoint_id: int, 
    request: schemas.ApiExecuteRequest, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """执行接口"""
    import requests
    import time
    
    # 获取接口并加载项目成员信息
    endpoint = db.query(models.ApiEndpoint).options(
        joinedload(models.ApiEndpoint.project).joinedload(models.Project.members)
    ).filter(models.ApiEndpoint.id == endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail="接口不存在")
    
    # 检查权限：只有项目成员可以执行接口
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, endpoint.project, "执行接口")
    
    # 获取环境
    environment = db.query(models.ApiEnvironment).filter(models.ApiEnvironment.id == request.environment_id).first()
    if not environment:
        raise HTTPException(status_code=404, detail="环境不存在")
    
    # 获取测试数据（如果提供）
    test_data = None
    if request.test_data_id:
        test_data = db.query(models.ApiTestData).filter(models.ApiTestData.id == request.test_data_id).first()
    
    # 初始化变量上下文（用于模板替换）
    context: Dict[str, Any] = {}
    if request.global_variables:
        context.update(request.global_variables)
        print(f"🔍 执行单个接口 - 从请求加载变量: {request.global_variables}")
    else:
        print(f"🔍 执行单个接口 - 请求中没有 global_variables")
    
    # 构建请求URL
    base_url = environment.base_url.rstrip('/')
    path = endpoint.path.lstrip('/')
    full_url = f"{base_url}/{path}"
    
    # 处理路径参数（支持变量替换）
    path_params = None
    if test_data and test_data.path_params:
        path_params = _render_template(test_data.path_params, context)
    elif request.path_params:
        path_params = _render_template(request.path_params, context)
    
    if path_params:
        for key, value in path_params.items():
            full_url = full_url.replace(f"{{{key}}}", str(value))
    
    # 构建请求头（支持变量替换）
    headers = {}
    if environment.headers:
        # 环境配置的 headers 也需要渲染
        env_headers = _render_template(environment.headers, context)
        headers.update(env_headers)
        print(f"🔍 环境配置 Headers (渲染后): {env_headers}")
    if test_data and test_data.headers:
        test_headers = _render_template(test_data.headers, context)
        headers.update(test_headers)
        print(f"🔍 测试数据 Headers (渲染后): {test_headers}")
    if request.headers:
        print(f"🔍 请求 Headers (渲染前): {request.headers}")
        request_headers = _render_template(request.headers, context)
        headers.update(request_headers)
        print(f"🔍 请求 Headers (渲染后): {request_headers}")
    
    # 构建查询参数（支持变量替换）
    query_params = {}
    if test_data and test_data.query_params:
        test_query = _render_template(test_data.query_params, context)
        query_params.update(test_query)
    if request.query_params:
        request_query = _render_template(request.query_params, context)
        query_params.update(request_query)
    
    # 构建请求体（支持变量替换）
    body = None
    if test_data and test_data.body:
        body = _render_template(test_data.body, context)
    if request.body:
        body = _render_template(request.body, context)
    
    # 调试日志：输出最终的 headers 和 context
    print(f"🔍 执行单个接口 - 最终变量上下文: {context}")
    print(f"🔍 执行单个接口 - 最终 Headers: {headers}")
    
    # 执行请求
    start_time = time.time()
    success = False
    response_status = None
    response_headers = None
    response_body = None
    error_message = None
    
    try:
        if query_params:
            response = requests.request(
                method=endpoint.method.upper(),
                url=full_url,
                headers=headers,
                json=body if body else None,
                params=query_params,
                timeout=30
            )
        else:
            response = requests.request(
                method=endpoint.method.upper(),
                url=full_url,
                headers=headers,
                json=body if body else None,
                timeout=30
            )
        
        response_time = int((time.time() - start_time) * 1000)
        response_status = response.status_code
        response_headers = dict(response.headers)
        
        # 尝试解析响应体
        try:
            response_body = response.text
            if len(response_body) > 10000:  # 限制响应体长度
                response_body = response_body[:10000] + "... (truncated)"
        except:
            response_body = str(response.content)
        
        # 解析响应体为JSON（用于断言验证）
        json_body = None
        try:
            json_body = response.json()
        except:
            json_body = None
        
        # 获取断言列表（优先使用请求中的断言，否则使用测试数据中的断言）
        assertions_list = request.assertions if request.assertions else (test_data.assertions if test_data and test_data.assertions else [])
        
        # 验证断言
        success = True
        assertion_errors = []
        
        if assertions_list:
            for assertion in assertions_list:
                assertion_type = assertion.get('type')
                operator = assertion.get('operator')
                target = assertion.get('target')
                expected = assertion.get('expected')
                
                if assertion_type == 'status_code':
                    # 状态码断言
                    actual_value = response_status
                    if not _check_assertion(actual_value, operator, expected):
                        success = False
                        assertion_errors.append(f"状态码断言失败: 期望 {expected}，实际 {actual_value}")
                
                elif assertion_type == 'json_path':
                    # JSON路径断言
                    if json_body and target:
                        # 先渲染 expected 值（处理变量替换，如 STR($CreateValue)）
                        if expected:
                            try:
                                # 对于单个接口执行，使用 request.global_variables 作为 context
                                exec_context = {}
                                if request.global_variables:
                                    exec_context.update(request.global_variables)
                                expected_rendered = _render_template(expected, exec_context)
                                # 如果渲染后是字符串且以引号开头结尾，去掉引号
                                if isinstance(expected_rendered, str) and expected_rendered.startswith('"') and expected_rendered.endswith('"'):
                                    expected_rendered = expected_rendered[1:-1]
                                expected = expected_rendered
                            except Exception as e:
                                print(f"🔍 渲染 expected 值失败: {e}")
                        
                        actual_value = _extract_json_path(json_body, target)
                        print(f"🔍 JSON路径断言: path={target}, expected={expected}, actual={actual_value}")
                        
                        if not _check_assertion(actual_value, operator, expected):
                            success = False
                            assertion_errors.append(f"JSON路径断言失败: {target} 期望 {expected}，实际 {_format_value_for_display(actual_value)}")
                    else:
                        success = False
                        assertion_errors.append(f"JSON路径断言失败: 无法提取路径 {target}")
                
                elif assertion_type == 'response_time':
                    # 响应时间断言
                    actual_value = response_time
                    if not _check_assertion(actual_value, operator, expected):
                        success = False
                        assertion_errors.append(f"响应时间断言失败: 期望 {expected}ms，实际 {actual_value}ms")
                
                elif assertion_type == 'contains':
                    # 包含断言
                    if json_body:
                        body_str = json.dumps(json_body) if isinstance(json_body, dict) else str(json_body)
                        if expected and expected not in body_str:
                            success = False
                            assertion_errors.append(f"包含断言失败: 响应体中不包含 {expected}")
                    else:
                        if expected and expected not in str(response_body):
                            success = False
                            assertion_errors.append(f"包含断言失败: 响应体中不包含 {expected}")
        else:
            # 如果没有断言，使用默认逻辑：状态码在200-299之间
            success = 200 <= response_status < 300
        
        # 如果有期望状态码，检查是否匹配
        expected_status = None
        if test_data:
            expected_status = test_data.expected_status
        if expected_status and response_status != expected_status:
            success = False
            error_message = f"期望状态码 {expected_status}，实际 {response_status}"
        
        # 如果有断言错误，设置错误消息
        if assertion_errors:
            error_message = "\n".join(assertion_errors)
    
    except Exception as e:
        response_time = int((time.time() - start_time) * 1000)
        error_message = str(e)
        success = False
    
    # 保存执行记录
    record = models.ApiExecutionRecord(
        endpoint_id=endpoint_id,
        test_data_id=request.test_data_id,
        environment_id=request.environment_id,
        request_url=full_url,
        request_method=endpoint.method.upper(),
        request_headers=headers,
        request_query_params=query_params if query_params else None,
        request_path_params=path_params if path_params else None,
        request_body=body,
        response_status=response_status,
        response_headers=response_headers,
        response_body=response_body,
        response_time=response_time,
        success=success,
        error_message=error_message
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    
    return record


# ==================== 辅助函数 ====================
def _extract_json_path(data: Any, path: str) -> Any:
    """从JSON数据中提取路径值（支持简单的点号分隔路径和数组索引，如 data.code 或 data.list[0].name）"""
    if not path or not data:
        return None
    
    try:
        # 处理路径，支持数组索引，如 data.list[0].name
        # 将路径分割，但保留数组索引部分
        import re
        # 使用正则表达式分割路径，保留数组索引
        parts = re.split(r'\.(?!\d)', path)  # 按点分割，但不分割数字后的点
        
        value = data
        for part in parts:
            if not part:
                continue
            
            # 检查是否包含数组索引，如 list[0]
            if '[' in part and ']' in part:
                # 提取键名和索引
                key_match = re.match(r'^([^\[]+)\[(\d+)\]$', part)
                if key_match:
                    key_name = key_match.group(1)
                    index = int(key_match.group(2))
                    
                    # 从字典中获取列表
                    if isinstance(value, dict) and key_name in value:
                        value = value[key_name]
                        # 从列表中获取指定索引的元素
                        if isinstance(value, list) and 0 <= index < len(value):
                            value = value[index]
                        else:
                            return None
                    else:
                        return None
                else:
                    return None
            else:
                # 普通键访问
                if isinstance(value, dict) and part in value:
                    value = value[part]
                elif isinstance(value, list) and len(value) > 0:
                    # 如果是列表，尝试访问第一个元素
                    value = value[0]
                    if isinstance(value, dict) and part in value:
                        value = value[part]
                    else:
                        return None
                else:
                    return None
        
        return value
    except Exception as e:
        print(f"🔍 _extract_json_path 错误: path={path}, error={e}")
        return None

def _format_value_for_display(value: Any) -> str:
    """格式化值用于显示，将 None 显示为 null（与 JSON 保持一致）"""
    if value is None:
        return 'null'
    return str(value)

def _check_assertion(actual: Any, operator: str, expected: Any) -> bool:
    """检查断言是否通过 - 直接比较值，不做额外转换"""
    # 如果expected为None或空字符串，返回False
    if expected is None or expected == '':
        return False
    
    # 处理用户输入 "null" 字符串的情况：JSON 中的 null 在 Python 中被解析为 None
    # 如果用户输入字符串 "null"（不区分大小写），将其转换为 None 进行比较
    if isinstance(expected, str):
        expected_trimmed = expected.strip()
        if expected_trimmed.lower() == 'null':
            expected = None
        # 处理 JSON 字符串的情况：如果 expected 是 JSON 字符串（如 "[]", "{}", "[1,2,3]" 等），尝试解析
        elif expected_trimmed.startswith('[') or expected_trimmed.startswith('{'):
            try:
                import json
                expected_parsed = json.loads(expected_trimmed)
                expected = expected_parsed
            except (json.JSONDecodeError, ValueError):
                # 解析失败，保持原值
                pass
    
    # 对于数值比较，尝试转换为相同类型
    # 如果一个是数字，另一个是字符串形式的数字，转换为数字比较
    if operator in ('eq', 'ne'):
        # 尝试将字符串转换为数字进行比较
        try:
            if isinstance(actual, (int, float)) and isinstance(expected, str):
                # actual 是数字，expected 是字符串，尝试转换 expected
                if '.' in expected:
                    expected_num = float(expected)
                else:
                    expected_num = int(expected)
                if operator == 'eq':
                    return actual == expected_num
                else:
                    return actual != expected_num
            elif isinstance(expected, (int, float)) and isinstance(actual, str):
                # expected 是数字，actual 是字符串，尝试转换 actual
                if '.' in actual:
                    actual_num = float(actual)
                else:
                    actual_num = int(actual)
                if operator == 'eq':
                    return actual_num == expected
                else:
                    return actual_num != expected
        except (ValueError, TypeError):
            # 转换失败，使用原始值比较
            pass
    
    # 直接比较值，使用 Python 的 == 操作符
    if operator == 'eq':
        return actual == expected
    elif operator == 'ne':
        return actual != expected
    elif operator == 'gt':
        try:
            return float(actual) > float(expected)
        except:
            return False
    elif operator == 'gte':
        try:
            return float(actual) >= float(expected)
        except:
            return False
    elif operator == 'lt':
        try:
            return float(actual) < float(expected)
        except:
            return False
    elif operator == 'lte':
        try:
            return float(actual) <= float(expected)
        except:
            return False
    elif operator == 'contains':
        return str(expected) in str(actual)
    elif operator == 'not_contains':
        return str(expected) not in str(actual)
    else:
        return False

# ==================== 接口流程测试 ====================
def _render_template(value: Any, context: Dict[str, Any]) -> Any:
    """使用 $ 关键字语法渲染数据，支持 $API[N].path 和 $var 语法，同时兼容 {{ var }} 语法
    支持 NUM() 和 STR() 格式：
    - NUM($API[N].path) 或 NUM($var) - 返回数字值（不加引号）
    - STR($API[N].path) 或 STR($var) - 返回字符串值（自动添加引号）
    """
    if isinstance(value, str):
        def extract_and_format(expr: str):
            """提取值并返回 (值, 是否找到)"""
            extracted_value = None
            found = False
            
            # 支持 $API[N].path 语法，例如 $API[2].code 或 $API[2].data.code
            api_match = re.match(r'\$?API\[(\d+)\]\.(.+)', expr)
            if api_match:
                api_index = int(api_match.group(1))
                path = api_match.group(2)
                # 从 context 中获取 API[N] 对象
                api_key = f"API[{api_index}]"
                api_data = context.get(api_key)
                if api_data is not None:
                    # 从 api_data 中提取路径值
                    extracted_value = _extract_value(api_data, path)
                    found = extracted_value is not None
            # 支持 $var 语法（局部变量或全局变量）
            elif expr.startswith('$'):
                var_name = expr[1:].strip()
                extracted_value = context.get(var_name)
                found = extracted_value is not None
            else:
                # 传统方式：直接使用变量名（兼容旧语法）
                extracted_value = context.get(expr)
                found = extracted_value is not None
            
            return extracted_value, found
        
        def format_value(val: Any, return_string: bool = False):
            """根据值的类型格式化输出
            return_string: 如果为 True，返回带引号的字符串（用于 JSON 字符串替换）；如果为 False，返回原始值（用于字典/列表）
            """
            if return_string:
                # 返回字符串格式（用于在 JSON 字符串中替换）
                if val is None:
                    return 'null'
                # 如果是数字、布尔值，直接转换为字符串（不需要引号）
                if isinstance(val, (int, float, bool)):
                    return str(val)
                # 如果是字符串，需要加引号（但需要转义内部的双引号和反斜杠）
                elif isinstance(val, str):
                    # 转义字符串中的双引号和反斜杠
                    escaped = val.replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n').replace('\r', '\\r')
                    return f'"{escaped}"'
                # 如果是对象或数组，序列化为 JSON 字符串
                else:
                    import json
                    json_str = json.dumps(val, ensure_ascii=False)
                    return json_str
            else:
                # 返回原始值（用于字典/列表）
                return val
        
        # 处理 NUM() 和 STR() 格式（优先级最高）
        def repl_num_api(match: re.Match) -> str:
            """处理 NUM($API[N].path) 格式，返回数字值（不加引号）"""
            api_index = int(match.group(1))
            path = match.group(2)
            api_key = f"API[{api_index}]"
            api_data = context.get(api_key)
            if api_data is not None:
                extracted_value = _extract_value(api_data, path)
                if extracted_value is not None:
                    try:
                        if isinstance(extracted_value, bool):
                            num_value = 1 if extracted_value else 0
                        elif isinstance(extracted_value, (int, float)):
                            num_value = extracted_value
                        else:
                            num_value = float(extracted_value)
                        return str(int(num_value) if isinstance(num_value, float) and num_value.is_integer() else num_value)
                    except (ValueError, TypeError):
                        return '0'
            return '0'
        
        def repl_num_var(match: re.Match) -> str:
            """处理 NUM($var) 格式，返回数字值（不加引号）"""
            var_name = match.group(1)
            extracted_value = context.get(var_name)
            if extracted_value is not None:
                try:
                    if isinstance(extracted_value, bool):
                        num_value = 1 if extracted_value else 0
                    elif isinstance(extracted_value, (int, float)):
                        num_value = extracted_value
                    else:
                        num_value = float(extracted_value)
                    return str(int(num_value) if isinstance(num_value, float) and num_value.is_integer() else num_value)
                except (ValueError, TypeError):
                    return '0'
            return '0'
        
        def repl_str_api(match: re.Match) -> str:
            """处理 STR($API[N].path) 格式，返回字符串值（自动添加引号）"""
            api_index = int(match.group(1))
            path = match.group(2)
            api_key = f"API[{api_index}]"
            api_data = context.get(api_key)
            if api_data is not None:
                extracted_value = _extract_value(api_data, path)
                if extracted_value is not None:
                    import json
                    return json.dumps(str(extracted_value))
            return '""'
        
        def repl_str_var(match: re.Match) -> str:
            """处理 STR($var) 格式，返回字符串值（自动添加引号）"""
            var_name = match.group(1)
            extracted_value = context.get(var_name)
            if extracted_value is not None:
                import json
                return json.dumps(str(extracted_value))
            return '""'
        
        # 先处理 NUM($API[N].path) 和 NUM($var) 格式
        value = re.sub(r'NUM\(\$?API\[(\d+)\]\.([a-zA-Z0-9_.]+)\)', repl_num_api, value)
        value = re.sub(r'NUM\(\$([a-zA-Z_][a-zA-Z0-9_]*)\)', repl_num_var, value)
        
        # 再处理 STR($API[N].path) 和 STR($var) 格式
        value = re.sub(r'STR\(\$?API\[(\d+)\]\.([a-zA-Z0-9_.]+)\)', repl_str_api, value)
        value = re.sub(r'STR\(\$([a-zA-Z_][a-zA-Z0-9_]*)\)', repl_str_var, value)
        
        # 处理 $API[N].path 语法（不在引号内）
        def repl_api(match: re.Match) -> str:
            full_expr = match.group(0)  # 例如 $API[1].data
            expr = full_expr.lstrip('$')  # 移除开头的 $
            extracted_value, found = extract_and_format(expr)
            if found:
                return format_value(extracted_value, return_string=True)
            return full_expr  # 如果找不到，返回原始表达式
        
        # 处理 $var 语法（变量名）
        def repl_var(match: re.Match) -> str:
            full_expr = match.group(0)  # 例如 $Tenant
            var_name = match.group(1)  # 例如 Tenant
            extracted_value = context.get(var_name)
            if extracted_value is not None:
                return format_value(extracted_value, return_string=True)
            return full_expr  # 如果找不到，返回原始表达式
        
        # 先处理 $API[N].path 语法（更具体的模式优先）
        # 匹配 $API[数字].路径，路径可以包含字母、数字、点、下划线
        value = re.sub(r'\$API\[(\d+)\]\.([a-zA-Z0-9_.]+)', repl_api, value)
        
        # 再处理 $变量名 语法（但不能是 $API[ 开头，因为已经处理过了）
        # 匹配 $ 后跟变量名（字母、数字、下划线），但不能是 API[ 开头
        value = re.sub(r'\$([a-zA-Z_][a-zA-Z0-9_]*)(?![\[\.])', repl_var, value)
        
        # 兼容旧的 {{ var }} 语法
        def repl_old(match: re.Match) -> str:
            expr = match.group(1).strip()
            extracted_value, found = extract_and_format(expr)
            if found:
                return format_value(extracted_value, return_string=True)
            return match.group(0)  # 如果找不到，返回原始占位符
        value = re.sub(r"\{\{\s*([^{}]+)\s*\}\}", repl_old, value)
        
        return value
    if isinstance(value, dict):
        # 对于字典，递归处理每个值
        result = {}
        for k, v in value.items():
            # 如果值是字符串且包含模板语法
            if isinstance(v, str) and ('$' in v or '{{' in v or 'NUM(' in v or 'STR(' in v):
                # 进行模板渲染
                rendered = _render_template(v, context)
                # 如果渲染结果是字符串且是带引号的 JSON 字符串，解析它
                if isinstance(rendered, str) and len(rendered) >= 2 and rendered.startswith('"') and rendered.endswith('"'):
                    import json
                    try:
                        # 解析 JSON 字符串，去掉外层引号，得到实际值
                        result[k] = json.loads(rendered)
                    except:
                        # 解析失败，保持原样
                        result[k] = rendered
                else:
                    # 不是带引号的 JSON 字符串，直接使用（可能是数字、布尔值等）
                    result[k] = rendered
            else:
                # 递归处理
                result[k] = _render_template(v, context)
        return result
    if isinstance(value, list):
        return [_render_template(item, context) for item in value]
    return value


def _extract_value(data: Any, path: str) -> Any:
    """按照点路径提取值，支持列表下标"""
    current = data
    for part in path.split('.'):
        if isinstance(current, dict):
            current = current.get(part)
        elif isinstance(current, list):
            try:
                idx = int(part)
                current = current[idx]
            except (ValueError, IndexError):
                return None
        else:
            return None
    return current


@app.get("/api/api-flows")
def list_api_flows(
    project_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    is_favorite: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=10000),
    db: Session = Depends(get_db)
):
    """获取接口流程列表（服务端分页）"""
    query = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project),
        joinedload(models.ApiTestFlow.environment)
    ).order_by(models.ApiTestFlow.updated_at.desc())
    if project_id:
        query = query.filter(models.ApiTestFlow.project_id == project_id)
    if keyword:
        keyword_lower = keyword.lower()
        query = query.filter(func.lower(models.ApiTestFlow.name).contains(keyword_lower))
    if is_favorite is not None:
        query = query.filter(models.ApiTestFlow.is_favorite == is_favorite)
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}


@app.get("/api/api-flows/{flow_id}", response_model=schemas.ApiTestFlow)
def get_api_flow(flow_id: int, db: Session = Depends(get_db)):
    flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project),
        joinedload(models.ApiTestFlow.environment)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    return flow


@app.put("/api/api-flows/{flow_id}/favorite")
def toggle_favorite_flow(
    flow_id: int, 
    is_favorite: bool = Query(...), 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """收藏/取消收藏流程"""
    flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project).joinedload(models.Project.members)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    # 检查权限：只有项目成员可以收藏流程
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, flow.project, "收藏流程")
    
    flow.is_favorite = is_favorite
    db.commit()
    db.refresh(flow)
    return {"message": "操作成功", "is_favorite": flow.is_favorite}


@app.post("/api/api-flows", response_model=schemas.ApiTestFlow)
def create_api_flow(
    flow: schemas.ApiTestFlowCreate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    # 校验项目存在并加载成员信息
    project = db.query(models.Project).options(
        joinedload(models.Project.members)
    ).filter(models.Project.id == flow.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查权限：只有项目成员可以创建流程
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, project, "创建流程")
    
    # 检查同一项目下流程名称是否已存在
    existing_flow = db.query(models.ApiTestFlow).filter(
        models.ApiTestFlow.project_id == flow.project_id,
        models.ApiTestFlow.name == flow.name
    ).first()
    if existing_flow:
        raise HTTPException(status_code=400, detail="该项目下已存在同名流程")
    
    db_flow = models.ApiTestFlow(**flow.model_dump())
    db.add(db_flow)
    db.commit()
    db.refresh(db_flow)
    # 重新加载以包含关联的项目信息
    db_flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project),
        joinedload(models.ApiTestFlow.environment)
    ).filter(models.ApiTestFlow.id == db_flow.id).first()
    return db_flow


@app.put("/api/api-flows/{flow_id}", response_model=schemas.ApiTestFlow)
def update_api_flow(
    flow_id: int, 
    flow: schemas.ApiTestFlowUpdate, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    db_flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project).joinedload(models.Project.members)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    if not db_flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    # 检查权限：只有项目成员可以更新流程
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_flow.project, "更新流程")
    update_data = flow.model_dump(exclude_unset=True)
    
    # 特别处理 steps：确保即使为空列表或包含空对象的步骤也能被保存
    # 将 Pydantic 模型转换为字典，以便 JSON 序列化
    if 'steps' in flow.model_dump(exclude_unset=False):
        if flow.steps is not None:
            # 将 FlowStep 对象列表转换为字典列表
            # 如果已经是字典，直接使用；如果是Pydantic模型，转换为字典
            steps_list = []
            for step in flow.steps:
                if hasattr(step, 'model_dump'):
                    # Pydantic 模型，转换为字典
                    step_dict = step.model_dump()
                elif isinstance(step, dict):
                    # 已经是字典，直接使用
                    step_dict = step
                else:
                    # 其他情况，尝试转换为字典
                    step_dict = dict(step) if hasattr(step, '__dict__') else step
                
                # 确保所有参数字段都被包含（包括空对象 {}）
                # 如果字段是 undefined，不包含；如果是空对象 {}，必须包含
                final_step = {}
                for key in ['endpoint_id', 'environment_id', 'test_data_id', 'alias', 'path_params', 'query_params', 'headers', 'body', 'assertions']:
                    if key in step_dict:
                        final_step[key] = step_dict[key]
                    elif hasattr(step, key):
                        value = getattr(step, key)
                        if value is not None:  # 包括空对象 {}
                            final_step[key] = value
                
                steps_list.append(final_step)
            update_data['steps'] = steps_list
            print(f"DEBUG: 保存 steps 数据，共 {len(steps_list)} 个步骤")
            for i, step in enumerate(steps_list):
                print(f"DEBUG: 步骤 {i}: endpoint_id={step.get('endpoint_id')}, headers={step.get('headers')}, assertions={step.get('assertions')}")
        else:
            update_data['steps'] = None
    
    if 'project_id' in update_data:
        project = db.query(models.Project).filter(models.Project.id == update_data['project_id']).first()
        if not project:
            raise HTTPException(status_code=404, detail="项目不存在")
    
    # 检查流程名称唯一性（如果更新了名称）
    if 'name' in update_data:
        project_id = update_data.get('project_id', db_flow.project_id)
        existing_flow = db.query(models.ApiTestFlow).filter(
            models.ApiTestFlow.project_id == project_id,
            models.ApiTestFlow.name == update_data['name'],
            models.ApiTestFlow.id != flow_id  # 排除当前流程
        ).first()
        if existing_flow:
            raise HTTPException(status_code=400, detail="该项目下已存在同名流程")
    
    # 调试：输出要更新的数据
    print(f"DEBUG: 更新流程 {flow_id}，update_data keys: {list(update_data.keys())}")
    if 'steps' in update_data:
        print(f"DEBUG: steps 数据长度: {len(update_data['steps']) if update_data['steps'] else 0}")
        if update_data['steps']:
            for i, step in enumerate(update_data['steps']):
                print(f"DEBUG: 步骤 {i}: endpoint_id={step.get('endpoint_id')}, alias={step.get('alias')}, headers={step.get('headers')}")
    
    for key, value in update_data.items():
        setattr(db_flow, key, value)
    
    # 调试：确认保存前的数据
    print(f"DEBUG: 保存前 db_flow.steps 类型: {type(db_flow.steps)}, 长度: {len(db_flow.steps) if db_flow.steps else 0}")
    
    db.commit()
    db.refresh(db_flow)
    
    # 调试：确认保存后的数据
    print(f"DEBUG: 保存后 db_flow.steps 类型: {type(db_flow.steps)}, 长度: {len(db_flow.steps) if db_flow.steps else 0}")
    if db_flow.steps:
        for i, step in enumerate(db_flow.steps):
            print(f"DEBUG: 保存后步骤 {i}: endpoint_id={step.get('endpoint_id')}, alias={step.get('alias')}, headers={step.get('headers')}")
    # 重新加载以包含关联的项目信息
    db_flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project),
        joinedload(models.ApiTestFlow.environment)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    return db_flow


@app.delete("/api/api-flows/{flow_id}")
def delete_api_flow(
    flow_id: int, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    db_flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project).joinedload(models.Project.members)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    if not db_flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    # 检查权限：只有项目成员可以删除流程
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, db_flow.project, "删除流程")
    
    # 检查依赖：是否有测试任务使用该流程
    task_items = db.query(models.TestTaskItem).filter(
        models.TestTaskItem.item_type == 'flow',
        models.TestTaskItem.item_id == flow_id
    ).all()
    if task_items:
        task_names = []
        for item in task_items:
            task = db.query(models.TestTask).filter(models.TestTask.id == item.task_id).first()
            if task:
                task_names.append(task.name)
        raise HTTPException(
            status_code=400,
            detail=f"无法删除流程：该流程正在被测试任务使用（{', '.join(task_names)}），请先从测试任务中移除该流程"
        )
    
    # 手动删除关联的导出记录（避免外键约束问题）
    db.query(models.FlowExportRecord).filter(models.FlowExportRecord.flow_id == flow_id).delete()
    
    db.delete(db_flow)
    db.commit()
    return {"message": "流程已删除"}


@app.get("/api/api-flows/{flow_id}/variables", response_model=List[schemas.FlowVariable])
def get_flow_variables(flow_id: int, db: Session = Depends(get_db)):
    """获取流程变量列表"""
    flow = db.query(models.ApiTestFlow).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    return flow.variables


@app.post("/api/api-flows/{flow_id}/variables", response_model=List[schemas.FlowVariable])
def save_flow_variables(
    flow_id: int,
    request: schemas.FlowVariableBatchRequest,
    db: Session = Depends(get_db)
):
    """批量保存流程变量"""
    # 使用 joinedload 确保加载 variables 关联
    flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.variables)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    # 获取现有变量，以 (flow_id, key.strip()) 为唯一标识（统一去除空格）
    # 确保访问 flow.variables 以触发加载
    existing_vars = {
        (v.flow_id, v.key.strip()): v for v in flow.variables
    }
    print(f"🔍 现有变量数量: {len(existing_vars)}")
    for (fid, key), v in existing_vars.items():
        print(f"🔍 现有变量: flow_id={fid}, key='{key}', id={v.id}, value={v.value}")
    
    # 获取请求中的变量 key 集合（去除空白）
    request_keys = {var_data.key.strip() for var_data in request.variables if var_data.key.strip()}
    print(f"🔍 请求的变量 keys: {request_keys}")
    
    # 删除不在请求中的变量（根据 key 判断，统一去除空格后比较）
    to_delete = [
        v for (fid, key), v in existing_vars.items()
        if fid == flow_id and key not in request_keys
    ]
    print(f"🔍 要删除的变量数量: {len(to_delete)}")
    for v in to_delete:
        print(f"🔍 删除变量: id={v.id}, key='{v.key}'")
        db.delete(v)
    
    # 更新或创建变量（根据 flow_id 和 key 判断是否存在）
    for var_data in request.variables:
        key = var_data.key.strip()
        if not key:
            continue  # 跳过空的 key
        
        # 根据 flow_id 和 key（已去除空格）查找现有变量
        existing_var = existing_vars.get((flow_id, key))
        
        if existing_var:
            # 更新现有变量
            print(f"🔍 更新变量: key='{key}', id={existing_var.id}, 旧值='{existing_var.value}', 新值='{var_data.value.strip()}'")
            existing_var.value = var_data.value.strip()
        else:
            # 创建新变量
            print(f"🔍 创建新变量: key='{key}', value='{var_data.value.strip()}'")
            new_var = models.FlowVariable(
                flow_id=flow_id,
                key=key,
                value=var_data.value.strip()
            )
            db.add(new_var)
            # 更新 existing_vars 以便后续检查重复的 key
            existing_vars[(flow_id, key)] = new_var
    
    db.commit()
    # 重新加载变量列表（使用 joinedload 确保加载）
    db.refresh(flow)
    # 重新查询以确保获取最新数据
    flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.variables)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    return flow.variables


@app.delete("/api/api-flows/{flow_id}/variables/{variable_id}")
def delete_flow_variable(flow_id: int, variable_id: int, db: Session = Depends(get_db)):
    """删除流程变量"""
    flow = db.query(models.ApiTestFlow).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    var = db.query(models.FlowVariable).filter(
        models.FlowVariable.id == variable_id,
        models.FlowVariable.flow_id == flow_id
    ).first()
    if not var:
        raise HTTPException(status_code=404, detail="变量不存在")
    
    db.delete(var)
    db.commit()
    return {"message": "删除成功"}


@app.post("/api/api-flows/{flow_id}/execute")
def execute_api_flow(
    flow_id: int,
    request: schemas.FlowExecuteRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    # 使用 joinedload 确保加载 variables 和 project.members 关联
    flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.variables),
        joinedload(models.ApiTestFlow.project).joinedload(models.Project.members)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    # 检查权限：只有项目成员可以执行流程
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, flow.project, "执行流程")

    # 初始化上下文，优先加载流程变量
    context: Dict[str, Any] = {}
    # 从流程变量表加载（确保触发加载）
    # 先访问 flow.variables 以确保 joinedload 生效
    variables_list = list(flow.variables) if flow.variables else []
    print(f"🔍 流程变量数量: {len(variables_list)}")
    for var in variables_list:
        context[var.key] = var.value
        print(f"🔍 加载变量: {var.key} = {var.value}")
    # 兼容旧的global_variables字段
    if flow.global_variables:
        # 确保 global_variables 是字典类型
        if isinstance(flow.global_variables, dict):
            context.update(flow.global_variables)
        elif isinstance(flow.global_variables, str):
            # 如果是字符串，尝试解析为JSON
            try:
                import json
                parsed_vars = json.loads(flow.global_variables)
                if isinstance(parsed_vars, dict):
                    context.update(parsed_vars)
            except:
                pass
        print(f"🔍 从global_variables加载: {flow.global_variables}")
        print(f"🔍 加载后的context: {context}")
    # 请求中的变量会覆盖流程变量
    if request.global_variables:
        context.update(request.global_variables)
        print(f"🔍 从请求加载变量: {request.global_variables}")
    
    # 调试日志：输出变量上下文
    print(f"🔍 最终变量上下文: {context}")

    results = []
    overall_success = True
    
    # 获取执行配置中的 failAction，优先使用请求中的，否则使用流程保存的，默认为 'stop'
    fail_action = request.failAction
    if not fail_action and flow.executionConfig:
        fail_action = flow.executionConfig.get("failAction", "stop")
    if not fail_action:
        fail_action = "stop"
    
    # 获取步骤间延迟时间（毫秒）
    step_delay = request.delay or 0
    if step_delay > 0:
        print(f"🔍 步骤间延迟设置为: {step_delay}ms")

    # 重要：每个步骤的参数和执行都是完全独立的
    # 即使多个步骤使用相同的接口（endpoint_id），每个步骤也有自己独立的参数
    # 参数保存在 flow.steps 数组中，通过索引（idx）访问，确保独立性
    # 过滤掉被禁用的步骤（enabled为False的步骤）
    enabled_steps = [step for step in (flow.steps or []) if step.get("enabled") is not False]
    
    for idx, step in enumerate(enabled_steps):
        endpoint = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.id == step.get("endpoint_id")).first()
        if not endpoint:
            raise HTTPException(status_code=400, detail=f"第 {idx+1} 步接口不存在")

        # 确定环境
        environment_id = step.get("environment_id") or request.environment_id or flow.environment_id
        if not environment_id:
            raise HTTPException(status_code=400, detail="请为流程或步骤选择环境")
        environment = db.query(models.ApiEnvironment).filter(models.ApiEnvironment.id == environment_id).first()
        if not environment:
            raise HTTPException(status_code=404, detail=f"第 {idx+1} 步环境不存在")

        # 获取测试数据（可选）
        test_data = None
        test_data_id = step.get("test_data_id")
        if test_data_id:
            test_data = db.query(models.ApiTestData).filter(models.ApiTestData.id == test_data_id).first()
        else:
            test_data = db.query(models.ApiTestData).filter(models.ApiTestData.endpoint_id == endpoint.id).first()

        # 组装请求数据，优先级：环境默认 < 测试数据 < 步骤自定义参数（已模板渲染）
        # 重要：每个步骤使用自己的参数（step.get("headers")等），不会影响其他步骤
        # 即使第2个和第5个步骤使用相同的接口，它们也有完全独立的参数
        headers = {}
        if environment.headers:
            headers.update(environment.headers)
        if test_data and test_data.headers:
            headers.update(test_data.headers)
        # 步骤自己的headers优先级最高，会覆盖测试数据和环境配置
        if step.get("headers") is not None:
            # 调试日志：输出渲染前后的headers
            original_headers = step["headers"]
            rendered_headers = _render_template(step["headers"], context)
            print(f"🔍 步骤 {idx+1} (endpoint_id={step.get('endpoint_id')}, alias={step.get('alias')}) Headers 渲染前: {original_headers}")
            print(f"🔍 步骤 {idx+1} (endpoint_id={step.get('endpoint_id')}, alias={step.get('alias')}) Headers 渲染后: {rendered_headers}")
            # 如果步骤的headers是空对象 {}，表示用户明确清空了headers，应该使用空对象
            if isinstance(rendered_headers, dict):
                headers = rendered_headers  # 使用步骤自己的headers，完全覆盖
            else:
                headers.update(rendered_headers if isinstance(rendered_headers, dict) else {})

        # 路径参数：步骤自己的参数优先级最高
        path_params = test_data.path_params if test_data else None
        if step.get("path_params") is not None:
            # 如果步骤有path_params（包括空对象 {}），使用步骤的参数
            path_params = _render_template(step["path_params"], context)

        # 查询参数：步骤自己的参数优先级最高
        query_params = test_data.query_params if test_data else None
        if step.get("query_params") is not None:
            # 如果步骤有query_params（包括空对象 {}），使用步骤的参数
            query_params = _render_template(step["query_params"], context)

        # 请求体：步骤自己的参数优先级最高
        body = test_data.body if test_data else None
        if step.get("body") is not None:
            # 如果步骤有body（包括空对象 {}），使用步骤的参数
            body = _render_template(step["body"], context)
        
        # 调试日志：确认每个步骤使用自己的参数
        print(f"🔍 步骤 {idx+1} (endpoint_id={step.get('endpoint_id')}, alias={step.get('alias')}) 使用的参数:")
        print(f"  - path_params: {path_params}")
        print(f"  - query_params: {query_params}")
        print(f"  - headers: {headers}")
        print(f"  - body: {body}")

        # 拼接 URL 并替换路径参数
        base_url = environment.base_url.rstrip('/')
        path = endpoint.path.lstrip('/')
        full_url = f"{base_url}/{path}"
        if path_params:
            for key, value in path_params.items():
                full_url = full_url.replace(f"{{{key}}}", str(value))

        start_time = time.time()
        step_success = False
        response_status = None
        response_headers = None
        response_body_text = None
        error_message = None
        json_body = None
        assertions_list = []  # 初始化断言列表，确保在异常情况下也能使用

        try:
            # 调试日志：输出实际发送的请求headers
            print(f"🔍 步骤 {idx+1} 实际发送的 Headers: {headers}")
            response = requests.request(
                method=endpoint.method.upper(),
                url=full_url,
                headers=headers,
                params=query_params,
                json=body if body is not None else None,
                timeout=30
            )
            response_time = int((time.time() - start_time) * 1000)
            response_status = response.status_code
            response_headers = dict(response.headers)
            response_body_text = response.text
            try:
                json_body = response.json()
            except Exception:
                json_body = None

            # 检查断言
            # 重要：每个步骤使用自己的断言列表，完全独立
            # 即使多个步骤使用相同的接口，每个步骤也有自己独立的断言
            # 重要：每个步骤使用自己的断言列表，完全独立
            # 即使多个步骤使用相同的接口，每个步骤也有自己独立的断言
            step_assertions = step.get("assertions")
            if step_assertions is None:
                assertions_list = []
            elif isinstance(step_assertions, list):
                assertions_list = step_assertions
            else:
                assertions_list = []
            
            print(f"🔍 步骤 {idx+1} (endpoint_id={step.get('endpoint_id')}, alias={step.get('alias')}) 使用的断言: {assertions_list}")
            
            print(f"🔍 步骤 {idx+1} (endpoint_id={step.get('endpoint_id')}) 使用的断言: {assertions_list}")
            
            assertion_errors = []
            step_success = True
            
            if assertions_list and len(assertions_list) > 0:
                # 有断言时，初始化为True，只有断言失败时才设为False
                step_success = True
                for assertion in assertions_list:
                    assertion_type = assertion.get('type')
                    operator = assertion.get('operator')
                    target = assertion.get('target')
                    expected = assertion.get('expected')
                    
                    # 跳过无效的断言
                    if not assertion_type or not operator:
                        continue
                    
                    if assertion_type == 'status_code':
                        # 状态码断言
                        actual_value = response_status
                        # 确保expected不为None或空字符串
                        if expected is None or expected == '':
                            step_success = False
                            assertion_errors.append(f"状态码断言失败: 期望值不能为空")
                            continue
                        # 检查断言（_check_assertion会将两者都转为字符串比较）
                        # 确保expected是字符串类型，以便正确比较
                        expected_str = str(expected).strip()
                        # 确保actual_value也是整数类型（response_status已经是整数）
                        assertion_passed = _check_assertion(actual_value, operator, expected_str)
                        if not assertion_passed:
                            step_success = False
                            assertion_errors.append(f"状态码断言失败: 期望 {expected_str}，实际 {actual_value}")
                        # 断言通过时，确保step_success保持为True
                        else:
                            # 断言通过，但需要确保step_success保持为True（防止被其他逻辑覆盖）
                            pass
                    
                    elif assertion_type == 'json_path':
                        # JSON路径断言
                        if json_body and target:
                            # 先渲染 expected 值（处理变量替换，如 STR($CreateValue)）
                            if expected:
                                try:
                                    expected_rendered = _render_template(expected, context)
                                    # 如果渲染后是字符串且以引号开头结尾，去掉引号
                                    if isinstance(expected_rendered, str) and expected_rendered.startswith('"') and expected_rendered.endswith('"'):
                                        expected_rendered = expected_rendered[1:-1]
                                    expected = expected_rendered
                                except Exception as e:
                                    print(f"🔍 渲染 expected 值失败: {e}")
                            
                            actual_value = _extract_json_path(json_body, target)
                            print(f"🔍 JSON路径断言: path={target}, expected={expected}, actual={actual_value}")
                            
                            if not _check_assertion(actual_value, operator, expected):
                                step_success = False
                                assertion_errors.append(f"JSON路径断言失败: {target} 期望 {expected}，实际 {_format_value_for_display(actual_value)}")
                        else:
                            step_success = False
                            assertion_errors.append(f"JSON路径断言失败: 无法提取路径 {target}")
                    
                    elif assertion_type == 'response_time':
                        # 响应时间断言
                        actual_value = response_time
                        if not _check_assertion(actual_value, operator, expected):
                            step_success = False
                            assertion_errors.append(f"响应时间断言失败: 期望 {expected}ms，实际 {actual_value}ms")
                    
                    elif assertion_type == 'contains':
                        # 包含断言
                        if json_body:
                            body_str = json.dumps(json_body) if isinstance(json_body, dict) else str(json_body)
                            if expected and expected not in body_str:
                                step_success = False
                                assertion_errors.append(f"包含断言失败: 响应体中不包含 {expected}")
                        else:
                            if expected and expected not in str(response_body_text):
                                step_success = False
                                assertion_errors.append(f"包含断言失败: 响应体中不包含 {expected}")
                
                # 如果有断言错误，设置错误消息
                if assertion_errors:
                    error_message = "\n".join(assertion_errors)
            else:
                # 如果没有断言，使用默认逻辑：状态码在200-299之间
                step_success = 200 <= response_status < 300
        except Exception as exc:  # pragma: no cover
            response_time = int((time.time() - start_time) * 1000)
            error_message = str(exc)
            step_success = False

        # 将当前接口的响应体存储到 context 中，支持 API[N] 语法
        api_key = f"API[{idx + 1}]"
        context[api_key] = json_body or {}
        print(f"🔍 步骤 {idx+1} 存储响应体到 context[{api_key}]")
        
        # 变量提取：从当前接口提取（step_index 为 None、0 或等于当前接口序号）
        for rule in step.get("extracts") or []:
            rule_step_index = rule.get("step_index")
            # 如果 step_index 为 None、0 或等于当前接口序号，从当前接口提取
            if not rule_step_index or rule_step_index == 0 or rule_step_index == (idx + 1):
                value = _extract_value(json_body or {}, rule.get("path", ""))
                if rule.get("name"):
                    context[rule["name"]] = value
                    print(f"🔍 步骤 {idx+1} 从当前接口提取变量: {rule.get('name')} = {value}")

        results.append({
            "index": idx + 1,
            "endpoint_id": endpoint.id,
            "endpoint_name": endpoint.name,
            "url": full_url,
            "method": endpoint.method,
            "success": step_success,
            "status": response_status,
            "response_time": response_time,
            "error_message": error_message,
            "alias": step.get("alias"),
            "extracted": {rule.get("name"): _extract_value(json_body or {}, rule.get("path", "")) for rule in step.get("extracts") or []},
            # 请求信息
            "request_headers": headers,
            "request_path_params": path_params,
            "request_query_params": query_params,
            "request_body": body,
            "request_assertions": assertions_list,  # 添加断言信息
            # 响应信息
            "response_headers": response_headers,
            "response_body": response_body_text
        })
        if not step_success:
            overall_success = False
            if fail_action == "stop":
                # 停止执行
                break
            # 如果 fail_action == "continue"，继续执行下一个步骤
        
        # 步骤间延迟（不是最后一个步骤时）
        if step_delay > 0 and idx < len(enabled_steps) - 1:
            print(f"🔍 步骤 {idx+1} 执行完成，延迟 {step_delay}ms 后执行下一步骤")
            time.sleep(step_delay / 1000.0)  # 转换为秒

    return {
        "success": overall_success,
        "results": results,
        "context": context
    }


# ==================== 流程导出和导入 ====================

@app.post("/api/api-flows/{flow_id}/export", response_model=schemas.FlowExportRecord)
def export_api_flow(flow_id: int, db: Session = Depends(get_db)):
    """导出流程到数据库"""
    flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project),
        joinedload(models.ApiTestFlow.environment),
        joinedload(models.ApiTestFlow.variables)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    # 构建导出数据
    export_data = {
        "version": "1.0",
        "exportTime": datetime.now().isoformat(),
        "flow": {
            "name": flow.name,
            "description": flow.description,
            "project_id": flow.project_id,
            "environment_id": flow.environment_id,
            "global_variables": flow.global_variables or {},
            "steps": flow.steps or []
        },
        "variables": [
            {"key": v.key, "value": v.value}
            for v in flow.variables
        ]
    }
    
    # 生成导出名称（带时间戳）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c for c in flow.name if c.isalnum() or c in (' ', '-', '_')).strip()
    if not safe_name:
        safe_name = f"flow_{flow_id}"
    export_name = f"{safe_name}_{timestamp}"
    
    # 保存到数据库
    export_record = models.FlowExportRecord(
        flow_id=flow_id,
        name=export_name,
        export_data=export_data
    )
    db.add(export_record)
    db.commit()
    db.refresh(export_record)
    
    return export_record


@app.get("/api/api-flows/{flow_id}/exports", response_model=List[schemas.FlowExportRecord])
def get_flow_exports(flow_id: int, db: Session = Depends(get_db)):
    """获取流程的所有导出记录"""
    flow = db.query(models.ApiTestFlow).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    exports = db.query(models.FlowExportRecord).filter(
        models.FlowExportRecord.flow_id == flow_id
    ).order_by(models.FlowExportRecord.created_at.desc()).all()
    
    return exports


@app.get("/api/api-flows/{flow_id}/exports/{export_id}", response_model=schemas.FlowExportRecord)
def get_flow_export(flow_id: int, export_id: int, db: Session = Depends(get_db)):
    """获取单个导出记录"""
    export_record = db.query(models.FlowExportRecord).filter(
        models.FlowExportRecord.id == export_id,
        models.FlowExportRecord.flow_id == flow_id
    ).first()
    if not export_record:
        raise HTTPException(status_code=404, detail="导出记录不存在")
    return export_record


@app.post("/api/api-flows/{flow_id}/import/{export_id}")
def import_api_flow(
    flow_id: int, 
    export_id: int, 
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """从导出记录导入流程数据"""
    flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project).joinedload(models.Project.members)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    if not flow:
        raise HTTPException(status_code=404, detail="流程不存在")
    
    # 检查权限：只有项目成员可以导入流程
    user = db.query(models.User).filter(models.User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")
    check_project_member_permission(user, flow.project, "导入流程")
    
    export_record = db.query(models.FlowExportRecord).filter(
        models.FlowExportRecord.id == export_id,
        models.FlowExportRecord.flow_id == flow_id
    ).first()
    if not export_record:
        raise HTTPException(status_code=404, detail="导出记录不存在")
    
    export_data = export_record.export_data
    
    # 更新流程基本信息（不更新名称，保持当前流程名称）
    if "flow" in export_data:
        flow_data = export_data["flow"]
        # 注意：不更新流程名称，保持当前流程的名称不变
        # if "name" in flow_data:
        #     flow.name = flow_data["name"]
        if "description" in flow_data:
            flow.description = flow_data.get("description")
        if "project_id" in flow_data:
            flow.project_id = flow_data["project_id"]
        if "environment_id" in flow_data:
            flow.environment_id = flow_data.get("environment_id")
        if "global_variables" in flow_data:
            flow.global_variables = flow_data["global_variables"]
        if "steps" in flow_data:
            flow.steps = flow_data["steps"]
    
    # 更新流程变量
    if "variables" in export_data:
        # 删除现有变量
        db.query(models.FlowVariable).filter(
            models.FlowVariable.flow_id == flow_id
        ).delete()
        
        # 创建新变量
        for var_data in export_data["variables"]:
            if var_data.get("key") and var_data.get("value"):
                new_var = models.FlowVariable(
                    flow_id=flow_id,
                    key=var_data["key"],
                    value=var_data["value"]
                )
                db.add(new_var)
    
    db.commit()
    db.refresh(flow)
    
    # 重新加载以包含关联信息
    flow = db.query(models.ApiTestFlow).options(
        joinedload(models.ApiTestFlow.project),
        joinedload(models.ApiTestFlow.environment),
        joinedload(models.ApiTestFlow.variables)
    ).filter(models.ApiTestFlow.id == flow_id).first()
    
    return flow


@app.delete("/api/api-flows/{flow_id}/exports/{export_id}")
def delete_flow_export(flow_id: int, export_id: int, db: Session = Depends(get_db)):
    """删除导出记录"""
    export_record = db.query(models.FlowExportRecord).filter(
        models.FlowExportRecord.id == export_id,
        models.FlowExportRecord.flow_id == flow_id
    ).first()
    if not export_record:
        raise HTTPException(status_code=404, detail="导出记录不存在")
    
    db.delete(export_record)
    db.commit()
    return {"message": "删除成功"}

# ==================== 执行记录 ====================

@app.get("/api/api-execution-records", response_model=List[schemas.ApiExecutionRecord])
def get_api_execution_records(
    endpoint_id: Optional[int] = None,
    environment_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """获取执行记录"""
    query = db.query(models.ApiExecutionRecord).order_by(models.ApiExecutionRecord.executed_at.desc())
    
    if endpoint_id:
        query = query.filter(models.ApiExecutionRecord.endpoint_id == endpoint_id)
    
    if environment_id:
        query = query.filter(models.ApiExecutionRecord.environment_id == environment_id)
    
    return query.offset(skip).limit(limit).all()

@app.get("/api/api-execution-records/{record_id}", response_model=schemas.ApiExecutionRecord)
def get_api_execution_record(record_id: int, db: Session = Depends(get_db)):
    """获取单个执行记录"""
    record = db.query(models.ApiExecutionRecord).filter(models.ApiExecutionRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="执行记录不存在")
    return record

# ==================== Swagger同步和上传 ====================

# 常见 OpenAPI/Swagger 文档路径（Springdoc / FastAPI / 旧版 Swagger2）
_COMMON_OPENAPI_PATHS = ("/v3/api-docs", "/openapi.json", "/v2/api-docs")


def _fetch_openapi_spec_with_fallbacks(base_url: str, swagger_path: str) -> tuple:
    """
    先请求用户指定的 swagger_path；若 404 或返回非 OpenAPI/Swagger JSON，再依次尝试常见路径。
    401/403 不重试（需服务端允许匿名访问文档或单独配置）。
    返回 (spec_dict, 实际生效的路径)。
    """
    base = base_url.rstrip("/")
    raw = (swagger_path or "/v3/api-docs").strip()
    primary = raw if raw.startswith("/") else f"/{raw}"

    candidates = []
    seen = set()
    for p in (primary, *_COMMON_OPENAPI_PATHS):
        if p not in seen:
            seen.add(p)
            candidates.append(p)

    errors = []
    for p in candidates:
        url = f"{base}{p}"
        try:
            r = requests.get(url, timeout=30)
        except requests.RequestException as e:
            errors.append(f"{url}: {e}")
            continue

        if r.status_code == 404:
            errors.append(f"{url}: 404")
            continue

        if r.status_code in (401, 403):
            raise HTTPException(
                status_code=400,
                detail=(
                    f"无法访问Swagger文档: {url} 返回 {r.status_code}。"
                    " 请允许匿名访问文档或改用可访问的文档地址。"
                ),
            )

        try:
            r.raise_for_status()
        except requests.RequestException as e:
            errors.append(f"{url}: HTTP {r.status_code}")
            continue

        try:
            spec = r.json()
        except ValueError:
            errors.append(f"{url}: 响应非 JSON")
            continue

        if isinstance(spec, dict) and ("openapi" in spec or "swagger" in spec):
            return spec, p

        errors.append(f"{url}: 非 OpenAPI/Swagger 文档结构")

    msg = "无法获取有效的 OpenAPI/Swagger 文档。已依次尝试: " + ", ".join(candidates) + "。"
    if errors:
        msg += " 详情: " + "; ".join(errors[:8])
        if len(errors) > 8:
            msg += "…"
    raise HTTPException(status_code=400, detail=msg)


@app.post("/api/api-endpoints/sync")
def sync_swagger_from_environment(
    environment_id: int = Query(...),
    project_id: int = Query(...),
    swagger_path: str = Query("/v3/api-docs"),
    db: Session = Depends(get_db)
):
    """从环境同步Swagger接口（自动兼容 Springdoc /v3/api-docs 与 FastAPI /openapi.json 等）"""
    # 获取环境
    environment = db.query(models.ApiEnvironment).filter(models.ApiEnvironment.id == environment_id).first()
    if not environment:
        raise HTTPException(status_code=404, detail="环境不存在")
    
    # 检查项目是否存在
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    base_url = environment.base_url.rstrip("/")
    
    try:
        spec, resolved_swagger_path = _fetch_openapi_spec_with_fallbacks(base_url, swagger_path)
        
        # 解析接口
        parser = OpenAPIParser(spec)
        apis = parser.parse()
        
        # 调试：检查第一个接口的数据结构
        if apis and len(apis) > 0:
            import logging
            import json
            logger = logging.getLogger(__name__)
            first_api = apis[0]
            logger.info(f"解析的第一个接口: {first_api.get('name')}")
            logger.info(f"  - parameters 数量: {len(first_api.get('parameters', []))}")
            logger.info(f"  - parameters 内容: {json.dumps(first_api.get('parameters', [])[:2], ensure_ascii=False) if len(first_api.get('parameters', [])) > 0 else '[]'}")
            logger.info(f"  - request_body: {bool(first_api.get('request_body'))}")
            if first_api.get('request_body'):
                req_body = first_api.get('request_body', {})
                logger.info(f"  - request_body.schema: {bool(req_body.get('schema'))}")
                if req_body.get('schema'):
                    schema = req_body.get('schema', {})
                    logger.info(f"  - schema.type: {schema.get('type')}")
                    logger.info(f"  - schema.properties: {bool(schema.get('properties'))}")
        
        # 删除该项目下所有旧接口及其测试数据
        deleted_count = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.project_id == project_id).count()
        old_endpoints = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.project_id == project_id).all()
        old_ids = [e.id for e in old_endpoints]
        if old_ids:
            db.query(models.ApiTestData).filter(models.ApiTestData.endpoint_id.in_(old_ids)).delete(synchronize_session=False)
        db.query(models.ApiEndpoint).filter(models.ApiEndpoint.project_id == project_id).delete()
        db.commit()
        
        # 先保存所有接口
        saved_count = 0
        created_endpoints = []
        for api_data in apis:
            # 确保 parameters 是列表，request_body 是字典或 None
            parameters = api_data.get('parameters')
            if parameters is None:
                parameters = []
            elif not isinstance(parameters, list):
                parameters = []
            
            request_body = api_data.get('request_body')
            if request_body is not None and not isinstance(request_body, dict):
                request_body = None
            
            db_endpoint = models.ApiEndpoint(
                project_id=project_id,
                name=api_data['name'],
                path=api_data['path'],
                method=api_data['method'],
                description=api_data.get('description', ''),
                tags=api_data.get('tags', []),
                parameters=parameters,  # 确保是列表
                request_body=request_body  # 确保是字典或 None
            )
            db.add(db_endpoint)
            created_endpoints.append(db_endpoint)
            saved_count += 1
        
        db.commit()  # 先提交，确保数据已保存

        # 然后为每个接口生成测试数据（使用已保存的数据库对象，就像 platform 那样）
        default_created = 0
        import logging
        logger = logging.getLogger(__name__)
        
        for endpoint in created_endpoints:
            # 刷新对象，确保数据已加载
            db.refresh(endpoint)
            
            # 直接使用数据库模型对象，就像 platform 那样
            # SQLAlchemy 会自动将 JSON 字段反序列化为 Python 对象
            import json
            logger.info(f"生成测试数据 - 接口: {endpoint.name} (ID: {endpoint.id})")
            logger.info(f"  - parameters类型: {type(endpoint.parameters).__name__}")
            logger.info(f"  - parameters值: {json.dumps(endpoint.parameters, ensure_ascii=False)[:500] if endpoint.parameters else 'None'}")
            logger.info(f"  - parameters数量: {len(endpoint.parameters) if isinstance(endpoint.parameters, list) else 0}")
            logger.info(f"  - request_body类型: {type(endpoint.request_body).__name__}")
            logger.info(f"  - request_body值: {json.dumps(endpoint.request_body, ensure_ascii=False)[:500] if endpoint.request_body else 'None'}")
            
            # 检查数据是否有效
            if endpoint.parameters is None:
                logger.warning(f"  ⚠️ endpoint.parameters 是 None，应该是一个列表")
            elif not isinstance(endpoint.parameters, list):
                logger.warning(f"  ⚠️ endpoint.parameters 不是列表，类型: {type(endpoint.parameters).__name__}")
            
            if endpoint.request_body is None:
                logger.info(f"  ℹ️ endpoint.request_body 是 None（这是正常的，如果接口没有请求体）")
            elif not isinstance(endpoint.request_body, dict):
                logger.warning(f"  ⚠️ endpoint.request_body 不是字典，类型: {type(endpoint.request_body).__name__}")
            
            # 使用 TestDataGenerator 生成测试数据（直接使用模型对象）
            try:
                test_data = TestDataGenerator.generate_test_data(endpoint)
            except Exception as e:
                logger.error(f"  ❌ 生成测试数据时出错: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                continue
            
            # 调试：打印生成的测试数据
            logger.info(f"生成的测试数据:")
            logger.info(f"  - path_params: {json.dumps(test_data.get('path_params'), ensure_ascii=False)}")
            logger.info(f"  - query_params: {json.dumps(test_data.get('query_params'), ensure_ascii=False)}")
            logger.info(f"  - headers: {json.dumps(test_data.get('headers'), ensure_ascii=False)}")
            logger.info(f"  - body: {json.dumps(test_data.get('body'), ensure_ascii=False)[:200] if test_data.get('body') else '{}'}")
            
            # 获取生成的数据
            path_params = test_data.get('path_params')
            query_params = test_data.get('query_params')
            headers = test_data.get('headers')
            body = test_data.get('body')
            
            # 调试：打印保存前的数据
            logger.info(f"保存前的数据:")
            logger.info(f"  - path_params: {path_params} (类型: {type(path_params).__name__}, 是否为空: {not path_params if isinstance(path_params, dict) else path_params is None})")
            logger.info(f"  - query_params: {query_params} (类型: {type(query_params).__name__}, 是否为空: {not query_params if isinstance(query_params, dict) else query_params is None})")
            logger.info(f"  - headers: {headers} (类型: {type(headers).__name__}, 是否为空: {not headers if isinstance(headers, dict) else headers is None})")
            logger.info(f"  - body: {body} (类型: {type(body).__name__}, 是否为空: {not body if isinstance(body, dict) else body is None})")
            
            # 创建测试数据记录（直接使用生成的数据，就像 platform 那样）
            # 注意：即使数据是空字典 {}，也要保存（SQLAlchemy JSON 字段可以处理）
            default_data = models.ApiTestData(
                endpoint_id=endpoint.id,
                name=endpoint.path,  # 使用接口路径作为名称
                path_params=path_params,  # 直接使用，不判断是否为空
                query_params=query_params,
                headers=headers,
                body=body,
                expected_status=200
            )
            db.add(default_data)
            default_created += 1
            
            # 调试：打印保存后的对象
            db.flush()  # 刷新以获取 ID
            logger.info(f"测试数据已创建 (ID: {default_data.id})")
            logger.info(f"  - path_params: {default_data.path_params}")
            logger.info(f"  - query_params: {default_data.query_params}")
            logger.info(f"  - headers: {default_data.headers}")
            logger.info(f"  - body: {default_data.body}")
        
        db.commit()
        
        return {
            "message": f"同步成功：删除 {deleted_count} 个旧接口，导入 {saved_count} 个新接口，生成 {default_created} 条默认测试数据（文档路径: {resolved_swagger_path}）",
            "deleted_count": deleted_count,
            "imported_count": saved_count,
            "swagger_path_used": resolved_swagger_path,
        }
        
    except HTTPException:
        raise
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"文档解析失败: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"同步失败: {str(e)}")

@app.post("/api/api-endpoints/upload")
async def upload_swagger_file(
    file: UploadFile = File(...),
    project_id: int = Query(...),
    db: Session = Depends(get_db)
):
    """上传并解析Swagger文件"""
    if not file.filename.endswith('.json'):
        raise HTTPException(status_code=400, detail="只支持JSON格式的Swagger文件")
    
    # 检查项目是否存在
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 读取文件内容
    content = await file.read()
    
    # 解析Swagger
    try:
        apis = parse_swagger_file(content, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # 删除该项目下所有旧接口及其测试数据
    deleted_count = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.project_id == project_id).count()
    old_endpoints = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.project_id == project_id).all()
    old_ids = [e.id for e in old_endpoints]
    if old_ids:
        db.query(models.ApiTestData).filter(models.ApiTestData.endpoint_id.in_(old_ids)).delete(synchronize_session=False)
    db.query(models.ApiEndpoint).filter(models.ApiEndpoint.project_id == project_id).delete()
    db.commit()
    
    # 先保存所有接口
    saved_count = 0
    created_endpoints = []
    for api_data in apis:
        # 确保 parameters 是列表，request_body 是字典或 None
        parameters = api_data.get('parameters')
        if parameters is None:
            parameters = []
        elif not isinstance(parameters, list):
            parameters = []
        
        request_body = api_data.get('request_body')
        if request_body is not None and not isinstance(request_body, dict):
            request_body = None
        
        db_endpoint = models.ApiEndpoint(
            project_id=project_id,
            name=api_data['name'],
            path=api_data['path'],
            method=api_data['method'],
            description=api_data.get('description', ''),
            tags=api_data.get('tags', []),
            parameters=parameters,  # 确保是列表
            request_body=request_body  # 确保是字典或 None
        )
        db.add(db_endpoint)
        created_endpoints.append(db_endpoint)
        saved_count += 1
    
    db.commit()  # 先提交，确保数据已保存

    # 然后为每个接口生成测试数据（使用已保存的数据库对象，就像 platform 那样）
    default_created = 0
    import logging
    logger = logging.getLogger(__name__)
    
    for endpoint in created_endpoints:
        # 刷新对象，确保数据已加载
        db.refresh(endpoint)
        
        # 直接使用数据库模型对象，就像 platform 那样
        import json
        logger.info(f"上传文件 - 生成测试数据 - 接口: {endpoint.name} (ID: {endpoint.id})")
        logger.info(f"  - parameters类型: {type(endpoint.parameters).__name__}")
        logger.info(f"  - parameters值: {json.dumps(endpoint.parameters, ensure_ascii=False)[:500] if endpoint.parameters else 'None'}")
        logger.info(f"  - parameters数量: {len(endpoint.parameters) if isinstance(endpoint.parameters, list) else 0}")
        logger.info(f"  - request_body类型: {type(endpoint.request_body).__name__}")
        logger.info(f"  - request_body值: {json.dumps(endpoint.request_body, ensure_ascii=False)[:500] if endpoint.request_body else 'None'}")
        
        # 使用 TestDataGenerator 生成测试数据（直接使用模型对象）
        test_data = TestDataGenerator.generate_test_data(endpoint)
        
        # 调试：打印生成的测试数据
        logger.info(f"生成的测试数据:")
        logger.info(f"  - path_params: {json.dumps(test_data.get('path_params'), ensure_ascii=False)}")
        logger.info(f"  - query_params: {json.dumps(test_data.get('query_params'), ensure_ascii=False)}")
        logger.info(f"  - headers: {json.dumps(test_data.get('headers'), ensure_ascii=False)}")
        logger.info(f"  - body: {json.dumps(test_data.get('body'), ensure_ascii=False)[:200] if test_data.get('body') else '{}'}")
        
        # 获取生成的数据
        path_params = test_data.get('path_params')
        query_params = test_data.get('query_params')
        headers = test_data.get('headers')
        body = test_data.get('body')
        
        # 调试：打印保存前的数据
        logger.info(f"保存前的数据:")
        logger.info(f"  - path_params: {path_params} (类型: {type(path_params).__name__}, 是否为空: {not path_params if isinstance(path_params, dict) else path_params is None})")
        logger.info(f"  - query_params: {query_params} (类型: {type(query_params).__name__}, 是否为空: {not query_params if isinstance(query_params, dict) else query_params is None})")
        logger.info(f"  - headers: {headers} (类型: {type(headers).__name__}, 是否为空: {not headers if isinstance(headers, dict) else headers is None})")
        logger.info(f"  - body: {body} (类型: {type(body).__name__}, 是否为空: {not body if isinstance(body, dict) else body is None})")
        
        # 创建测试数据记录（直接使用生成的数据，就像 platform 那样）
        # 注意：即使数据是空字典 {}，也要保存（SQLAlchemy JSON 字段可以处理）
        default_data = models.ApiTestData(
            endpoint_id=endpoint.id,
            name=endpoint.path,  # 使用接口路径作为名称
            path_params=path_params,  # 直接使用，不判断是否为空
            query_params=query_params,
            headers=headers,
            body=body,
            expected_status=200
        )
        db.add(default_data)
        default_created += 1
        
        # 调试：打印保存后的对象
        db.flush()  # 刷新以获取 ID
        logger.info(f"测试数据已创建 (ID: {default_data.id})")
        logger.info(f"  - path_params: {default_data.path_params}")
        logger.info(f"  - query_params: {default_data.query_params}")
        logger.info(f"  - headers: {default_data.headers}")
        logger.info(f"  - body: {default_data.body}")
    
    db.commit()
    
    return {
        "message": f"上传成功：删除 {deleted_count} 个旧接口，导入 {saved_count} 个新接口，生成 {default_created} 条默认测试数据",
        "deleted_count": deleted_count,
        "imported_count": saved_count,
        "filename": file.filename
    }

# ==================== 测试任务管理 ====================

@app.get("/api/test-tasks")
def get_test_tasks(
    project_id: Optional[int] = Query(None),
    project_ids: Optional[str] = Query(None, description="多个项目ID，用逗号分隔"),
    keyword: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    is_favorite: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=10000),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取测试任务列表"""
    require_permission(current_user.role, "apitest", "read")
    
    query = db.query(models.TestTask).options(
        joinedload(models.TestTask.project),
        joinedload(models.TestTask.items)
    )
    
    # 支持多项目ID过滤（优先使用 project_ids）
    if project_ids:
        try:
            project_id_list = [int(id.strip()) for id in project_ids.split(',') if id.strip()]
            if project_id_list:
                query = query.filter(models.TestTask.project_id.in_(project_id_list))
        except ValueError:
            pass  # 如果解析失败，忽略该参数
    elif project_id:
        query = query.filter(models.TestTask.project_id == project_id)
    
    if keyword:
        query = query.filter(
            or_(
                models.TestTask.name.contains(keyword),
                models.TestTask.description.contains(keyword)
            )
        )
    
    if status:
        query = query.filter(models.TestTask.status == status)
    
    if is_favorite is not None:
        query = query.filter(models.TestTask.is_favorite == is_favorite)

    total = query.count()
    items = query.order_by(models.TestTask.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "items": items, "page": page, "page_size": page_size}


@app.get("/api/test-tasks/{task_id}", response_model=schemas.TestTask)
def get_test_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取测试任务详情"""
    require_permission(current_user.role, "apitest", "read")
    
    task = db.query(models.TestTask).options(
        joinedload(models.TestTask.project),
        joinedload(models.TestTask.items)
    ).filter(models.TestTask.id == task_id).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="测试任务不存在")
    
    return task


@app.post("/api/test-tasks", response_model=schemas.TestTask)
def create_test_task(
    task: schemas.TestTaskCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建测试任务"""
    require_permission(current_user.role, "apitest", "create")
    
    # 验证项目是否存在
    project = db.query(models.Project).filter(models.Project.id == task.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 创建任务
    db_task = models.TestTask(
        name=task.name,
        project_id=task.project_id,
        description=task.description,
        status='idle',
        cron_expression=task.cron_expression,
        environment_id=task.environment_id
    )
    db.add(db_task)
    db.flush()
    
    # 创建任务项
    if task.items:
        for idx, item in enumerate(task.items):
            # 验证接口或流程是否存在
            if item.item_type == 'api':
                endpoint = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.id == item.item_id).first()
                if not endpoint:
                    raise HTTPException(status_code=404, detail=f"接口 ID {item.item_id} 不存在")
            elif item.item_type == 'flow':
                flow = db.query(models.ApiTestFlow).filter(models.ApiTestFlow.id == item.item_id).first()
                if not flow:
                    raise HTTPException(status_code=404, detail=f"流程 ID {item.item_id} 不存在")
            
            db_item = models.TestTaskItem(
                task_id=db_task.id,
                item_type=item.item_type,
                item_id=item.item_id,
                sort_order=item.sort_order if item.sort_order > 0 else idx
            )
            db.add(db_item)
    
    db.commit()
    db.refresh(db_task)
    
    # 重新加载关联数据
    db_task = db.query(models.TestTask).options(
        joinedload(models.TestTask.project),
        joinedload(models.TestTask.items)
    ).filter(models.TestTask.id == db_task.id).first()
    
    return db_task


@app.put("/api/test-tasks/{task_id}", response_model=schemas.TestTask)
def update_test_task(
    task_id: int,
    task: schemas.TestTaskUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新测试任务"""
    require_permission(current_user.role, "apitest", "update")
    
    db_task = db.query(models.TestTask).options(
        joinedload(models.TestTask.items)
    ).filter(models.TestTask.id == task_id).first()
    
    if not db_task:
        raise HTTPException(status_code=404, detail="测试任务不存在")
    
    # 更新基本信息
    update_data = task.model_dump(exclude_unset=True, exclude={'items'})
    for field, value in update_data.items():
        setattr(db_task, field, value)
    
    # 更新任务项
    if 'items' in task.model_dump(exclude_unset=True):
        # 删除旧的任务项
        db.query(models.TestTaskItem).filter(models.TestTaskItem.task_id == task_id).delete()
        
        # 创建新的任务项
        if task.items:
            for idx, item in enumerate(task.items):
                # 验证接口或流程是否存在
                if item.item_type == 'api':
                    endpoint = db.query(models.ApiEndpoint).filter(models.ApiEndpoint.id == item.item_id).first()
                    if not endpoint:
                        raise HTTPException(status_code=404, detail=f"接口 ID {item.item_id} 不存在")
                elif item.item_type == 'flow':
                    flow = db.query(models.ApiTestFlow).filter(models.ApiTestFlow.id == item.item_id).first()
                    if not flow:
                        raise HTTPException(status_code=404, detail=f"流程 ID {item.item_id} 不存在")
                
                db_item = models.TestTaskItem(
                    task_id=task_id,
                    item_type=item.item_type,
                    item_id=item.item_id,
                    sort_order=item.sort_order if item.sort_order > 0 else idx
                )
                db.add(db_item)
    
    db.commit()
    db.refresh(db_task)
    
    # 重新加载关联数据
    db_task = db.query(models.TestTask).options(
        joinedload(models.TestTask.project),
        joinedload(models.TestTask.items)
    ).filter(models.TestTask.id == task_id).first()
    
    return db_task


@app.delete("/api/test-tasks/{task_id}")
def delete_test_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除测试任务"""
    require_permission(current_user.role, "apitest", "delete")
    
    db_task = db.query(models.TestTask).filter(models.TestTask.id == task_id).first()
    if not db_task:
        raise HTTPException(status_code=404, detail="测试任务不存在")
    
    db.delete(db_task)
    db.commit()
    
    return {"message": "测试任务已删除"}


@app.post("/api/test-tasks/{task_id}/toggle-favorite")
def toggle_test_task_favorite(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """切换测试任务收藏状态"""
    require_permission(current_user.role, "apitest", "read")
    
    db_task = db.query(models.TestTask).filter(models.TestTask.id == task_id).first()
    if not db_task:
        raise HTTPException(status_code=404, detail="测试任务不存在")
    
    db_task.is_favorite = not db_task.is_favorite
    db.commit()
    
    return {"message": "收藏状态已更新", "is_favorite": db_task.is_favorite}


@app.post("/api/test-tasks/{task_id}/execute", response_model=schemas.TestTaskExecution)
def execute_test_task(
    task_id: int,
    request: schemas.TestTaskExecutionRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """执行测试任务"""
    require_permission(current_user.role, "apitest", "execute")
    
    # 获取任务
    task = db.query(models.TestTask).options(
        joinedload(models.TestTask.items)
    ).filter(models.TestTask.id == task_id).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="测试任务不存在")
    
    if task.status == 'running':
        raise HTTPException(status_code=400, detail="任务正在执行中")
    
    # 验证环境
    environment = db.query(models.ApiEnvironment).filter(models.ApiEnvironment.id == request.environment_id).first()
    if not environment:
        raise HTTPException(status_code=404, detail="环境不存在")
    
    # 创建执行记录
    execution = models.TestTaskExecution(
        task_id=task_id,
        environment_id=request.environment_id,
        status='running',
        total_count=len(task.items) if task.items else 0,
        success_count=0,
        failed_count=0,
        execution_results=[]
    )
    db.add(execution)
    
    # 更新任务状态
    task.status = 'running'
    db.commit()
    db.refresh(execution)
    
    # 在后台执行任务（异步）
    import threading
    def run_task():
        db_session = SessionLocal()
        try:
            results = []
            success_count = 0
            failed_count = 0
            
            # 重新加载任务和items（在后台线程中使用新的session）
            db_task = db_session.query(models.TestTask).options(
                joinedload(models.TestTask.items)
            ).filter(models.TestTask.id == task_id).first()
            
            if not db_task:
                raise Exception("任务不存在")
            
            # 按顺序执行每个任务项
            sorted_items = sorted(db_task.items, key=lambda x: x.sort_order) if db_task.items else []
            for item in sorted_items:
                try:
                    if item.item_type == 'api':
                        # 执行接口测试
                        endpoint = db_session.query(models.ApiEndpoint).filter(models.ApiEndpoint.id == item.item_id).first()
                        if not endpoint:
                            results.append({
                                "item_type": "api",
                                "item_id": item.item_id,
                                "item_name": f"接口 {item.item_id}",
                                "success": False,
                                "error_message": "接口不存在"
                            })
                            failed_count += 1
                            continue
                        
                        # 获取默认测试数据
                        test_data = db_session.query(models.ApiTestData).filter(
                            models.ApiTestData.endpoint_id == item.item_id
                        ).first()
                        
                        # 使用任务选择的环境，而不是接口自己的环境
                        # 构建执行请求
                        from schemas import ApiExecuteRequest
                        execute_request = ApiExecuteRequest(
                            environment_id=request.environment_id,
                            test_data_id=test_data.id if test_data else None,
                            global_variables={}
                        )
                        
                        # 调用现有的接口执行逻辑
                        api_result = {
                            "item_type": "api",
                            "item_id": item.item_id,
                            "item_name": endpoint.name or endpoint.path,
                            "success": False,
                            "status_code": None,
                            "error_message": None,
                            "execution_time": None,
                            "details": {}
                        }
                        
                        try:
                            # 获取环境
                            env = db_session.query(models.ApiEnvironment).filter(
                                models.ApiEnvironment.id == request.environment_id
                            ).first()
                            
                            if not env:
                                raise Exception(f"环境不存在 (ID: {request.environment_id})")
                            
                            # 构建URL
                            base_url = env.base_url.rstrip('/')
                            path = endpoint.path.lstrip('/')
                            url = f"{base_url}/{path}"
                            
                            # 处理路径参数
                            path_params = {}
                            if test_data and test_data.path_params:
                                path_params = test_data.path_params
                                for key, value in path_params.items():
                                    url = url.replace(f'{{{key}}}', str(value))
                            
                            # 准备请求参数
                            headers = {}
                            if env.headers:
                                headers.update(env.headers)
                            if test_data and test_data.headers:
                                headers.update(test_data.headers)
                            
                            # 应用 Header 替换
                            if request.header_replacements:
                                for replacement in request.header_replacements:
                                    if replacement.key and replacement.value:
                                        headers[replacement.key] = replacement.value
                            
                            query_params = {}
                            if test_data and test_data.query_params:
                                query_params = test_data.query_params
                            
                            body = None
                            if endpoint.method.upper() in ['POST', 'PUT', 'PATCH']:
                                if test_data and test_data.body:
                                    body = test_data.body
                            
                            # 获取断言信息
                            assertions_list = []
                            if test_data and test_data.assertions:
                                assertions_list = test_data.assertions
                            
                            # 应用断言替换（如果有替换配置，则使用替换的断言）
                            if request.assertion_replacements:
                                assertions_list = [
                                    {
                                        "type": a.type,
                                        "target": a.target,
                                        "operator": a.operator,
                                        "expected": a.expected
                                    }
                                    for a in request.assertion_replacements
                                    if a.type and a.operator and a.expected is not None
                                ]
                                print(f"🔍 测试任务断言替换: {assertions_list}")
                            
                            # 执行请求
                            import requests
                            request_kwargs = {
                                'method': endpoint.method.upper(),
                                'url': url,
                                'headers': headers,
                                'timeout': 30
                            }
                            
                            if query_params:
                                request_kwargs['params'] = query_params
                            
                            if body:
                                request_kwargs['json'] = body
                            
                            # 记录请求详情
                            api_result["details"] = {
                                "request_url": url,
                                "request_method": endpoint.method.upper(),
                                "request_headers": headers,
                                "request_query_params": query_params,
                                "request_path_params": path_params,
                                "request_body": body,
                                "request_assertions": assertions_list,
                                "environment": {
                                    "id": env.id,
                                    "name": env.name,
                                    "base_url": env.base_url
                                }
                            }
                            
                            start_time = time.time()
                            try:
                                response = requests.request(**request_kwargs)
                                response_time = int((time.time() - start_time) * 1000)
                                
                                # 尝试解析响应体
                                response_body = None
                                try:
                                    response_body = response.text
                                    if len(response_body) > 5000:
                                        response_body = response_body[:5000] + "... (truncated)"
                                except:
                                    response_body = str(response.content)[:5000] if hasattr(response, 'content') else None
                                
                                api_result["status_code"] = response.status_code
                                api_result["execution_time"] = response_time
                                
                                # 更新响应详情
                                api_result["details"].update({
                                    "response_status": response.status_code,
                                    "response_headers": dict(response.headers),
                                    "response_body": response_body,
                                    "response_time": response_time
                                })
                                
                                # 验证断言
                                assertion_success = True
                                assertion_errors = []
                                
                                print(f"🔍 开始验证断言，断言数量: {len(assertions_list) if assertions_list else 0}")
                                
                                if assertions_list:
                                    # 尝试解析响应体为 JSON
                                    json_body = None
                                    try:
                                        json_body = response.json()
                                    except:
                                        pass
                                    
                                    for assertion in assertions_list:
                                        assertion_type = assertion.get('type')
                                        operator = assertion.get('operator')
                                        target = assertion.get('target')
                                        expected = assertion.get('expected')
                                        
                                        if assertion_type == 'status_code':
                                            # 状态码断言
                                            actual_value = response.status_code
                                            if not _check_assertion(actual_value, operator, expected):
                                                assertion_success = False
                                                assertion_errors.append(f"状态码断言失败: 期望 {expected}，实际 {actual_value}")
                                        
                                        elif assertion_type == 'json_path':
                                            # JSON路径断言
                                            if json_body and target:
                                                actual_value = _extract_json_path(json_body, target)
                                                print(f"🔍 JSON路径断言: path={target}, expected={expected}, actual={actual_value}, operator={operator}")
                                                check_result = _check_assertion(actual_value, operator, expected)
                                                print(f"🔍 断言检查结果: {check_result}")
                                                if not check_result:
                                                    assertion_success = False
                                                    assertion_errors.append(f"JSON路径断言失败: {target} 期望 {expected}，实际 {_format_value_for_display(actual_value)}")
                                            else:
                                                assertion_success = False
                                                assertion_errors.append(f"JSON路径断言失败: 无法提取路径 {target}")
                                        
                                        elif assertion_type == 'response_time':
                                            # 响应时间断言
                                            actual_value = response_time
                                            if not _check_assertion(actual_value, operator, expected):
                                                assertion_success = False
                                                assertion_errors.append(f"响应时间断言失败: 期望 {expected}ms，实际 {actual_value}ms")
                                        
                                        elif assertion_type == 'contains':
                                            # 包含断言
                                            if json_body:
                                                body_str = json.dumps(json_body) if isinstance(json_body, dict) else str(json_body)
                                                if expected and expected not in body_str:
                                                    assertion_success = False
                                                    assertion_errors.append(f"包含断言失败: 响应体中不包含 {expected}")
                                            else:
                                                if expected and expected not in str(response_body):
                                                    assertion_success = False
                                                    assertion_errors.append(f"包含断言失败: 响应体中不包含 {expected}")
                                    
                                    # 记录断言结果
                                    api_result["details"]["assertion_results"] = assertion_errors if assertion_errors else ["所有断言通过"]
                                    api_result["success"] = assertion_success
                                    
                                    if not assertion_success:
                                        api_result["error_message"] = "; ".join(assertion_errors)
                                        failed_count += 1
                                    else:
                                        success_count += 1
                                else:
                                    # 如果没有断言，使用默认逻辑：HTTP 状态码判断
                                    api_result["success"] = response.status_code < 400
                                    if not api_result["success"]:
                                        api_result["error_message"] = f"HTTP {response.status_code}: {response_body[:200] if response_body else '无响应内容'}"
                                        failed_count += 1
                                    else:
                                        success_count += 1
                            
                            except requests.exceptions.Timeout:
                                response_time = int((time.time() - start_time) * 1000)
                                api_result["execution_time"] = response_time
                                api_result["error_message"] = f"请求超时 (超过30秒)"
                                api_result["details"]["error_type"] = "timeout"
                                failed_count += 1
                            
                            except requests.exceptions.ConnectionError as e:
                                response_time = int((time.time() - start_time) * 1000)
                                api_result["execution_time"] = response_time
                                api_result["error_message"] = f"连接错误: {str(e)}"
                                api_result["details"]["error_type"] = "connection_error"
                                failed_count += 1
                            
                            except requests.exceptions.RequestException as e:
                                response_time = int((time.time() - start_time) * 1000)
                                api_result["execution_time"] = response_time
                                api_result["error_message"] = f"请求异常: {str(e)}"
                                api_result["details"]["error_type"] = "request_exception"
                                failed_count += 1
                        
                        except Exception as exec_error:
                            import traceback
                            error_msg = str(exec_error)
                            traceback.print_exc()
                            api_result["error_message"] = f"执行失败: {error_msg}"
                            api_result["details"]["error_type"] = "execution_error"
                            api_result["details"]["error_traceback"] = traceback.format_exc()
                            failed_count += 1
                        
                        finally:
                            # 确保结果被记录
                            results.append(api_result)
                    
                    elif item.item_type == 'flow':
                        # 执行流程测试
                        flow = db_session.query(models.ApiTestFlow).filter(models.ApiTestFlow.id == item.item_id).first()
                        if not flow:
                            results.append({
                                "item_type": "flow",
                                "item_id": item.item_id,
                                "item_name": f"流程 {item.item_id}",
                                "success": False,
                                "error_message": "流程不存在"
                            })
                            failed_count += 1
                            continue
                        
                        # 使用任务选择的环境执行流程
                        # 构建执行请求
                        from schemas import FlowExecuteRequest
                        flow_execute_request = FlowExecuteRequest(
                            environment_id=request.environment_id,
                            global_variables={}
                        )
                        
                        # 流程执行结果
                        flow_result = {
                            "item_type": "flow",
                            "item_id": item.item_id,
                            "item_name": flow.name,
                            "success": False,
                            "execution_time": None,
                            "error_message": None,
                            "details": {
                                "steps": []
                            }
                        }
                        
                        try:
                            # 调用流程执行逻辑（简化版，直接执行流程的每个步骤）
                            # 获取环境
                            env = db_session.query(models.ApiEnvironment).filter(
                                models.ApiEnvironment.id == request.environment_id
                            ).first()
                            
                            if not env:
                                raise Exception(f"环境不存在 (ID: {request.environment_id})")
                            
                            # 初始化上下文（用于变量替换和模板渲染）
                            context: Dict[str, Any] = {}
                            # 从流程变量加载（如果有）
                            if flow.global_variables:
                                if isinstance(flow.global_variables, dict):
                                    context.update(flow.global_variables)
                                elif isinstance(flow.global_variables, str):
                                    try:
                                        import json
                                        parsed_vars = json.loads(flow.global_variables)
                                        if isinstance(parsed_vars, dict):
                                            context.update(parsed_vars)
                                    except:
                                        pass
                            
                            # 获取执行配置
                            fail_action = "stop"
                            if hasattr(flow, 'executionConfig') and flow.executionConfig:
                                if isinstance(flow.executionConfig, dict):
                                    fail_action = flow.executionConfig.get("failAction", "stop")
                            elif hasattr(flow, 'execution_config') and flow.execution_config:
                                if isinstance(flow.execution_config, dict):
                                    fail_action = flow.execution_config.get("failAction", "stop")
                            
                            # 执行流程的每个步骤（过滤掉被禁用的步骤）
                            enabled_steps = [step for step in (flow.steps or []) if step.get("enabled") is not False]
                            flow_success = True
                            flow_error = None
                            total_flow_time = 0
                            
                            for step_idx, step in enumerate(enabled_steps):
                                if not step.get("enabled", True):
                                    continue
                                
                                step_result = {
                                    "step_index": step_idx + 1,
                                    "step_name": step.get("alias") or f"步骤 {step_idx + 1}",
                                    "success": False,
                                    "error_message": None,
                                    "execution_time": None,
                                    "details": {}
                                }
                                
                                try:
                                    step_endpoint_id = step.get("endpoint_id")
                                    if not step_endpoint_id:
                                        step_result["error_message"] = "步骤中未指定接口ID"
                                        flow_result["details"]["steps"].append(step_result)
                                        flow_success = False
                                        continue
                                    
                                    step_endpoint = db_session.query(models.ApiEndpoint).filter(
                                        models.ApiEndpoint.id == step_endpoint_id
                                    ).first()
                                    
                                    if not step_endpoint:
                                        step_result["error_message"] = f"步骤中的接口 {step_endpoint_id} 不存在"
                                        flow_result["details"]["steps"].append(step_result)
                                        flow_success = False
                                        continue
                                    
                                    # 使用任务选择的环境，而不是步骤自己的环境
                                    step_env = env
                                    
                                    # 获取步骤的测试数据
                                    step_test_data_id = step.get("test_data_id")
                                    step_test_data = None
                                    if step_test_data_id:
                                        step_test_data = db_session.query(models.ApiTestData).filter(
                                            models.ApiTestData.id == step_test_data_id
                                        ).first()
                                    
                                    # 从之前的步骤中提取变量（如果有）
                                    # 注意：这里我们需要从 flow_result["details"]["steps"] 中提取之前步骤的响应
                                    for prev_step_result in flow_result["details"]["steps"]:
                                        prev_step_idx = prev_step_result.get("step_index", 0)
                                        if prev_step_idx > 0 and prev_step_idx < step_idx + 1:
                                            # 将之前步骤的响应体存储到 context 中
                                            prev_response_body = prev_step_result.get("details", {}).get("response_body")
                                            if prev_response_body:
                                                try:
                                                    import json
                                                    prev_json_body = json.loads(prev_response_body) if isinstance(prev_response_body, str) else prev_response_body
                                                    context[f"API[{prev_step_idx}]"] = prev_json_body or {}
                                                except:
                                                    pass
                                            
                                            # 从之前步骤的提取变量中获取
                                            prev_extracted = prev_step_result.get("details", {}).get("extracted", {})
                                            if prev_extracted:
                                                context.update(prev_extracted)
                                    
                                    # 组装请求数据（和 execute_api_flow 中的逻辑一致，支持模板渲染）
                                    headers = {}
                                    if step_env.headers:
                                        headers.update(step_env.headers)
                                    if step_test_data and step_test_data.headers:
                                        headers.update(step_test_data.headers)
                                    if step.get("headers") is not None:
                                        rendered_headers = _render_template(step["headers"], context)
                                        if isinstance(rendered_headers, dict):
                                            headers = rendered_headers
                                        else:
                                            headers.update(rendered_headers if isinstance(rendered_headers, dict) else {})
                                    
                                    # 路径参数：步骤自己的参数优先级最高，支持模板渲染
                                    path_params = step_test_data.path_params if step_test_data else None
                                    if step.get("path_params") is not None:
                                        path_params = _render_template(step["path_params"], context)
                                    
                                    # 查询参数：步骤自己的参数优先级最高，支持模板渲染
                                    query_params = step_test_data.query_params if step_test_data else None
                                    if step.get("query_params") is not None:
                                        query_params = _render_template(step["query_params"], context)
                                    
                                    # 请求体：步骤自己的参数优先级最高，支持模板渲染
                                    body = step_test_data.body if step_test_data else None
                                    if step.get("body") is not None:
                                        body = _render_template(step["body"], context)
                                    
                                    # 拼接 URL 并替换路径参数
                                    base_url = step_env.base_url.rstrip('/')
                                    path = step_endpoint.path.lstrip('/')
                                    step_url = f"{base_url}/{path}"
                                    if path_params:
                                        for key, value in path_params.items():
                                            step_url = step_url.replace(f"{{{key}}}", str(value))
                                    
                                    # 记录步骤请求详情
                                    step_result["details"] = {
                                        "request_url": step_url,
                                        "request_method": step_endpoint.method.upper(),
                                        "request_headers": headers,
                                        "request_path_params": path_params,
                                        "request_query_params": query_params,
                                        "request_body": body,
                                        "endpoint_id": step_endpoint_id,
                                        "endpoint_name": step_endpoint.name or step_endpoint.path
                                    }
                                    
                                    # 执行请求
                                    import requests
                                    import json
                                    step_request_kwargs = {
                                        'method': step_endpoint.method.upper(),
                                        'url': step_url,
                                        'headers': headers,
                                        'timeout': 30
                                    }
                                    
                                    if query_params:
                                        step_request_kwargs['params'] = query_params
                                    
                                    if body is not None:
                                        step_request_kwargs['json'] = body
                                    
                                    step_start_time = time.time()
                                    try:
                                        step_response = requests.request(**step_request_kwargs)
                                        step_response_time = int((time.time() - step_start_time) * 1000)
                                        total_flow_time += step_response_time
                                        
                                        # 尝试解析响应体
                                        step_response_body = None
                                        try:
                                            step_response_body = step_response.text
                                            if len(step_response_body) > 5000:
                                                step_response_body = step_response_body[:5000] + "... (truncated)"
                                        except:
                                            step_response_body = str(step_response.content)[:5000] if hasattr(step_response, 'content') else None
                                        
                                        step_result["execution_time"] = step_response_time
                                        step_result["status_code"] = step_response.status_code
                                        
                                        # 获取断言信息
                                        step_assertions_list = step.get("assertions") or []
                                        
                                        # 解析响应体为JSON（用于断言验证）
                                        step_json_body = None
                                        try:
                                            step_json_body = step_response.json()
                                        except:
                                            step_json_body = None
                                        
                                        step_result["details"].update({
                                            "response_status": step_response.status_code,
                                            "response_headers": dict(step_response.headers),
                                            "response_body": step_response_body,
                                            "response_time": step_response_time,
                                            "request_assertions": step_assertions_list
                                        })
                                        
                                        # 验证断言
                                        if step_assertions_list and len(step_assertions_list) > 0:
                                            step_assertion_success = True
                                            step_assertion_errors = []
                                            
                                            for assertion in step_assertions_list:
                                                assertion_type = assertion.get('type')
                                                operator = assertion.get('operator')
                                                target = assertion.get('target')
                                                expected = assertion.get('expected')
                                                
                                                # 跳过无效的断言
                                                if not assertion_type or not operator:
                                                    continue
                                                
                                                if assertion_type == 'status_code':
                                                    # 状态码断言
                                                    actual_value = step_response.status_code
                                                    if expected is None or expected == '':
                                                        step_assertion_success = False
                                                        step_assertion_errors.append(f"状态码断言失败: 期望值不能为空")
                                                        continue
                                                    expected_str = str(expected).strip()
                                                    if not _check_assertion(actual_value, operator, expected_str):
                                                        step_assertion_success = False
                                                        step_assertion_errors.append(f"状态码断言失败: 期望 {expected_str}，实际 {actual_value}")
                                                
                                                elif assertion_type == 'json_path':
                                                    # JSON路径断言
                                                    if step_json_body and target:
                                                        # 渲染 expected 值（处理变量替换）
                                                        if expected:
                                                            try:
                                                                expected_rendered = _render_template(expected, context)
                                                                if isinstance(expected_rendered, str) and expected_rendered.startswith('"') and expected_rendered.endswith('"'):
                                                                    expected_rendered = expected_rendered[1:-1]
                                                                expected = expected_rendered
                                                            except Exception as e:
                                                                print(f"🔍 渲染 expected 值失败: {e}")
                                                        
                                                        actual_value = _extract_json_path(step_json_body, target)
                                                        print(f"🔍 流程JSON路径断言: path={target}, expected={expected}, actual={actual_value}, operator={operator}")
                                                        
                                                        if not _check_assertion(actual_value, operator, expected):
                                                            step_assertion_success = False
                                                            step_assertion_errors.append(f"JSON路径断言失败: {target} 期望 {expected}，实际 {_format_value_for_display(actual_value)}")
                                                    else:
                                                        step_assertion_success = False
                                                        step_assertion_errors.append(f"JSON路径断言失败: 无法提取路径 {target}")
                                                
                                                elif assertion_type == 'response_time':
                                                    # 响应时间断言
                                                    actual_value = step_response_time
                                                    if not _check_assertion(actual_value, operator, expected):
                                                        step_assertion_success = False
                                                        step_assertion_errors.append(f"响应时间断言失败: 期望 {expected}ms，实际 {actual_value}ms")
                                                
                                                elif assertion_type == 'contains':
                                                    # 包含断言
                                                    if step_json_body:
                                                        body_str = json.dumps(step_json_body) if isinstance(step_json_body, dict) else str(step_json_body)
                                                        if expected and expected not in body_str:
                                                            step_assertion_success = False
                                                            step_assertion_errors.append(f"包含断言失败: 响应体中不包含 {expected}")
                                                    else:
                                                        if expected and expected not in str(step_response_body):
                                                            step_assertion_success = False
                                                            step_assertion_errors.append(f"包含断言失败: 响应体中不包含 {expected}")
                                            
                                            # 设置步骤成功状态
                                            step_result["success"] = step_assertion_success
                                            step_result["details"]["assertion_results"] = step_assertion_errors if step_assertion_errors else ["所有断言通过"]
                                            
                                            if not step_assertion_success:
                                                step_result["error_message"] = "; ".join(step_assertion_errors)
                                                flow_success = False
                                                flow_error = f"步骤 {step_idx + 1} 断言失败: {'; '.join(step_assertion_errors)}"
                                                if fail_action == "stop":
                                                    flow_result["details"]["steps"].append(step_result)
                                                    break
                                        else:
                                            # 没有断言，使用HTTP状态码判断
                                            step_result["success"] = step_response.status_code < 400
                                            if step_response.status_code >= 400:
                                                step_result["error_message"] = f"HTTP {step_response.status_code}: {step_response_body[:200] if step_response_body else '无响应内容'}"
                                                flow_success = False
                                                flow_error = f"步骤 {step_idx + 1} 失败: HTTP {step_response.status_code}"
                                                if fail_action == "stop":
                                                    flow_result["details"]["steps"].append(step_result)
                                                    break
                                    
                                    except requests.exceptions.Timeout:
                                        step_response_time = int((time.time() - step_start_time) * 1000)
                                        total_flow_time += step_response_time
                                        step_result["execution_time"] = step_response_time
                                        step_result["error_message"] = "请求超时 (超过30秒)"
                                        step_result["details"]["error_type"] = "timeout"
                                        step_result["success"] = False
                                        flow_success = False
                                        flow_error = f"步骤 {step_idx + 1} 超时"
                                        if fail_action == "stop":
                                            break
                                    
                                    except requests.exceptions.ConnectionError as e:
                                        step_response_time = int((time.time() - step_start_time) * 1000)
                                        total_flow_time += step_response_time
                                        step_result["execution_time"] = step_response_time
                                        step_result["error_message"] = f"连接错误: {str(e)}"
                                        step_result["details"]["error_type"] = "connection_error"
                                        step_result["success"] = False
                                        flow_success = False
                                        flow_error = f"步骤 {step_idx + 1} 连接失败: {str(e)}"
                                        if fail_action == "stop":
                                            break
                                    
                                    except requests.exceptions.RequestException as e:
                                        step_response_time = int((time.time() - step_start_time) * 1000)
                                        total_flow_time += step_response_time
                                        step_result["execution_time"] = step_response_time
                                        step_result["error_message"] = f"请求异常: {str(e)}"
                                        step_result["details"]["error_type"] = "request_exception"
                                        step_result["success"] = False
                                        flow_success = False
                                        flow_error = f"步骤 {step_idx + 1} 请求异常: {str(e)}"
                                        if fail_action == "stop":
                                            break
                                    
                                    except Exception as exc:
                                        step_response_time = int((time.time() - step_start_time) * 1000) if 'step_start_time' in locals() else 0
                                        total_flow_time += step_response_time
                                        step_result["execution_time"] = step_response_time
                                        step_result["error_message"] = f"执行异常: {str(exc)}"
                                        step_result["details"]["error_type"] = "execution_error"
                                        step_result["success"] = False
                                        flow_success = False
                                        flow_error = f"步骤 {step_idx + 1} 执行异常: {str(exc)}"
                                        if fail_action == "stop":
                                            break
                                
                                except Exception as step_error:
                                    import traceback
                                    step_result["error_message"] = f"步骤执行异常: {str(step_error)}"
                                    step_result["details"]["error_type"] = "step_execution_error"
                                    step_result["details"]["error_traceback"] = traceback.format_exc()
                                    flow_success = False
                                    flow_error = f"步骤 {step_idx + 1} 执行异常: {str(step_error)}"
                                
                                finally:
                                    # 将步骤结果添加到流程结果中
                                    flow_result["details"]["steps"].append(step_result)
                            
                            # 将流程的每个步骤转换为独立的执行结果项（用于前端展示）
                            # 这样前端可以将流程拆分成多个单接口来展示
                            for step_result_item in flow_result["details"]["steps"]:
                                step_execution_result = {
                                    "item_type": "api",  # 标记为api类型，以便前端统一展示
                                    "item_id": step_result_item.get("details", {}).get("endpoint_id", 0),
                                    "item_name": step_result_item.get("step_name", ""),
                                    "success": step_result_item.get("success", False),
                                    "status_code": step_result_item.get("status_code"),
                                    "error_message": step_result_item.get("error_message"),
                                    "execution_time": step_result_item.get("execution_time"),
                                    "details": {
                                        "request_url": step_result_item.get("details", {}).get("request_url", ""),
                                        "request_method": step_result_item.get("details", {}).get("request_method", ""),
                                        "request_headers": step_result_item.get("details", {}).get("request_headers", {}),
                                        "request_path_params": step_result_item.get("details", {}).get("request_path_params", {}),
                                        "request_query_params": step_result_item.get("details", {}).get("request_query_params", {}),
                                        "request_body": step_result_item.get("details", {}).get("request_body"),
                                        "request_assertions": step_result_item.get("details", {}).get("request_assertions", []),
                                        "response_status": step_result_item.get("details", {}).get("response_status"),
                                        "response_headers": step_result_item.get("details", {}).get("response_headers", {}),
                                        "response_body": step_result_item.get("details", {}).get("response_body"),
                                        "response_time": step_result_item.get("details", {}).get("response_time", 0)
                                    }
                                }
                                results.append(step_execution_result)
                            
                            # 流程整体结果（不再添加到results，因为已经将每个步骤作为独立项添加了）
                            flow_result["success"] = flow_success
                            flow_result["execution_time"] = total_flow_time
                            flow_result["error_message"] = flow_error
                            flow_result["details"]["environment"] = {
                                "id": env.id,
                                "name": env.name,
                                "base_url": env.base_url
                            }
                            
                            # 注意：流程的整体结果不再添加到 results 中，因为我们已经将每个步骤作为独立的项添加了
                            # 这样前端就可以将流程拆分成多个单接口来展示
                            # 统计成功和失败数量
                            for step_result_item in flow_result["details"]["steps"]:
                                if step_result_item.get("success", False):
                                    success_count += 1
                                else:
                                    failed_count += 1
                        
                        except Exception as flow_error:
                            import traceback
                            error_msg = str(flow_error)
                            traceback.print_exc()
                            flow_result["error_message"] = f"执行失败: {error_msg}"
                            flow_result["details"]["error_type"] = "flow_execution_error"
                            flow_result["details"]["error_traceback"] = traceback.format_exc()
                            failed_count += 1
                        
                        finally:
                            # 注意：流程的整体结果不再添加到 results 中，因为我们已经将每个步骤作为独立的项添加了
                            # results.append(flow_result)  # 已移除
                            pass
                
                except Exception as e:
                    import traceback
                    error_msg = str(e)
                    traceback.print_exc()
                    results.append({
                        "item_type": item.item_type,
                        "item_id": item.item_id,
                        "item_name": f"{item.item_type} {item.item_id}",
                        "success": False,
                        "error_message": error_msg
                    })
                    failed_count += 1
            
            # 更新执行记录（确保所有结果都被记录）
            try:
                db_execution = db_session.query(models.TestTaskExecution).filter(
                    models.TestTaskExecution.id == execution.id
                ).first()
                if db_execution:
                    db_execution.status = 'success' if failed_count == 0 else 'failed'
                    db_execution.success_count = success_count
                    db_execution.failed_count = failed_count
                    db_execution.total_count = len(results)  # 确保总数正确
                    # 确保 execution_results 是 JSON 格式（列表）
                    if results:
                        db_execution.execution_results = results  # 保存所有执行结果
                        print(f"✅ 保存执行记录: task_id={task_id}, execution_id={execution.id}, results_count={len(results)}, success={success_count}, failed={failed_count}")
                        # 打印第一个结果的示例（用于调试）
                        if results:
                            import json
                            print(f"   执行结果示例: {json.dumps(results[0], ensure_ascii=False, indent=2)[:500]}")
                    else:
                        db_execution.execution_results = []  # 如果没有结果，保存空数组
                        print(f"⚠️ 执行记录为空: task_id={task_id}, execution_id={execution.id}")
                    db_execution.completed_at = datetime.now()
                    if failed_count > 0 and not db_execution.error_message:
                        # 如果有失败，记录汇总错误信息
                        failed_items = [r for r in results if not r.get("success", False)]
                        error_summary = f"共 {failed_count} 个接口/流程执行失败"
                        if failed_items:
                            error_summary += f"，失败项: {', '.join([r.get('item_name', '未知') for r in failed_items[:5]])}"
                            if len(failed_items) > 5:
                                error_summary += f" 等"
                        db_execution.error_message = error_summary
                    db_session.commit()
                
                # 更新任务状态
                db_task = db_session.query(models.TestTask).filter(models.TestTask.id == task_id).first()
                if db_task:
                    db_task.status = 'success' if failed_count == 0 else 'failed'
                    db_session.commit()
            except Exception as commit_error:
                import traceback
                print(f"❌ 更新执行记录失败: {commit_error}")
                traceback.print_exc()
                # 尝试回滚并重新提交
                try:
                    db_session.rollback()
                    db_session.commit()
                except:
                    pass
        
        except Exception as e:
            import traceback
            error_msg = str(e)
            traceback.print_exc()
            # 即使出现异常，也要尝试保存已收集的结果
            try:
                db_execution = db_session.query(models.TestTaskExecution).filter(
                    models.TestTaskExecution.id == execution.id
                ).first()
                if db_execution:
                    # 保存已收集的结果（即使不完整）
                    if results:
                        db_execution.execution_results = results
                        db_execution.success_count = success_count
                        db_execution.failed_count = failed_count
                        db_execution.total_count = len(results)
                    else:
                        db_execution.execution_results = []  # 如果没有结果，保存空数组
                        db_execution.total_count = 0
                    db_execution.status = 'failed'
                    db_execution.error_message = f"任务执行异常: {error_msg}"
                    db_execution.completed_at = datetime.now()
                    print(f"❌ 保存异常执行记录: task_id={task_id}, execution_id={execution.id}, results_count={len(results) if results else 0}")
                    db_session.commit()
                
                db_task = db_session.query(models.TestTask).filter(models.TestTask.id == task_id).first()
                if db_task:
                    db_task.status = 'failed'
                    db_session.commit()
            except Exception as save_error:
                import traceback
                print(f"❌ 保存异常执行记录失败: {save_error}")
                traceback.print_exc()
        finally:
            try:
                db_session.close()
            except:
                pass
    
    thread = threading.Thread(target=run_task)
    thread.daemon = True
    thread.start()
    
    return execution


@app.get("/api/test-tasks/{task_id}/executions", response_model=List[schemas.TestTaskExecutionSummary])
def get_test_task_executions(
    task_id: int,
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取测试任务执行记录列表（不包含详细结果，仅返回摘要信息）"""
    require_permission(current_user.role, "apitest", "read")
    
    # 只查询需要的列，排除 execution_results 大字段以提高性能
    executions = db.query(
        models.TestTaskExecution.id,
        models.TestTaskExecution.task_id,
        models.TestTaskExecution.environment_id,
        models.TestTaskExecution.status,
        models.TestTaskExecution.total_count,
        models.TestTaskExecution.success_count,
        models.TestTaskExecution.failed_count,
        models.TestTaskExecution.error_message,
        models.TestTaskExecution.started_at,
        models.TestTaskExecution.completed_at
    ).filter(
        models.TestTaskExecution.task_id == task_id
    ).order_by(models.TestTaskExecution.started_at.desc()).limit(limit).all()
    
    # 将查询结果转换为字典列表
    result = []
    for execution in executions:
        result.append({
            "id": execution.id,
            "task_id": execution.task_id,
            "environment_id": execution.environment_id,
            "status": execution.status,
            "total_count": execution.total_count,
            "success_count": execution.success_count,
            "failed_count": execution.failed_count,
            "error_message": execution.error_message,
            "started_at": execution.started_at,
            "completed_at": execution.completed_at
        })
    
    return result


@app.get("/api/test-tasks/{task_id}/executions/{execution_id}", response_model=schemas.TestTaskExecution)
def get_test_task_execution(
    task_id: int,
    execution_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取测试任务执行记录详情"""
    require_permission(current_user.role, "apitest", "read")
    
    execution = db.query(models.TestTaskExecution).filter(
        models.TestTaskExecution.id == execution_id,
        models.TestTaskExecution.task_id == task_id
    ).first()
    
    if not execution:
        raise HTTPException(status_code=404, detail="执行记录不存在")
    
    return execution


# ==================== 测试文件管理 ====================

# 文件上传目录配置
# 优先使用环境变量 UPLOAD_DIR，否则使用默认路径
# Docker 部署时通过环境变量指定持久化卷路径（如 /opt/bug-uploads）
#
# 目录结构（按文件类型和名称组织）：
#   uploads/ 或 /opt/bug-uploads/
#     ├── local/                    # 本地上传的文件
#     │   ├── 测试数据1/            # 文件管理中设置的"名称"
#     │   │   └── a1b2c3d4.json     # 实际文件（UUID命名）
#     │   ├── 用户信息/
#     │   │   └── x9y8z7w6.csv
#     │   └── ...
#     └── flow/                     # 流程导出的文件
#         ├── 登录流程/
#         │   └── f1e2d3c4.json
#         └── ...
#
UPLOAD_BASE_DIR = os.environ.get("UPLOAD_DIR", os.path.join(os.path.dirname(__file__), "uploads"))
UPLOAD_DIR_LOCAL = os.path.join(UPLOAD_BASE_DIR, "local")
UPLOAD_DIR_FLOW = os.path.join(UPLOAD_BASE_DIR, "flow")
os.makedirs(UPLOAD_DIR_LOCAL, exist_ok=True)
os.makedirs(UPLOAD_DIR_FLOW, exist_ok=True)

# Bug 图片存储目录配置
# Docker 部署时映射到 /opt/bug-images，本地开发时使用 backend/images
# 目录结构：
#   images/ 或 /opt/bug-images/
#     ├── BUG-001/                  # 以缺陷编号为文件夹名称
#     │   ├── a1b2c3d4.png
#     │   └── x9y8z7w6.jpg
#     ├── BUG-002/
#     │   └── ...
#     └── ...
#
BUG_IMAGE_DIR = os.environ.get("BUG_IMAGE_DIR", os.path.join(os.path.dirname(__file__), "images"))
os.makedirs(BUG_IMAGE_DIR, exist_ok=True)

def get_upload_dir(file_type: str) -> str:
    """根据文件类型获取对应的基础上传目录"""
    if file_type == "flow":
        return UPLOAD_DIR_FLOW
    return UPLOAD_DIR_LOCAL

@app.get("/api/test-files", response_model=schemas.TestFileList)
def get_test_files(
    keyword: Optional[str] = None,
    file_type: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取测试文件列表"""
    require_permission(current_user.role, "apitest", "read")
    
    query = db.query(models.TestFile)
    
    if keyword:
        query = query.filter(
            (models.TestFile.name.contains(keyword)) |
            (models.TestFile.description.contains(keyword)) |
            (models.TestFile.file_name.contains(keyword))
        )
    
    if file_type:
        query = query.filter(models.TestFile.file_type == file_type)
    
    total = query.count()
    items = query.order_by(models.TestFile.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return {"items": items, "total": total}

@app.get("/api/test-files/{file_id}", response_model=schemas.TestFile)
def get_test_file(
    file_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取单个测试文件详情"""
    require_permission(current_user.role, "apitest", "read")
    
    test_file = db.query(models.TestFile).filter(models.TestFile.id == file_id).first()
    if not test_file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    return test_file

@app.post("/api/test-files", response_model=schemas.TestFile)
async def create_test_file(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    file_type: str = Form("local"),
    file: Optional[UploadFile] = File(None),
    file_content: Optional[str] = Form(None),
    flow_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """创建测试文件（上传本地文件或保存流程导出）
    
    文件存储目录结构：
    - local/{名称}/: 本地上传的文件（.json, .txt, .csv, .xml 等）
    - flow/{名称}/: 流程导出的文件（.json）
    
    例如：
    - local/测试数据1/a1b2c3d4.json
    - flow/登录流程/f1e2d3c4.json
    """
    require_permission(current_user.role, "apitest", "write")
    
    import uuid
    import re
    
    file_path = None
    file_size = None
    mime_type = None
    actual_file_name = None
    content_json = None
    
    # 获取对应类型的上传目录
    base_upload_dir = get_upload_dir(file_type)
    
    # 清理文件夹名称（移除不安全字符）
    safe_folder_name = re.sub(r'[<>:"/\\|?*]', '_', name.strip())
    if not safe_folder_name:
        safe_folder_name = "unnamed"
    
    # 创建以"名称"命名的子文件夹
    upload_dir = os.path.join(base_upload_dir, safe_folder_name)
    os.makedirs(upload_dir, exist_ok=True)
    
    if file_type == "local" and file:
        # 本地上传文件
        actual_file_name = file.filename
        mime_type = file.content_type
        
        # 生成唯一文件名，保留原始扩展名
        ext = os.path.splitext(file.filename)[1] if file.filename else ""
        unique_name = f"{uuid.uuid4().hex}{ext}"
        file_path = os.path.join(upload_dir, unique_name)
        
        # 保存文件到文件系统
        content = await file.read()
        file_size = len(content)
        with open(file_path, "wb") as f:
            f.write(content)
            
    elif file_type == "flow" and file_content:
        # 流程导出内容 - 同时保存到文件系统和数据库
        actual_file_name = f"{name}.json"
        mime_type = "application/json"
        
        try:
            content_json = json.loads(file_content)
            file_size = len(file_content.encode('utf-8'))
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="无效的JSON内容")
        
        # 生成唯一文件名并保存到文件系统
        unique_name = f"{uuid.uuid4().hex}.json"
        file_path = os.path.join(upload_dir, unique_name)
        
        # 格式化 JSON 并保存
        formatted_content = json.dumps(content_json, ensure_ascii=False, indent=2)
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(formatted_content)
    else:
        raise HTTPException(status_code=400, detail="请提供文件或流程内容")
    
    test_file = models.TestFile(
        name=name,
        description=description,
        file_type=file_type,
        file_name=actual_file_name,
        file_path=file_path,
        file_content=content_json,  # 流程导出同时保存到数据库（便于快速查询）
        file_size=file_size,
        mime_type=mime_type,
        flow_id=flow_id,
        created_by=current_user.id
    )
    
    db.add(test_file)
    db.commit()
    db.refresh(test_file)
    
    return test_file

@app.put("/api/test-files/{file_id}", response_model=schemas.TestFile)
def update_test_file(
    file_id: int,
    request: schemas.TestFileUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """更新测试文件信息"""
    require_permission(current_user.role, "apitest", "write")
    
    test_file = db.query(models.TestFile).filter(models.TestFile.id == file_id).first()
    if not test_file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    if request.name is not None:
        test_file.name = request.name
    if request.description is not None:
        test_file.description = request.description
    
    db.commit()
    db.refresh(test_file)
    
    return test_file

@app.delete("/api/test-files/{file_id}")
def delete_test_file(
    file_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """删除测试文件"""
    require_permission(current_user.role, "apitest", "write")
    
    test_file = db.query(models.TestFile).filter(models.TestFile.id == file_id).first()
    if not test_file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    # 删除物理文件和空文件夹
    if test_file.file_path and os.path.exists(test_file.file_path):
        try:
            # 获取文件所在目录
            file_dir = os.path.dirname(test_file.file_path)
            
            # 删除文件
            os.remove(test_file.file_path)
            
            # 如果文件夹为空，删除文件夹
            if os.path.isdir(file_dir) and not os.listdir(file_dir):
                os.rmdir(file_dir)
        except Exception as e:
            print(f"删除文件失败: {e}")
    
    db.delete(test_file)
    db.commit()
    
    return {"message": "删除成功"}

@app.get("/api/test-files/{file_id}/download")
def download_test_file(
    file_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """下载测试文件
    
    优先从文件系统读取，如果文件不存在则尝试从数据库读取（兼容旧数据）
    """
    require_permission(current_user.role, "apitest", "read")
    
    test_file = db.query(models.TestFile).filter(models.TestFile.id == file_id).first()
    if not test_file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    from fastapi.responses import FileResponse, Response
    
    # 优先从文件系统读取
    if test_file.file_path and os.path.exists(test_file.file_path):
        return FileResponse(
            path=test_file.file_path,
            filename=test_file.file_name,
            media_type=test_file.mime_type or "application/octet-stream"
        )
    
    # 兼容旧数据：如果文件系统中没有，尝试从数据库读取（仅流程导出类型）
    if test_file.file_type == "flow" and test_file.file_content:
        content = json.dumps(test_file.file_content, ensure_ascii=False, indent=2)
        return Response(
            content=content,
            media_type="application/json",
            headers={
                "Content-Disposition": f'attachment; filename="{test_file.file_name}"'
            }
        )
    
    raise HTTPException(status_code=404, detail="文件内容不存在")

@app.get("/api/test-files/{file_id}/content")
def get_test_file_content(
    file_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
):
    """获取测试文件内容（用于导入）
    
    优先从文件系统读取，如果文件不存在则尝试从数据库读取（兼容旧数据）
    """
    require_permission(current_user.role, "apitest", "read")
    
    test_file = db.query(models.TestFile).filter(models.TestFile.id == file_id).first()
    if not test_file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    # 优先从文件系统读取
    if test_file.file_path and os.path.exists(test_file.file_path):
        try:
            with open(test_file.file_path, "r", encoding="utf-8") as f:
                content = f.read()
                return json.loads(content)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="文件内容不是有效的JSON格式")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"读取文件失败: {str(e)}")
    
    # 兼容旧数据：如果文件系统中没有，尝试从数据库读取（仅流程导出类型）
    if test_file.file_type == "flow" and test_file.file_content:
        return test_file.file_content
    
    raise HTTPException(status_code=404, detail="文件内容不存在")


# ==================== 健康检查 ====================

@app.get("/")
def root():
    return {"message": "缺陷管理系统API服务运行中", "version": "1.0.0"}

@app.get("/health")
def health_check():
    return {"status": "healthy"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=43211)

